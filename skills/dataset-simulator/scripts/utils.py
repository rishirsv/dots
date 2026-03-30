"""Shared utilities for data-room-simulator scripts."""

import json
import random
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Dict, List, Optional

import numpy as np
import pandas as pd

# --- Section directory constants ---

SECTIONS = {
    "corporate": "1.0-corporate",
    "financial": "2.0-financial",
    "tax": "3.0-tax",
    "commercial": "4.0-commercial",
    "operations": "5.0-operations",
    "hr": "6.0-hr",
    "technology": "7.0-technology",
    "legal": "8.0-legal",
    "regulatory": "9.0-regulatory",
    "real_estate": "10.0-real-estate",
    "insurance": "11.0-insurance",
    "process": "12.0-process",
}

SECTION_DIRS = list(SECTIONS.values())

SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent
PROFILES_DIR = PROJECT_DIR / "references" / "profiles"
TEMPLATES_DIR = PROJECT_DIR / "references" / "templates"


def load_deal_state(output_dir: Path) -> dict:
    """Load deal_state.json from the output directory.

    Args:
        output_dir: Path to the output directory containing deal_state.json

    Returns:
        Dictionary with deal state configuration

    Raises:
        FileNotFoundError: If deal_state.json not found
    """
    deal_state_path = output_dir / "deal_state.json"
    if not deal_state_path.exists():
        raise FileNotFoundError(f"deal_state.json not found in {output_dir}")
    with open(deal_state_path) as f:
        return json.load(f)


def load_profile_for_industry(industry: str) -> dict:
    """Load a profile JSON by industry name from the standard profiles directory.

    Args:
        industry: Industry name (e.g., 'dental', 'saas', 'construction')

    Returns:
        Dictionary with industry profile

    Raises:
        FileNotFoundError: If profile not found
    """
    profile_path = PROFILES_DIR / f"{industry}.json"
    if not profile_path.exists():
        raise FileNotFoundError(f"Profile not found: {profile_path}")
    with open(profile_path) as f:
        return json.load(f)


def round_currency(value: float) -> float:
    """Round to 2 decimal places using banker's rounding.

    Args:
        value: Numeric value to round

    Returns:
        Rounded value to 2 decimal places
    """
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def split_amount_exact(total: float, parts: int) -> List[float]:
    """Split a total into N parts that sum exactly to the total.

    Uses Dirichlet distribution for realistic variation,
    then adjusts the last element to eliminate rounding error.

    Args:
        total: Total amount to split
        parts: Number of parts to split into

    Returns:
        List of rounded amounts that sum exactly to total
    """
    if parts <= 0:
        return []
    if parts == 1:
        return [round_currency(total)]
    weights = np.random.dirichlet(np.ones(parts))
    amounts = [round_currency(total * w) for w in weights]
    amounts[-1] = round_currency(total - sum(amounts[:-1]))
    return amounts


def generate_ein() -> str:
    """Generate a realistic-looking EIN (XX-XXXXXXX).

    Returns:
        EIN string in format XX-XXXXXXX
    """
    return f"{random.randint(10, 99)}-{random.randint(1000000, 9999999)}"


def identify_compensation_accounts(profile: dict) -> List[str]:
    """Identify compensation-related account codes from a profile's chart of accounts.

    Searches COGS and OpEx accounts for salary/payroll/compensation keywords.

    Args:
        profile: Industry profile dictionary

    Returns:
        List of account code strings for compensation-related accounts
    """
    keywords = ["salary", "payroll", "compensation", "staff", "labor", "wage", "hygiene", "provider"]
    comp_accounts = []
    coa = profile.get("chart_of_accounts", profile.get("financial_model", {}).get("chart_of_accounts", {}))
    for category in ["cogs", "opex"]:
        for acct in coa.get(category, []):
            name = acct.get("name", "").lower()
            if any(kw in name for kw in keywords):
                comp_accounts.append(acct["code"])
    return comp_accounts


def save_to_excel(
    df: pd.DataFrame,
    filename: str,
    output_dir: Path,
    sheet_name: str = "Sheet1"
) -> Path:
    """Save a DataFrame to Excel in the given output directory.

    Creates the directory if it doesn't exist.

    Args:
        df: DataFrame to save
        filename: Name of the Excel file to create
        output_dir: Path to output directory
        sheet_name: Name of the worksheet (default: "Sheet1")

    Returns:
        Path to the created Excel file
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    filepath = output_dir / filename
    df.to_excel(filepath, index=False, sheet_name=sheet_name)
    return filepath


def save_to_excel_styled(
    df: pd.DataFrame,
    filepath: Path,
    sheet_name: str = "Sheet1"
) -> None:
    """Save a DataFrame to Excel with professional formatting (headers, borders, auto-width).

    Creates parent directories if they don't exist.
    Uses openpyxl for styling:
    - Blue header with white text
    - Thin borders around all cells
    - Auto-adjusted column widths
    - Text wrapping for headers

    Args:
        df: DataFrame to save
        filepath: Full path to the Excel file to create
        sheet_name: Name of the worksheet (default: "Sheet1")
    """
    filepath.parent.mkdir(parents=True, exist_ok=True)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl required for save_to_excel_styled. Install with: pip install openpyxl")

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Apply header styling
        for col_num, col_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Apply borders and alignment to data
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Auto-adjust column widths
        for col_num, col_title in enumerate(df.columns, 1):
            max_length = max(
                df[col_title].astype(str).apply(len).max(),
                len(str(col_title))
            )
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)


def get_tb_path(output_dir: Path) -> Path:
    """Get the path to the trial balance file.

    Args:
        output_dir: Path to the output directory

    Returns:
        Path to trial_balance.xlsx
    """
    return output_dir / SECTIONS["financial"] / "trial_balance.xlsx"


def load_trial_balance(output_dir: Path) -> pd.DataFrame:
    """Load trial balance from financial outputs.

    Args:
        output_dir: Path to the output directory

    Returns:
        DataFrame with trial balance data

    Raises:
        FileNotFoundError: If trial balance not found
    """
    tb_path = get_tb_path(output_dir)
    if not tb_path.exists():
        raise FileNotFoundError(f"Trial balance not found: {tb_path}")
    return pd.read_excel(tb_path)
