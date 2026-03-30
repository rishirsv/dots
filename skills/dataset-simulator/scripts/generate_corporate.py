#!/usr/bin/env python3
"""
Corporate Section Generator for M&A Data Room Simulator.

Generates Section 1.0 Corporate artifacts (entity chart, cap table, debt schedule,
bank account list, related party register) by reading deal_state.json.

Usage:
    python3 scripts/generate_corporate.py --output-dir output/
"""

import argparse
import json
import random
from datetime import datetime, timedelta
from pathlib import Path

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: pandas and openpyxl required. Install with: pip install pandas openpyxl")
    exit(1)

try:
    from faker import Faker
except ImportError:
    print("Error: faker package required. Install with: pip install faker")
    exit(1)

# Initialize Faker
fake = Faker()

# Constants
SCRIPT_DIR = Path(__file__).parent
PROJECT_DIR = SCRIPT_DIR.parent

BANK_NAMES = [
    "JPMorgan Chase", "Bank of America", "Wells Fargo", "Citibank",
    "Goldman Sachs", "Morgan Stanley", "Silicon Valley Bank", "Comerica",
    "PNC Bank", "U.S. Bank", "Truist Bank", "Fifth Third Bank"
]

MAJOR_BANKS = ["JPMorgan Chase", "Bank of America", "Wells Fargo", "Citibank"]

LENDER_TYPES = ["Bank", "Credit Union", "Finance Company", "Alternative Lender"]

LOAN_TYPES = [
    "revolver", "term_loan", "line_of_credit", "equipment", "bonding"
]

COLLATERAL_TYPES = [
    "accounts_receivable", "inventory", "equipment", "real_estate",
    "intellectual_property", "personal_guarantee", "cash_deposit"
]


def load_deal_state(output_dir: Path) -> dict:
    """Load deal_state.json from output directory."""
    deal_state_path = output_dir / "deal_state.json"
    if not deal_state_path.exists():
        raise FileNotFoundError(f"deal_state.json not found: {deal_state_path}")
    with open(deal_state_path) as f:
        return json.load(f)


def write_excel_file(df: pd.DataFrame, filepath: Path, sheet_name: str = "Sheet1"):
    """Write DataFrame to Excel with formatting."""
    filepath.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Format worksheet
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        # Apply header styling
        for col_num, col_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Apply borders and alignment to data
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        # Auto-adjust column widths
        for col_num, col_title in enumerate(df.columns, 1):
            max_length = max(
                df[col_title].astype(str).apply(len).max(),
                len(col_title)
            )
            col_letter = get_column_letter(col_num)
            worksheet.column_dimensions[col_letter].width = min(max_length + 2, 50)


def generate_entity_chart(deal_state: dict) -> pd.DataFrame:
    """Generate entity chart from deal_state.entities."""
    company = deal_state["company"]
    entities = deal_state.get("entities", [])

    rows = []

    # Add parent company as first row
    parent_entity_id = f"E{len(entities)+1:04d}"
    rows.append({
        "entity_id": parent_entity_id,
        "entity_name": company["legal_name"],
        "entity_type": company["entity_type"],
        "jurisdiction": company["state_of_formation"],
        "state_of_formation": company["state_of_formation"],
        "status": "active",
        "parent_entity": "N/A",
        "ownership_pct": "100.0%",
        "ein": company.get("ein", generate_ein()),
        "date_formed": f"{company['founded']}-01-01",
        "purpose": "Operating company",
    })

    # Add subsidiaries
    for entity in entities:
        rows.append({
            "entity_id": entity["entity_id"],
            "entity_name": entity["name"],
            "entity_type": entity["type"],
            "jurisdiction": entity["jurisdiction"],
            "state_of_formation": entity["jurisdiction"],
            "status": entity.get("status", "active"),
            "parent_entity": parent_entity_id,
            "ownership_pct": f"{random.uniform(50, 100):.1f}%",
            "ein": generate_ein(),
            "date_formed": f"{company['founded'] + random.randint(0, 5)}-{random.randint(1,12):02d}-01",
            "purpose": "Operating subsidiary" if random.random() < 0.5 else "Holding company",
        })

    return pd.DataFrame(rows)


def generate_cap_table(deal_state: dict) -> pd.DataFrame:
    """Generate capitalization table from deal_state.ownership."""
    ownership = deal_state.get("ownership", {})
    shareholders = ownership.get("shareholders", [])
    option_pool_pct = ownership.get("option_pool_pct", 10)

    rows = []

    # Add shareholders
    for shareholder in shareholders:
        rows.append({
            "shareholder_name": shareholder["name"],
            "shareholder_type": shareholder.get("type", "individual"),
            "shares_held": random.randint(100000, 5000000),
            "pct_ownership": f"{shareholder['pct']:.2f}%",
            "share_class": random.choice(["common", "preferred"]),
            "vesting_status": "vested" if random.random() < 0.7 else "unvested",
            "grant_date": f"{datetime.now().year - random.randint(1, 10)}-{random.randint(1,12):02d}-{random.randint(1,28):02d}",
            "notes": "Founder" if random.random() < 0.3 else "Early investor",
        })

    # Add option pool
    rows.append({
        "shareholder_name": "Option Pool",
        "shareholder_type": "pool",
        "shares_held": random.randint(50000, 500000),
        "pct_ownership": f"{option_pool_pct:.2f}%",
        "share_class": "options",
        "vesting_status": "n_a",
        "grant_date": "N/A",
        "notes": "Reserved for employee grants",
    })

    return pd.DataFrame(rows)


def generate_debt_schedule(deal_state: dict) -> pd.DataFrame:
    """Generate debt schedule from ownership data."""
    ownership = deal_state.get("ownership", {})
    debt_to_equity = ownership.get("debt_to_equity_ratio", 0.5)
    annual_revenue = deal_state.get("financials_seed", {}).get("annual_revenue", 10000000)

    # Determine number of facilities based on debt ratio and company size
    if debt_to_equity < 0.3:
        num_facilities = 1
    elif debt_to_equity < 1.0:
        num_facilities = random.randint(1, 2)
    else:
        num_facilities = random.randint(2, 3)

    rows = []
    total_debt = annual_revenue * debt_to_equity / (1 + debt_to_equity)  # Back into debt from D/E

    facility_types = ["revolver", "term_loan", "line_of_credit"]

    for i in range(num_facilities):
        commitment = round(total_debt / num_facilities * random.uniform(0.8, 1.2), 0)
        drawn_pct = random.uniform(0.3, 0.95)
        outstanding = round(commitment * drawn_pct, 0)

        maturity_years = random.randint(2, 7)
        maturity_date = (datetime.now() + timedelta(days=365 * maturity_years)).strftime("%Y-%m-%d")

        rows.append({
            "facility_name": f"Facility {chr(65 + i)}",
            "lender": random.choice(BANK_NAMES),
            "facility_type": random.choice(facility_types),
            "commitment_amount": f"${commitment:,.0f}",
            "outstanding_balance": f"${outstanding:,.0f}",
            "interest_rate": f"{random.uniform(4.0, 9.0):.2f}%",
            "maturity_date": maturity_date,
            "collateral": ", ".join(random.sample(COLLATERAL_TYPES, random.randint(1, 3))),
            "covenant_summary": random.choice([
                "Maintain DSCR > 1.25x",
                "Minimum liquidity $500K",
                "Debt/EBITDA < 3.0x",
                "Maintain profitability",
            ]),
            "drawn_pct": f"{drawn_pct*100:.1f}%",
        })

    return pd.DataFrame(rows)


def generate_bank_account_list(deal_state: dict) -> pd.DataFrame:
    """Generate bank account list based on company size."""
    management = deal_state.get("management", [])
    company = deal_state["company"]

    # Extract CEO and CFO names
    ceo_name = "TBD"
    cfo_name = "TBD"
    for person in management:
        if "Chief Executive Officer" in person.get("title", ""):
            ceo_name = person["name"]
        elif "Chief Financial Officer" in person.get("title", ""):
            cfo_name = person["name"]

    # Determine number of accounts based on company size
    size = deal_state.get("metadata", {}).get("size", "mid")
    account_counts = {"small": 2, "mid": random.randint(3, 4), "large": random.randint(4, 5)}
    num_accounts = account_counts.get(size, 3)

    rows = []
    account_types = ["operating", "payroll", "savings", "escrow"]

    # Operating account at major bank (first)
    operating_bank = random.choice(MAJOR_BANKS)
    rows.append({
        "account_name": "Operating Account",
        "bank_name": operating_bank,
        "account_type": "operating",
        "account_number_last4": f"{random.randint(1000, 9999)}",
        "authorized_signers": f"{ceo_name}, {cfo_name}",
        "purpose": "Primary checking for operations",
    })

    # Payroll account (may be same or different bank)
    payroll_bank = operating_bank if random.random() < 0.4 else random.choice(MAJOR_BANKS)
    rows.append({
        "account_name": "Payroll Account",
        "bank_name": payroll_bank,
        "account_type": "payroll",
        "account_number_last4": f"{random.randint(1000, 9999)}",
        "authorized_signers": f"{cfo_name}, HR Manager",
        "purpose": "Employee payroll disbursement",
    })

    # Additional accounts
    for i in range(num_accounts - 2):
        account_type = random.choice(account_types[2:])
        rows.append({
            "account_name": f"{account_type.capitalize()} Account",
            "bank_name": random.choice(BANK_NAMES),
            "account_type": account_type,
            "account_number_last4": f"{random.randint(1000, 9999)}",
            "authorized_signers": f"{ceo_name}, {cfo_name}",
            "purpose": f"{account_type.replace('_', ' ').title()}",
        })

    return pd.DataFrame(rows)


def generate_related_party_register(deal_state: dict) -> pd.DataFrame:
    """Generate related party transaction register."""
    realism_mode = deal_state.get("metadata", {}).get("realism_mode", "realistic")
    management = deal_state.get("management", [])
    company = deal_state["company"]
    annual_revenue = deal_state.get("financials_seed", {}).get("annual_revenue", 10000000)

    # Determine number of related party transactions
    if realism_mode == "clean":
        num_transactions = random.randint(0, 1)
    elif realism_mode == "realistic":
        num_transactions = random.randint(1, 3)
    else:  # messy
        num_transactions = random.randint(2, 4)

    rows = []

    transaction_types_list = [
        ("lease", "Founder leasing building to company", 0.03),
        ("service", "Management company fees", 0.02),
        ("loan", "Shareholder loan", 0.05),
        ("lease", "Family member equipment lease", 0.01),
        ("service", "Consulting services", 0.015),
    ]

    for i in range(num_transactions):
        rel_name, rel_type_desc, revenue_pct = random.choice(transaction_types_list)

        annual_amount = round(annual_revenue * revenue_pct, 0)
        start_date = (datetime.now() - timedelta(days=random.randint(180, 1825))).strftime("%Y-%m-%d")

        # Pick related party
        if random.random() < 0.5 and management:
            counterparty = random.choice(management)["name"]
        else:
            counterparty = f"{fake.name()} (shareholder)"

        arms_length = "yes" if random.random() < 0.6 else (
            "under_review" if random.random() < 0.5 else "no"
        )

        rows.append({
            "counterparty_name": counterparty,
            "relationship": rel_type_desc,
            "transaction_type": rel_type_desc.split()[0].lower(),
            "annual_amount": f"${annual_amount:,.0f}",
            "terms": random.choice([
                "Monthly payment",
                "Quarterly invoicing",
                "Annual contract",
                "As-incurred basis"
            ]),
            "start_date": start_date,
            "arms_length": arms_length,
        })

    # If no transactions in clean mode, return empty DataFrame
    if not rows:
        rows.append({
            "counterparty_name": "N/A",
            "relationship": "None",
            "transaction_type": "N/A",
            "annual_amount": "$0",
            "terms": "N/A",
            "start_date": "N/A",
            "arms_length": "N/A",
        })

    return pd.DataFrame(rows)


def generate_ein() -> str:
    """Generate a fake but correctly formatted EIN."""
    return f"{random.randint(10, 99)}-{random.randint(1000000, 9999999)}"


def main():
    parser = argparse.ArgumentParser(
        description="Generate Section 1.0 Corporate artifacts from deal_state.json"
    )
    parser.add_argument(
        "--output-dir",
        required=True,
        help="Output directory containing deal_state.json and for writing corporate artifacts"
    )
    parser.add_argument(
        "--seed",
        type=int,
        help="Random seed for reproducibility"
    )

    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    corporate_dir = output_dir / "1.0-corporate"

    # Set seed if provided
    if args.seed:
        random.seed(args.seed)
        Faker.seed(args.seed)

    # Load deal state
    print(f"Loading deal_state.json from {output_dir}...")
    deal_state = load_deal_state(output_dir)

    company_name = deal_state["company"]["name"]
    print(f"Generating corporate artifacts for {company_name}...\n")

    # Create corporate directory
    corporate_dir.mkdir(parents=True, exist_ok=True)

    # Generate and write entity chart
    print("  Generating entity_chart.xlsx...")
    entity_chart = generate_entity_chart(deal_state)
    write_excel_file(entity_chart, corporate_dir / "entity_chart.xlsx")

    # Generate and write cap table
    print("  Generating cap_table.xlsx...")
    cap_table = generate_cap_table(deal_state)
    write_excel_file(cap_table, corporate_dir / "cap_table.xlsx")

    # Generate and write debt schedule
    print("  Generating debt_schedule.xlsx...")
    debt_schedule = generate_debt_schedule(deal_state)
    write_excel_file(debt_schedule, corporate_dir / "debt_schedule.xlsx")

    # Generate and write bank account list
    print("  Generating bank_account_list.xlsx...")
    bank_accounts = generate_bank_account_list(deal_state)
    write_excel_file(bank_accounts, corporate_dir / "bank_account_list.xlsx")

    # Generate and write related party register
    print("  Generating related_party_register.xlsx...")
    related_party = generate_related_party_register(deal_state)
    write_excel_file(related_party, corporate_dir / "related_party_register.xlsx")

    # Print summary
    print(f"\nGenerated corporate artifacts in {corporate_dir}:")
    print(f"  entity_chart.xlsx ({len(entity_chart)} entities)")
    print(f"  cap_table.xlsx ({len(cap_table)} shareholders/pool)")
    print(f"  debt_schedule.xlsx ({len(debt_schedule)} facilities)")
    print(f"  bank_account_list.xlsx ({len(bank_accounts)} accounts)")
    print(f"  related_party_register.xlsx ({len(related_party)} transactions)")
    print(f"\nDone!")


if __name__ == "__main__":
    main()
