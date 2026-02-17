"""Validate template v3 is deal-agnostic and structurally usable."""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


ROOT = Path(__file__).resolve().parent
WB_PATH = ROOT / "databook-template-v3.xlsx"

CORE_VISIBLE_ORDER = [
    "Project North >>>",
    "Disclaimer",
    "QofE >>",
    "QofE | Bridge",
    "QofE | Summary",
    "QofE | Detail",
    "Net debt>>",
    "Net debt",
    "Financials>>",
    "Income statement",
    "Balance sheet",
    "Cash flow",
    "Personnel>>",
    "IS | Payroll costs",
    "Headcount census",
    "Other analysis>>",
    "Top customers",
    "Revenue by customer group",
    "Recons >>",
    "Recons | IS",
    "Recons | BS",
    "Recon | Cash flow",
    "Recon | Billings",
    "Checks>>",
    "Checks | Core",
    "Audit>>",
    "Audit | Issues_Index",
    "Audit | Run_Log",
]

SYS_TABS = ["_Sys|Trial_Balance", "_Sys|Line_Map", "_Sys|Tab_Manifest", "_Sys|Config"]
MONTHLY_MANIFEST = {
    "Income statement": (9, "F"),
    "Balance sheet": (9, "F"),
    "QofE | Detail": (8, "E"),
    "Net debt": (9, "E"),
    "Recon | Billings": (8, "F"),
}

DEAL_KEYWORDS = ["everlink", "interac", "desjardins", "kyndryl", "cuca"]
REF_RE = re.compile(r"'([^']+)'!")


class Check:
    def __init__(self) -> None:
        self.errors: list[str] = []

    def ok(self, cond: bool, msg: str) -> None:
        if not cond:
            self.errors.append(msg)


def main() -> int:
    c = Check()
    if not WB_PATH.exists():
        print(f"FAIL\n- Missing workbook: {WB_PATH}")
        return 1

    wb = load_workbook(WB_PATH, data_only=False, read_only=False)

    visible = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    c.ok(visible == CORE_VISIBLE_ORDER, "Visible tab order mismatch")
    expected_all = CORE_VISIBLE_ORDER + SYS_TABS
    c.ok(wb.sheetnames == expected_all, "Workbook contains unexpected tabs or tab order drift")

    for t in SYS_TABS:
        c.ok(t in wb.sheetnames, f"Missing system tab {t}")
        if t in wb.sheetnames:
            c.ok(wb[t].sheet_state == "hidden", f"System tab not hidden: {t}")

    # 36-month headers
    for tab, (row, start_col) in MONTHLY_MANIFEST.items():
        ws = wb[tab]
        s = column_index_from_string(start_col)
        first = ws.cell(row, s).value
        second = ws.cell(row, s + 1).value
        trailing = ws.cell(row, s + 36).value
        c.ok(isinstance(first, str) and "ctl_month_start" in first, f"{tab}: first monthly header not ctl_month_start")
        c.ok(isinstance(second, str) and "EDATE(ctl_month_start,1)" in second, f"{tab}: second monthly header not contiguous")
        c.ok(trailing is None, f"{tab}: trailing header after 36 months not cleared")

    # deal keyword sweep on all tabs (materialized cells only for speed)
    keyword_hits = []
    for ws in wb.worksheets:
        for cell in ws._cells.values():
            v = cell.value
            if isinstance(v, str) and not v.startswith("="):
                low = v.lower()
                for k in DEAL_KEYWORDS:
                    if k in low:
                        keyword_hits.append(f"{ws.title}!{cell.coordinate}:{k}:{v[:80]}")
    c.ok(len(keyword_hits) == 0, "Deal-specific keywords remain:\n" + "\n".join(keyword_hits[:40]))

    # non-core dependency sweep on visible formulas (materialized cells only for speed)
    core = set(CORE_VISIBLE_ORDER + SYS_TABS)
    bad_refs = []
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        for cell in ws._cells.values():
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                refs = REF_RE.findall(v)
                if any(ref not in core for ref in refs):
                    bad_refs.append(f"{ws.title}!{cell.coordinate}:{v[:120]}")
    c.ok(len(bad_refs) == 0, "Formulas still reference non-core tabs:\n" + "\n".join(bad_refs[:40]))

    cfg = wb["_Sys|Config"]
    ctl_project_name = None
    for r in range(2, 200):
        key = cfg.cell(r, 1).value
        if key == "ctl_project_name":
            ctl_project_name = cfg.cell(r, 2).value
            break
    c.ok(ctl_project_name == "Project Name", "_Sys|Config ctl_project_name must be 'Project Name'")

    if c.errors:
        print("FAIL")
        for e in c.errors:
            print(f"- {e}")
        return 1

    print("PASS")
    print(f"Validated workbook: {WB_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
