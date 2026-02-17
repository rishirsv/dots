"""Build Databook Template v3 as a truly deal-agnostic North-style template.

Starts from template-v2 workbook and aggressively sanitizes deal-specific data,
keywords, and non-core formula dependencies while preserving North visual layout.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


ROOT = Path(__file__).resolve().parent
BASE_PATH = ROOT / "seeds-v2-base.xlsx"
OUT_PATH = ROOT / "databook-template-v3.xlsx"
REPORT_PATH = ROOT / "reports" / "v3-sanitization-report.md"

CORE_VISIBLE_ORDER = [
    "Project North >>>",
    "Disclaimer",
    "QofE >>",
    "QofE | Bridge",
    "QofE | Summary",
    "QofE | Detail",
    "Net debt>>",
    "Net debt",
    "Financials>>",
    "Income statement",
    "Balance sheet",
    "Cash flow",
    "Personnel>>",
    "IS | Payroll costs",
    "Headcount census",
    "Other analysis>>",
    "Top customers",
    "Revenue by customer group",
    "Recons >>",
    "Recons | IS",
    "Recons | BS",
    "Recon | Cash flow",
    "Recon | Billings",
    "Checks>>",
    "Checks | Core",
    "Audit>>",
    "Audit | Issues_Index",
    "Audit | Run_Log",
]

SYS_TABS = ["_Sys|Trial_Balance", "_Sys|Line_Map", "_Sys|Tab_Manifest", "_Sys|Config"]
SECTION_TABS = {t for t in CORE_VISIBLE_ORDER if t.endswith(">>")}
NON_DATA_TABS = SECTION_TABS | {"Disclaimer", "Checks | Core", "Audit | Issues_Index", "Audit | Run_Log"}

# Core tabs that should be treated as monthly grids (36 months) via manifest.
MONTHLY_MANIFEST = {
    "Income statement": (9, "F"),
    "Balance sheet": (9, "F"),
    "QofE | Detail": (8, "E"),
    "Net debt": (9, "E"),
    "Recon | Billings": (8, "F"),
}

DEAL_KEYWORDS = [
    "everlink",
    "project north",
    "interac",
    "desjardins",
    "kyndryl",
    "cuca",
    "management provided information",
]

FORMULA_SHEET_REF_RE = re.compile(r"'([^']+)'!")


def is_date_or_number(value) -> bool:
    """Return true when value is a numeric/date payload to be cleared."""

    return isinstance(value, (int, float, datetime, date))


def sanitize_banner(ws) -> None:
    """Normalize top banner cells for reusable template use."""

    if ws.title in SECTION_TABS:
        ws["A1"] = ws.title
        return

    # Keep subtitle/title row but genericize project/source.
    if ws["B1"].value is not None:
        ws["B1"] = "Project Name"
    if ws["B3"].value is not None:
        ws["B3"] = "Source: TBD"


def clear_comments_columns(ws) -> int:
    """Clear values below any detected comments/questions headers."""

    cleared = 0
    headers = []
    for r in range(6, 16):
        for c in range(1, min(ws.max_column, 220) + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str):
                low = v.lower()
                if "comment" in low or "question" in low:
                    headers.append((r, c))

    for hdr_row, hdr_col in headers:
        for row in range(hdr_row + 1, min(ws.max_row, 1600) + 1):
            cell = ws.cell(row, hdr_col)
            if cell.value is not None:
                cell.value = None
                cleared += 1

    return cleared


def clear_formula_refs_to_noncore(ws, core_set: set[str]) -> int:
    """Clear formulas that still depend on hidden/non-core deal-specific tabs."""

    cleared = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 1600), min_col=1, max_col=min(ws.max_column, 260)):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                refs = FORMULA_SHEET_REF_RE.findall(v)
                if any(ref not in core_set for ref in refs):
                    cell.value = None
                    cleared += 1
    return cleared


def strip_deal_keywords(ws) -> int:
    """Remove/neutralize deal-specific text in visible core tabs."""

    changed = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 1600), min_col=1, max_col=min(ws.max_column, 260)):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and not v.startswith("="):
                low = v.lower()
                if any(k in low for k in DEAL_KEYWORDS):
                    # Keep structure text only for high-level lines; otherwise blank it.
                    if cell.coordinate in {"B1", "B3"}:
                        continue
                    if len(v) > 45:
                        cell.value = "TBD"
                    else:
                        # Replace known tokens but keep readable label when possible.
                        nv = v
                        for k in DEAL_KEYWORDS:
                            pattern = re.compile(re.escape(k), re.IGNORECASE)
                            nv = pattern.sub("", nv)
                        nv = re.sub(r"\s{2,}", " ", nv).strip(" -_,")
                        cell.value = nv if nv else None
                    changed += 1
    return changed


def clear_payload_data(ws, monthly: bool = False, header_row: int | None = None, start_col: str | None = None) -> int:
    """Clear numbers/formulas from data regions while preserving layout labels."""

    cleared = 0

    # Monthly tabs: clear only monthly grid payload rows, keep descriptors on left.
    if monthly and header_row and start_col:
        start_idx = column_index_from_string(start_col)
        end_idx = start_idx + 35

        for r in range(header_row + 1, min(ws.max_row, 1600) + 1):
            left_text = " ".join(
                str(ws.cell(r, c).value or "")
                for c in range(1, min(start_idx, 8))
            ).lower()
            keep_row = "check" in left_text
            for c in range(start_idx, min(end_idx, ws.max_column) + 1):
                cell = ws.cell(r, c)
                if keep_row:
                    continue
                if cell.value is not None and (
                    isinstance(cell.value, str) and cell.value.startswith("=") or is_date_or_number(cell.value)
                ):
                    cell.value = None
                    cleared += 1
        return cleared

    # Non-monthly core tabs: clear broad numeric/formula payload from cols >= E rows >= 10.
    for r in range(10, min(ws.max_row, 1600) + 1):
        left_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, 7)).lower()
        keep_row = "check" in left_text or "total" in left_text
        for c in range(5, min(ws.max_column, 260) + 1):
            cell = ws.cell(r, c)
            if keep_row:
                continue
            if cell.value is not None and (
                isinstance(cell.value, str) and cell.value.startswith("=") or is_date_or_number(cell.value)
            ):
                cell.value = None
                cleared += 1

    return cleared


def enforce_36_month_headers(ws, header_row: int, start_col: str) -> None:
    """Reset monthly header row to control-driven 36 month sequence."""

    start_idx = column_index_from_string(start_col)
    for i in range(36):
        cell = ws.cell(header_row, start_idx + i)
        if i == 0:
            cell.value = "=ctl_month_start"
        else:
            cell.value = f"=EDATE(ctl_month_start,{i})"
    for c in range(start_idx + 36, start_idx + 180):
        ws.cell(header_row, c).value = None


def tighten_sys_config(wb) -> None:
    """Ensure locked monthly controls are present in `_Sys|Config`."""

    ws = wb["_Sys|Config"]
    mapping = {}
    row = 2
    while True:
        key = ws.cell(row, 1).value
        if key is None:
            break
        mapping[str(key)] = row
        row += 1

    if "ctl_monthly_period_count" in mapping:
        ws.cell(mapping["ctl_monthly_period_count"], 2).value = 36
    if "ctl_project_name" in mapping:
        ws.cell(mapping["ctl_project_name"], 2).value = "Project Name"


def prune_to_core_and_sys_tabs(wb) -> int:
    """Remove all non-core/non-system tabs so template carries no legacy deal sheets."""

    keep = set(CORE_VISIBLE_ORDER + SYS_TABS)
    removed = 0
    for name in list(wb.sheetnames):
        if name not in keep:
            wb.remove(wb[name])
            removed += 1
    return removed


def enforce_tab_states_and_order(wb) -> None:
    """Set final visibility and deterministic tab order."""

    for name in CORE_VISIBLE_ORDER:
        wb[name].sheet_state = "visible"
    for name in SYS_TABS:
        wb[name].sheet_state = "hidden"
    wb._sheets = [wb[name] for name in (CORE_VISIBLE_ORDER + SYS_TABS)]


def main() -> None:
    """Build template v3 from v2 baseline with robust sanitization."""

    if not BASE_PATH.exists():
        raise FileNotFoundError(f"Missing base workbook: {BASE_PATH}")

    wb = load_workbook(BASE_PATH, data_only=False)

    core_set = set(CORE_VISIBLE_ORDER + SYS_TABS)

    summary = {
        "cleared_comments": 0,
        "cleared_noncore_formulas": 0,
        "cleared_payload_cells": 0,
        "keyword_strips": 0,
        "removed_noncore_tabs": 0,
    }

    for ws in wb.worksheets:
        if ws.title in CORE_VISIBLE_ORDER and ws.title not in NON_DATA_TABS:
            sanitize_banner(ws)

            if ws.title in MONTHLY_MANIFEST:
                hdr_row, start_col = MONTHLY_MANIFEST[ws.title]
                enforce_36_month_headers(ws, hdr_row, start_col)
                summary["cleared_payload_cells"] += clear_payload_data(
                    ws,
                    monthly=True,
                    header_row=hdr_row,
                    start_col=start_col,
                )
            else:
                summary["cleared_payload_cells"] += clear_payload_data(ws, monthly=False)

            summary["cleared_comments"] += clear_comments_columns(ws)
            summary["cleared_noncore_formulas"] += clear_formula_refs_to_noncore(ws, core_set)
            summary["keyword_strips"] += strip_deal_keywords(ws)

    # Disclaimer may still contain deal-name references in legal text; neutralize those.
    if "Disclaimer" in wb.sheetnames:
        dws = wb["Disclaimer"]
        for row in dws.iter_rows(min_row=1, max_row=min(dws.max_row, 80), min_col=1, max_col=min(dws.max_column, 12)):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and not v.startswith("="):
                    low = v.lower()
                    if "everlink" in low or "project north" in low:
                        nv = re.sub(r"everlink", "Client Company", v, flags=re.IGNORECASE)
                        nv = re.sub(r"project north", "Target", nv, flags=re.IGNORECASE)
                        cell.value = nv
                        summary["keyword_strips"] += 1

    summary["removed_noncore_tabs"] = prune_to_core_and_sys_tabs(wb)
    tighten_sys_config(wb)
    enforce_tab_states_and_order(wb)
    wb.save(OUT_PATH)

    REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    REPORT_PATH.write_text(
        "\n".join(
            [
                "# Template V3 Sanitization Report",
                "",
                f"- Base: `{BASE_PATH}`",
                f"- Output: `{OUT_PATH}`",
                f"- Cleared comment cells: **{summary['cleared_comments']}**",
                f"- Cleared formulas referencing non-core tabs: **{summary['cleared_noncore_formulas']}**",
                f"- Cleared payload cells: **{summary['cleared_payload_cells']}**",
                f"- Text keyword sanitizations: **{summary['keyword_strips']}**",
                f"- Removed non-core hidden tabs: **{summary['removed_noncore_tabs']}**",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    print(f"Created: {OUT_PATH}")
    print(f"Report: {REPORT_PATH}")


if __name__ == "__main__":
    main()
