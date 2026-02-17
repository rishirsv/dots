"""Build the Databook Analyst v1 template workbook.

This script creates a deterministic `.xlsx` template aligned with the
Databook Analyst v1 plan. It is intentionally opinionated so future runs
produce the same tab order, anchors, and formula scaffolding.
"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_PATH = Path(__file__).resolve().parent / "databook-template-v1.xlsx"


VISIBLE_TABS = [
    "Project_Template>>",
    "Project|Disclaimer",
    "Config|Control",
    "Financials>>",
    "Financials|Income_Statement",
    "Financials|Balance_Sheet",
    "QofE>>",
    "QofE|Summary",
    "QofE|Detail",
    "Net_Debt>>",
    "Net_Debt|Summary",
    "Personnel>>",
    "Personnel|Payroll_Costs",
    "Personnel|Headcount_Census",
    "Other_Analysis>>",
    # Excel tab names are max length 31, so this is shortened from the long form.
    "Other_Analysis|Revenue_By_Cust",
    "Other_Analysis|Top_Customers",
    "Recons>>",
    "Recons|IS",
    "Recons|BS",
    "Recons|Cash_Flow",
    "Recons|Billings",
    "Checks>>",
    "Checks|Core",
    "Audit>>",
    "Audit|Issues_Index",
    "Audit|Run_Log",
]

HIDDEN_SYS_TABS = [
    "_Sys|Trial_Balance",
    "_Sys|Line_Map",
    "_Sys|Tab_Manifest",
    "_Sys|Config",
]

SECTION_TABS = {
    "Project_Template>>",
    "Financials>>",
    "QofE>>",
    "Net_Debt>>",
    "Personnel>>",
    "Other_Analysis>>",
    "Recons>>",
    "Checks>>",
    "Audit>>",
}

STANDARD_PERIOD_TABS = {
    "Financials|Income_Statement",
    "Financials|Balance_Sheet",
    "QofE|Summary",
    "QofE|Detail",
    "Net_Debt|Summary",
    "Recons|IS",
    "Recons|BS",
    "Recons|Cash_Flow",
    "Recons|Billings",
}

OUTPUT_TABS = {
    "Financials|Income_Statement",
    "Financials|Balance_Sheet",
    "QofE|Summary",
    "QofE|Detail",
    "Net_Debt|Summary",
    "Personnel|Payroll_Costs",
    "Personnel|Headcount_Census",
    "Other_Analysis|Revenue_By_Cust",
    "Other_Analysis|Top_Customers",
    "Recons|IS",
    "Recons|BS",
    "Recons|Cash_Flow",
    "Recons|Billings",
}

STATEMENT_TYPES = {
    "Financials|Income_Statement": "IS",
    "Financials|Balance_Sheet": "BS",
    "QofE|Summary": "IS",
    "QofE|Detail": "IS",
    "Net_Debt|Summary": "ND",
    "Personnel|Payroll_Costs": "IS",
    "Personnel|Headcount_Census": "HC",
    "Other_Analysis|Revenue_By_Cust": "SALES",
    "Other_Analysis|Top_Customers": "SALES",
    "Recons|IS": "IS",
    "Recons|BS": "BS",
    "Recons|Cash_Flow": "CF",
    "Recons|Billings": "BILLINGS",
}


@dataclass(frozen=True)
class StylePack:
    """Container for shared workbook styles."""

    banner_fill: PatternFill
    banner_font: Font
    title_font: Font
    header_fill: PatternFill
    header_font: Font
    thin_border: Border
    center: Alignment
    left: Alignment


def build_styles() -> StylePack:
    """Create shared styles to keep the workbook visual language consistent."""

    thin = Side(style="thin", color="D9D9D9")
    return StylePack(
        banner_fill=PatternFill(fill_type="solid", fgColor="0B5CAB"),
        banner_font=Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        title_font=Font(name="Calibri", size=12, bold=True, color="1F1F1F"),
        header_fill=PatternFill(fill_type="solid", fgColor="E8EEF7"),
        header_font=Font(name="Calibri", size=10, bold=True, color="1F1F1F"),
        thin_border=Border(left=thin, right=thin, top=thin, bottom=thin),
        center=Alignment(horizontal="center", vertical="center"),
        left=Alignment(horizontal="left", vertical="center"),
    )


def configure_dimensions(ws) -> None:
    """Set standard column widths and base row heights."""

    ws.column_dimensions["A"].width = 34
    for col in range(2, 14):
        ws.column_dimensions[get_column_letter(col)].width = 14
    ws.column_dimensions["N"].width = 26
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[7].height = 22
    ws.row_dimensions[10].height = 20


def section_tab(ws, title: str, styles: StylePack) -> None:
    """Create a lightweight section divider tab."""

    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="244062")
    ws["A1"].alignment = styles.left
    ws.row_dimensions[1].height = 28
    ws.column_dimensions["A"].width = 48
    ws.protection.sheet = True


def disclaimer_tab(ws, styles: StylePack) -> None:
    """Populate the disclaimer tab."""

    ws["A1"] = "Databook Disclaimer"
    ws["A1"].font = styles.title_font
    ws["A3"] = (
        "This workbook is prepared for transaction support purposes and may contain "
        "draft analyses. Use is restricted to the approved workstream team and client "
        "audience per engagement protocol."
    )
    ws["A5"] = "No external workbook links are permitted in this template."
    ws["A7"] = "All checks with FAIL status must be resolved before final client delivery."
    ws["A3"].alignment = Alignment(wrap_text=True, vertical="top")
    ws["A5"].alignment = styles.left
    ws["A7"].alignment = styles.left
    ws.column_dimensions["A"].width = 120
    ws.row_dimensions[3].height = 64
    ws.protection.sheet = True


def apply_banner(ws, subtitle: str, title: str, styles: StylePack) -> None:
    """Write common banner and title anchors for output tabs."""

    ws["A1"] = "=ctl_project_name"
    ws["A2"] = subtitle
    ws["A3"] = "Source: TBD"
    ws["A4"] = "WORKING_DRAFT"
    ws["A7"] = title
    ws["A8"] = '=ctl_currency&" x"&TEXT(ctl_units_scale,"0")'

    for row in range(1, 5):
        cell = ws.cell(row=row, column=1)
        cell.fill = styles.banner_fill
        cell.font = styles.banner_font
        cell.alignment = styles.left

    ws["A7"].font = styles.title_font
    ws["A8"].font = Font(name="Calibri", size=10, italic=True, color="3F3F3F")


def write_headers(ws, headers: Iterable[str], row: int, styles: StylePack) -> None:
    """Write standardized table headers for a tab."""

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = styles.header_fill
        cell.font = styles.header_font
        cell.border = styles.thin_border
        cell.alignment = styles.center


def add_output_tab_skeleton(ws, styles: StylePack) -> None:
    """Create reusable output scaffold, formulas, and check row."""

    title = ws.title.split("|", 1)[1].replace("_", " ")
    apply_banner(ws, subtitle=title, title=title, styles=styles)

    ws["A6"] = STATEMENT_TYPES.get(ws.title, "GEN")
    ws["A6"].font = Font(name="Calibri", size=9, color="7A7A7A")

    headers = ["Line_Item"] + [f"Period_{idx}" for idx in range(1, 13)] + ["Comments"]
    write_headers(ws, headers, row=10, styles=styles)

    if ws.title in STANDARD_PERIOD_TABS:
        for col in range(2, 14):
            cell = ws.cell(row=10, column=col)
            coord = f"${get_column_letter(col)}$10"
            cell.value = (
                "=IF(COLUMNS($B10:"
                f"{get_column_letter(col)}10)<=ctl_std_period_count,"
                "IF(ctl_std_periodicity=\"Monthly\","
                "EDATE(ctl_std_period_start,COLUMNS($B10:"
                f"{get_column_letter(col)}10)-1),"
                "EDATE(ctl_std_period_start,(COLUMNS($B10:"
                f"{get_column_letter(col)}10)-1)*3)),\"\")"
            )
            cell.number_format = "mmm-yy"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.comment = None

    line_items = {
        "IS": [
            "Revenue",
            "COGS",
            "Gross_Profit",
            "Personnel_Costs",
            "Other_Opex",
            "EBITDA",
            "Depreciation_Amortization",
            "EBIT",
            "Net_Financial_Items",
            "EBT",
            "Taxes",
            "Net_Income",
        ],
        "BS": [
            "Cash",
            "AR",
            "Inventory",
            "Other_Current_Assets",
            "Fixed_Assets",
            "Total_Assets",
            "AP",
            "Accruals",
            "Debt",
            "Other_Liabilities",
            "Equity",
            "Total_Liabilities_Equity",
        ],
        "ND": [
            "Cash_and_Cash_Equivalents",
            "Short_Term_Debt",
            "Long_Term_Debt",
            "Lease_Liabilities",
            "Other_Debt_Items",
            "Net_Debt",
        ],
        "CF": [
            "Net_Income",
            "Non_Cash_Adjustments",
            "Working_Capital_Change",
            "Capex",
            "Financing_Cash_Flow",
            "Net_Change_In_Cash",
        ],
        "BILLINGS": [
            "Billed_Revenue",
            "Credits",
            "Net_Billings",
            "Recognized_Revenue",
            "Variance",
        ],
        "HC": [
            "Headcount_Beginning",
            "New_Hires",
            "Leavers",
            "Headcount_Ending",
            "Payroll_Cost",
            "Cost_Per_Head",
        ],
        "SALES": [
            "Customer_Group_A",
            "Customer_Group_B",
            "Customer_Group_C",
            "Customer_Group_D",
            "Top_10",
            "Total_Sales",
        ],
        "GEN": [
            "Line_1",
            "Line_2",
            "Line_3",
            "Line_4",
            "Line_5",
        ],
    }

    statement_type = STATEMENT_TYPES.get(ws.title, "GEN")
    selected_lines = line_items.get(statement_type, line_items["GEN"])

    for idx, line in enumerate(selected_lines, start=11):
        ws.cell(row=idx, column=1, value=line)
        for col in range(2, 14):
            formula = (
                "=IF($A"
                f"{idx}=\"\",\"\",SUMIFS('_Sys|Trial_Balance'!$K:$K,"
                "'_Sys|Trial_Balance'!$A:$A,"
                f"{get_column_letter(col)}$10,'_Sys|Trial_Balance'!$I:$I,$A{idx},"
                "'_Sys|Trial_Balance'!$L:$L,$A$6))"
            )
            ws.cell(row=idx, column=col, value=formula)

    total_row = 31
    ws.cell(row=total_row, column=1, value="Total")
    ws.cell(row=total_row, column=1).font = Font(name="Calibri", size=10, bold=True)

    for col in range(2, 14):
        ws.cell(
            row=total_row,
            column=col,
            value=f"=SUM({get_column_letter(col)}11:{get_column_letter(col)}30)",
        )
        ws.cell(row=total_row, column=col).font = Font(name="Calibri", size=10, bold=True)

    check_row = 33
    ws.cell(row=check_row, column=1, value="Check")
    for col in range(2, 14):
        formula = (
            "=IF(ABS(SUMIFS('_Sys|Trial_Balance'!$K:$K,'_Sys|Trial_Balance'!$A:$A,"
            f"{get_column_letter(col)}$10,'_Sys|Trial_Balance'!$L:$L,$A$6)-"
            f"{get_column_letter(col)}$31)<=ctl_tolerance_abs,\"PASS\",\"FAIL\")"
        )
        ws.cell(row=check_row, column=col, value=formula)

    ws.freeze_panes = "B11"
    configure_dimensions(ws)


def style_grid(ws, min_row: int, max_row: int, min_col: int, max_col: int, styles: StylePack) -> None:
    """Apply grid styling for data regions."""

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = styles.thin_border
            if cell.column == 1:
                cell.alignment = styles.left
            else:
                cell.alignment = styles.center
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0"


def add_control_tab(ws, styles: StylePack) -> None:
    """Create control panel inputs, validations, and named-range anchors."""

    ws["A1"] = "Databook_Control_Panel"
    ws["A1"].font = styles.title_font

    controls = [
        ("ctl_project_name", "Project_North"),
        ("ctl_report_date", datetime(2026, 2, 17)),
        ("ctl_currency", "USD"),
        ("ctl_units_scale", 1000),
        ("ctl_sign_convention", "Source_Native"),
        ("ctl_tolerance_abs", 1),
        ("ctl_tolerance_pct", 0.001),
        ("ctl_preparer", ""),
        ("", ""),
        ("ctl_std_periodicity", "Monthly"),
        ("ctl_std_period_start", datetime(2024, 1, 31)),
        ("ctl_std_period_end", datetime(2024, 12, 31)),
        ("ctl_std_period_count", 12),
    ]

    start_row = 2
    named_cells: list[tuple[str, str]] = []

    for offset, (label, default_value) in enumerate(controls):
        row = start_row + offset
        if not label:
            continue
        ws.cell(row=row, column=1, value=label)
        value_cell = ws.cell(row=row, column=2, value=default_value)
        value_cell.fill = PatternFill(fill_type="solid", fgColor="FFFCE6")
        value_cell.border = styles.thin_border
        value_cell.alignment = styles.left
        if "date" in label or "start" in label or "end" in label:
            value_cell.number_format = "yyyy-mm-dd"
        if label == "ctl_tolerance_pct":
            value_cell.number_format = "0.00%"
        named_cells.append((label, value_cell.coordinate))

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 26
    ws["A12"] = "Standard period controls apply to IS/BS/QofE/Net Debt/Recons tabs only."
    ws["A12"].font = Font(name="Calibri", size=9, italic=True, color="4F4F4F")

    list_currency = DataValidation(type="list", formula1='"USD,CAD,EUR,GBP"', allow_blank=False)
    list_units = DataValidation(type="list", formula1='"1,1000,1000000"', allow_blank=False)
    list_sign = DataValidation(
        type="list",
        formula1='"Source_Native,Expenses_Negative,Expenses_Positive"',
        allow_blank=False,
    )
    list_periodicity = DataValidation(type="list", formula1='"Monthly,Quarterly"', allow_blank=False)

    ws.add_data_validation(list_currency)
    ws.add_data_validation(list_units)
    ws.add_data_validation(list_sign)
    ws.add_data_validation(list_periodicity)

    list_currency.add("B4")
    list_units.add("B5")
    list_sign.add("B6")
    list_periodicity.add("B11")

    for name, coord in named_cells:
        wb_ref = f"'Config|Control'!${coord[0]}${coord[1:]}"
        ws.parent.defined_names.add(DefinedName(name=name, attr_text=wb_ref))

    ws.freeze_panes = "A2"
    ws.protection.sheet = True
    # Unlock control value cells while keeping formulas/headings protected.
    for row in range(2, 15):
        ws.cell(row=row, column=2).protection = Protection(locked=False)


def add_trial_balance_tab(ws, styles: StylePack) -> None:
    """Build the minimal TB staging tab used for direct output translation."""

    headers = [
        "period_end",
        "account_code",
        "account_name",
        "amount",
        "entity",
        "source_file",
        "source_tab",
        "note",
        "map_line",
        "map_sign",
        "signed_amount",
        "map_statement_type",
    ]
    write_headers(ws, headers, row=1, styles=styles)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 28
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 20

    for row in range(2, 2001):
        ws.cell(row=row, column=9, value=f"=XLOOKUP(B{row},'_Sys|Line_Map'!$A:$A,'_Sys|Line_Map'!$C:$C,\"MISSING\",0)")
        ws.cell(row=row, column=10, value=f"=XLOOKUP(B{row},'_Sys|Line_Map'!$A:$A,'_Sys|Line_Map'!$D:$D,1,0)")
        ws.cell(row=row, column=11, value=f"=IF(D{row}=\"\",\"\",D{row}*J{row})")
        ws.cell(row=row, column=12, value=f"=XLOOKUP(B{row},'_Sys|Line_Map'!$A:$A,'_Sys|Line_Map'!$B:$B,\"MISSING\",0)")

    for row in range(2, 2001):
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = styles.thin_border

    ws.protection.sheet = True
    for row in range(2, 2001):
        for col in range(1, 9):
            ws.cell(row=row, column=col).protection = Protection(locked=False)


def add_line_map_tab(ws, styles: StylePack) -> None:
    """Build account-to-line mapping table used by XLOOKUP in TB tab."""

    headers = [
        "account_code",
        "statement_type",
        "report_line",
        "sign_factor",
        "active_flag",
        "notes",
    ]
    write_headers(ws, headers, row=1, styles=styles)

    seed_rows = [
        ("4000", "IS", "Revenue", 1, "Y", "Example revenue mapping"),
        ("5000", "IS", "COGS", -1, "Y", "Example cogs mapping"),
        ("6100", "IS", "Personnel_Costs", -1, "Y", "Example payroll mapping"),
        ("1000", "BS", "Cash", 1, "Y", "Example balance sheet mapping"),
        ("1200", "BS", "AR", 1, "Y", "Example balance sheet mapping"),
        ("2000", "BS", "AP", -1, "Y", "Example balance sheet mapping"),
        ("3000", "BS", "Equity", -1, "Y", "Example balance sheet mapping"),
        ("7000", "ND", "Short_Term_Debt", -1, "Y", "Example net debt mapping"),
        ("7100", "ND", "Long_Term_Debt", -1, "Y", "Example net debt mapping"),
        ("8000", "CF", "Net_Income", 1, "Y", "Example cash flow mapping"),
        ("9000", "BILLINGS", "Billed_Revenue", 1, "Y", "Example billings mapping"),
    ]

    for row_idx, values in enumerate(seed_rows, start=2):
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for row in range(2, 2001):
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = styles.thin_border
            ws.cell(row=row, column=col).protection = Protection(locked=False)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 36
    ws.freeze_panes = "A2"
    ws.protection.sheet = True


def add_checks_tab(ws, styles: StylePack) -> None:
    """Build core checks covering control, mappings, tie-outs, and conformance."""

    apply_banner(ws, subtitle="Checks Core", title="Checks Core", styles=styles)
    headers = [
        "Check_ID",
        "Description",
        "Status",
        "Actual",
        "Expected",
        "Severity",
        "Linked_Tab",
        "Notes",
    ]
    write_headers(ws, headers, row=10, styles=styles)

    checks = [
        ("CHK-001", "Required control values populated", "HIGH", "Config|Control"),
        ("CHK-002", "Unmapped TB accounts (MISSING)", "HIGH", "_Sys|Trial_Balance"),
        ("CHK-003", "No spaces in tab names", "MED", "_Sys|Tab_Manifest"),
        ("CHK-004", "Comments header exists across output tabs", "MED", "_Sys|Tab_Manifest"),
        ("CHK-005", "IS first period tie-out to TB", "HIGH", "Financials|Income_Statement"),
        ("CHK-006", "BS first period tie-out to TB", "HIGH", "Financials|Balance_Sheet"),
    ]

    start_row = 11
    for idx, (check_id, description, severity, linked_tab) in enumerate(checks, start=start_row):
        ws.cell(row=idx, column=1, value=check_id)
        ws.cell(row=idx, column=2, value=description)
        ws.cell(row=idx, column=6, value=severity)
        ws.cell(row=idx, column=7, value=linked_tab)

    ws["D11"] = "=COUNTBLANK('Config|Control'!$B$2:$B$9)"
    ws["E11"] = 0
    ws["C11"] = '=IF(D11=E11,"PASS","FAIL")'

    ws["D12"] = '=COUNTIF(\'_Sys|Trial_Balance\'!$I$2:$I$2000,"MISSING")+COUNTIF(\'_Sys|Trial_Balance\'!$L$2:$L$2000,"MISSING")'
    ws["E12"] = 0
    ws["C12"] = '=IF(D12=E12,"PASS","WARN")'

    ws["D13"] = '=COUNTIF(\'_Sys|Tab_Manifest\'!$A$2:$A$200,"* *")'
    ws["E13"] = 0
    ws["C13"] = '=IF(D13=E13,"PASS","FAIL")'

    ws["D14"] = (
        '=SUM(--(\'Financials|Income_Statement\'!N10<>"Comments"),--(\'Financials|Balance_Sheet\'!N10<>"Comments"),'
        '--(\'QofE|Summary\'!N10<>"Comments"),--(\'QofE|Detail\'!N10<>"Comments"),--(\'Net_Debt|Summary\'!N10<>"Comments"),'
        '--(\'Personnel|Payroll_Costs\'!N10<>"Comments"),--(\'Personnel|Headcount_Census\'!N10<>"Comments"),'
        '--(\'Other_Analysis|Revenue_By_Cust\'!N10<>"Comments"),--(\'Other_Analysis|Top_Customers\'!N10<>"Comments"),'
        '--(\'Recons|IS\'!N10<>"Comments"),--(\'Recons|BS\'!N10<>"Comments"),--(\'Recons|Cash_Flow\'!N10<>"Comments"),'
        '--(\'Recons|Billings\'!N10<>"Comments"))'
    )
    ws["E14"] = 0
    ws["C14"] = '=IF(D14=E14,"PASS","FAIL")'

    ws["D15"] = (
        '=ABS(SUMIFS(\'_Sys|Trial_Balance\'!$K:$K,\'_Sys|Trial_Balance\'!$A:$A,\'Financials|Income_Statement\'!$B$10,'
        '\'_Sys|Trial_Balance\'!$L:$L,"IS")-\'Financials|Income_Statement\'!$B$31)'
    )
    ws["E15"] = "=ctl_tolerance_abs"
    ws["C15"] = '=IF(D15<=E15,"PASS","FAIL")'

    ws["D16"] = (
        '=ABS(SUMIFS(\'_Sys|Trial_Balance\'!$K:$K,\'_Sys|Trial_Balance\'!$A:$A,\'Financials|Balance_Sheet\'!$B$10,'
        '\'_Sys|Trial_Balance\'!$L:$L,"BS")-\'Financials|Balance_Sheet\'!$B$31)'
    )
    ws["E16"] = "=ctl_tolerance_abs"
    ws["C16"] = '=IF(D16<=E16,"PASS","FAIL")'

    for row in range(11, 17):
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.border = styles.thin_border
            if col == 3:
                cell.alignment = styles.center
            else:
                cell.alignment = styles.left

    pass_fill = PatternFill(fill_type="solid", fgColor="E2F0D9")
    warn_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    fail_fill = PatternFill(fill_type="solid", fgColor="F8CBAD")
    ws.conditional_formatting.add("C11:C100", FormulaRule(formula=['C11="PASS"'], fill=pass_fill))
    ws.conditional_formatting.add("C11:C100", FormulaRule(formula=['C11="WARN"'], fill=warn_fill))
    ws.conditional_formatting.add("C11:C100", FormulaRule(formula=['C11="FAIL"'], fill=fail_fill))

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 28
    ws.column_dimensions["H"].width = 34
    ws.freeze_panes = "A11"
    ws.protection.sheet = True


def add_audit_tab(ws, styles: StylePack, kind: str) -> None:
    """Create audit tabs for issue tracking and run logs."""

    title = "Issues Index" if kind == "issues" else "Run Log"
    apply_banner(ws, subtitle=title, title=title, styles=styles)

    if kind == "issues":
        headers = [
            "issue_id",
            "sheet",
            "cell",
            "severity",
            "message",
            "linked_check",
            "timestamp_utc",
            "status",
        ]
    else:
        headers = [
            "run_id",
            "timestamp_utc",
            "mode",
            "tabs_touched",
            "checks_pass",
            "checks_warn",
            "checks_fail",
            "operator",
        ]

    write_headers(ws, headers, row=10, styles=styles)
    for row in range(11, 2001):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = styles.thin_border
            ws.cell(row=row, column=col).protection = Protection(locked=False)

    ws.freeze_panes = "A11"
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 24
    ws.protection.sheet = True


def add_sys_config_tab(ws, styles: StylePack) -> None:
    """Store finite configuration lists used by control logic and linting."""

    headers = ["config_key", "config_value"]
    write_headers(ws, headers, row=1, styles=styles)
    entries = [
        ("supported_currency", "USD,CAD,EUR,GBP"),
        ("supported_units", "1,1000,1000000"),
        ("supported_periodicity", "Monthly,Quarterly"),
        ("supported_status", "PASS,WARN,FAIL"),
        ("template_version", "v1.0.0"),
        ("created_utc", datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")),
    ]
    for row, (key, value) in enumerate(entries, start=2):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).border = styles.thin_border
        ws.cell(row=row, column=2).border = styles.thin_border

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40
    ws.protection.sheet = True


def add_tab_manifest(ws, styles: StylePack) -> None:
    """Publish authoritative tab metadata for write scopes and anchors."""

    headers = [
        "tab_name",
        "tab_type",
        "is_required",
        "is_agent_writable",
        "banner_anchor",
        "title_anchor",
        "units_anchor",
        "table_header_anchor",
        "comments_header_cell",
        "write_scope",
    ]
    write_headers(ws, headers, row=1, styles=styles)

    ordered_tabs = VISIBLE_TABS + HIDDEN_SYS_TABS
    for row, tab_name in enumerate(ordered_tabs, start=2):
        if tab_name in SECTION_TABS:
            tab_type = "section"
            writable = "FALSE"
        elif tab_name.startswith("_Sys|"):
            tab_type = "system"
            writable = "TRUE" if tab_name in {"_Sys|Trial_Balance", "_Sys|Line_Map"} else "FALSE"
        elif tab_name == "Project|Disclaimer":
            tab_type = "disclaimer"
            writable = "FALSE"
        elif tab_name == "Config|Control":
            tab_type = "config"
            writable = "TRUE"
        elif tab_name.startswith("Checks|"):
            tab_type = "checks"
            writable = "FALSE"
        elif tab_name.startswith("Audit|"):
            tab_type = "audit"
            writable = "TRUE"
        else:
            tab_type = "output"
            writable = "TRUE"

        ws.cell(row=row, column=1, value=tab_name)
        ws.cell(row=row, column=2, value=tab_type)
        ws.cell(row=row, column=3, value="TRUE")
        ws.cell(row=row, column=4, value=writable)

        if tab_type in {"output", "checks", "audit"}:
            ws.cell(row=row, column=5, value="A1:A4")
            ws.cell(row=row, column=6, value="A7")
            ws.cell(row=row, column=7, value="A8")
            ws.cell(row=row, column=8, value="A10")
            ws.cell(row=row, column=9, value="N10" if tab_type == "output" else "")
        else:
            ws.cell(row=row, column=5, value="")
            ws.cell(row=row, column=6, value="")
            ws.cell(row=row, column=7, value="")
            ws.cell(row=row, column=8, value="")
            ws.cell(row=row, column=9, value="")

        if tab_name == "_Sys|Trial_Balance":
            scope = "A2:H2000"
        elif tab_name == "_Sys|Line_Map":
            scope = "A2:F2000"
        elif tab_type == "output":
            scope = "B11:M200,N11:N200"
        elif tab_name == "Config|Control":
            scope = "B2:B16"
        elif tab_name.startswith("Audit|"):
            scope = "A11:H2000"
        else:
            scope = ""
        ws.cell(row=row, column=10, value=scope)

    for row in range(2, 2 + len(ordered_tabs)):
        for col in range(1, 11):
            ws.cell(row=row, column=col).border = styles.thin_border

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.column_dimensions["J"].width = 28
    ws.freeze_panes = "A2"
    ws.protection.sheet = True


def protect_output_sheet(ws) -> None:
    """Protect output tab while keeping data and comments entry ranges writable."""

    ws.protection.sheet = True
    for row in range(11, 201):
        for col in range(2, 14):
            ws.cell(row=row, column=col).protection = Protection(locked=False)
        ws.cell(row=row, column=14).protection = Protection(locked=False)


def main() -> None:
    """Generate the v1 workbook and save it to disk."""

    wb = Workbook()
    wb.remove(wb.active)
    styles = build_styles()

    # Create tabs in exact visible order.
    for tab_name in VISIBLE_TABS:
        wb.create_sheet(tab_name)

    # Create hidden system tabs after visible tabs.
    for tab_name in HIDDEN_SYS_TABS:
        ws = wb.create_sheet(tab_name)
        ws.sheet_state = "hidden"

    for ws in wb.worksheets:
        if ws.title in SECTION_TABS:
            section_tab(ws, ws.title, styles)
        elif ws.title == "Project|Disclaimer":
            disclaimer_tab(ws, styles)
        elif ws.title == "Config|Control":
            add_control_tab(ws, styles)
        elif ws.title in OUTPUT_TABS:
            add_output_tab_skeleton(ws, styles)
            style_grid(ws, min_row=10, max_row=33, min_col=1, max_col=14, styles=styles)
            protect_output_sheet(ws)
        elif ws.title == "Checks|Core":
            add_checks_tab(ws, styles)
        elif ws.title == "Audit|Issues_Index":
            add_audit_tab(ws, styles, kind="issues")
        elif ws.title == "Audit|Run_Log":
            add_audit_tab(ws, styles, kind="runlog")
        elif ws.title == "_Sys|Trial_Balance":
            add_trial_balance_tab(ws, styles)
        elif ws.title == "_Sys|Line_Map":
            add_line_map_tab(ws, styles)
        elif ws.title == "_Sys|Tab_Manifest":
            add_tab_manifest(ws, styles)
        elif ws.title == "_Sys|Config":
            add_sys_config_tab(ws, styles)

    wb.properties.title = "Databook Analyst Template v1"
    wb.properties.subject = "Deterministic Databook Analyst template"
    wb.properties.creator = "Codex"
    wb.properties.description = "North-conformant v1 template with simplified TB-to-output flow"

    wb.save(OUTPUT_PATH)
    print(f"Created: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
