"""Validate Databook Analyst v1 template conformance.

This script checks tab order, naming rules, hidden system tabs, banned formulas,
external links, and key anchor/header expectations.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

from openpyxl import load_workbook


TEMPLATE_PATH = Path(__file__).resolve().parent / "databook-template-v1.xlsx"

VISIBLE_TABS = [
    "Project_Template>>",
    "Project|Disclaimer",
    "Config|Control",
    "Financials>>",
    "Financials|Income_Statement",
    "Financials|Balance_Sheet",
    "QofE>>",
    "QofE|Summary",
    "QofE|Detail",
    "Net_Debt>>",
    "Net_Debt|Summary",
    "Personnel>>",
    "Personnel|Payroll_Costs",
    "Personnel|Headcount_Census",
    "Other_Analysis>>",
    "Other_Analysis|Revenue_By_Cust",
    "Other_Analysis|Top_Customers",
    "Recons>>",
    "Recons|IS",
    "Recons|BS",
    "Recons|Cash_Flow",
    "Recons|Billings",
    "Checks>>",
    "Checks|Core",
    "Audit>>",
    "Audit|Issues_Index",
    "Audit|Run_Log",
]

HIDDEN_SYS_TABS = [
    "_Sys|Trial_Balance",
    "_Sys|Line_Map",
    "_Sys|Tab_Manifest",
    "_Sys|Config",
]

OUTPUT_TABS = [
    "Financials|Income_Statement",
    "Financials|Balance_Sheet",
    "QofE|Summary",
    "QofE|Detail",
    "Net_Debt|Summary",
    "Personnel|Payroll_Costs",
    "Personnel|Headcount_Census",
    "Other_Analysis|Revenue_By_Cust",
    "Other_Analysis|Top_Customers",
    "Recons|IS",
    "Recons|BS",
    "Recons|Cash_Flow",
    "Recons|Billings",
]

CONTENT_RE = re.compile(r"^[A-Z][A-Za-z0-9_]*\|[A-Z][A-Za-z0-9_]*$")
SECTION_RE = re.compile(r"^[A-Z][A-Za-z0-9_]*>>$")
SYS_RE = re.compile(r"^_Sys\|[A-Z][A-Za-z0-9_]*$")
UNQUOTED_PIPE_REF_RE = re.compile(r"(?<!')([A-Za-z0-9_]+\|[A-Za-z0-9_]+)!")


class CheckRunner:
    """Collect and report validation checks."""

    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warnings: list[str] = []

    def check(self, condition: bool, message: str) -> None:
        """Record an error message when a condition is false."""

        if not condition:
            self.errors.append(message)

    def warn(self, condition: bool, message: str) -> None:
        """Record a warning message when a condition is false."""

        if not condition:
            self.warnings.append(message)



def validate() -> int:
    """Run all validations and return shell-friendly exit code."""

    runner = CheckRunner()
    wb = load_workbook(TEMPLATE_PATH, data_only=False)

    expected_order = VISIBLE_TABS + HIDDEN_SYS_TABS
    runner.check(wb.sheetnames == expected_order, "Sheet order does not match expected v1 contract")

    lower_seen = set()
    for name in wb.sheetnames:
        runner.check(" " not in name, f"Tab contains spaces: {name}")
        runner.check(len(name) <= 31, f"Tab exceeds 31-char Excel limit: {name}")
        lowered = name.lower()
        runner.check(lowered not in lower_seen, f"Duplicate tab name (case-insensitive): {name}")
        lower_seen.add(lowered)

        if name in HIDDEN_SYS_TABS:
            runner.check(SYS_RE.match(name) is not None, f"System tab naming mismatch: {name}")
        elif name.endswith(">>"):
            runner.check(SECTION_RE.match(name) is not None, f"Section tab naming mismatch: {name}")
        elif "|" in name:
            runner.check(CONTENT_RE.match(name) is not None, f"Content tab naming mismatch: {name}")

    for sys_tab in HIDDEN_SYS_TABS:
        runner.check(wb[sys_tab].sheet_state == "hidden", f"System tab must be hidden: {sys_tab}")

    # Anchor checks on output tabs
    for tab_name in OUTPUT_TABS:
        ws = wb[tab_name]
        runner.check(ws["A1"].value is not None, f"Missing A1 anchor on {tab_name}")
        runner.check(ws["A7"].value is not None, f"Missing A7 anchor on {tab_name}")
        runner.check(ws["A8"].value is not None, f"Missing A8 anchor on {tab_name}")
        runner.check(ws["A10"].value == "Line_Item", f"A10 header mismatch on {tab_name}")
        runner.check(ws["N10"].value == "Comments", f"Comments header missing on {tab_name}")

    # No external workbook links
    runner.check(len(getattr(wb, "_external_links", [])) == 0, "Workbook contains external links")

    banned_formula_tokens = ["VLOOKUP(", "INDEX(", "MATCH(", "OFFSET(", "INDIRECT("]
    hard_error_tokens = ["#REF!", "#VALUE!", "#NAME?", "#DIV/0!"]
    banned_hits: list[str] = []
    err_hits: list[str] = []
    unquoted_pipe_ref_hits: list[str] = []

    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_upper = value.upper()
                    for token in banned_formula_tokens:
                        if token in formula_upper:
                            banned_hits.append(f"{ws.title}!{cell.coordinate}: {token}")
                    if UNQUOTED_PIPE_REF_RE.search(value):
                        unquoted_pipe_ref_hits.append(f"{ws.title}!{cell.coordinate}: {value}")
                    for token in hard_error_tokens:
                        if token in formula_upper:
                            err_hits.append(f"{ws.title}!{cell.coordinate}: {token}")
                elif isinstance(value, str):
                    for token in hard_error_tokens:
                        if token in value.upper():
                            err_hits.append(f"{ws.title}!{cell.coordinate}: {token}")

    runner.check(not banned_hits, "Banned formula tokens found:\n" + "\n".join(banned_hits[:25]))
    runner.check(
        not unquoted_pipe_ref_hits,
        "Unquoted pipe-sheet references found:\n" + "\n".join(unquoted_pipe_ref_hits[:25]),
    )
    runner.check(not err_hits, "Formula error tokens found:\n" + "\n".join(err_hits[:25]))

    # Named ranges required by contract.
    names = {name for name, _ in wb.defined_names.items()}
    required_names = {
        "ctl_project_name",
        "ctl_report_date",
        "ctl_currency",
        "ctl_units_scale",
        "ctl_sign_convention",
        "ctl_tolerance_abs",
        "ctl_tolerance_pct",
        "ctl_preparer",
        "ctl_std_periodicity",
        "ctl_std_period_start",
        "ctl_std_period_end",
        "ctl_std_period_count",
    }
    missing = sorted(required_names - names)
    runner.check(not missing, "Missing required defined names: " + ", ".join(missing))

    # Soft warning: ensure the long Revenue tab was intentionally shortened.
    runner.warn(
        "Other_Analysis|Revenue_By_Customer_Group" not in wb.sheetnames,
        "Long Revenue_By_Customer_Group tab unexpectedly present",
    )

    if runner.warnings:
        print("WARNINGS:")
        for warning in runner.warnings:
            print(f"- {warning}")

    if runner.errors:
        print("FAIL")
        for error in runner.errors:
            print(f"- {error}")
        return 1

    print("PASS")
    print(f"Validated workbook: {TEMPLATE_PATH}")
    print(f"Sheet count: {len(wb.sheetnames)}")
    return 0


if __name__ == "__main__":
    sys.exit(validate())
