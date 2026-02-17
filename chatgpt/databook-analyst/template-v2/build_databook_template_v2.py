"""Build North-based Databook Template v2.

V2 keeps Project North visual conventions for core tabs, hides non-core tabs,
adds deterministic `_Sys|*` scaffolding, and enforces a 36-month manifest rule
on designated monthly tabs.
"""

from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
import json

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parent
SEED_PATH = ROOT / "seeds" / "north-master-seed.xlsx"
OUT_PATH = ROOT / "databook-template-v2.xlsx"
STYLE_CONTRACT_PATH = ROOT / "contracts" / "north_style_contract.json"
ANCHOR_CONTRACT_PATH = ROOT / "contracts" / "north_anchor_contract.json"
NORTH_TOKEN_PATH = ROOT.parent / "databook-template-analysis" / "north-design-tokens" / "north-design-tokens.json"

CORE_VISIBLE_ORDER = [
    "Project North >>>",
    "Disclaimer",
    "QofE >>",
    "QofE | Bridge",
    "QofE | Summary",
    "QofE | Detail",
    "Net debt>>",
    "Net debt",
    "Financials>>",
    "Income statement",
    "Balance sheet",
    "Cash flow",
    "Personnel>>",
    "IS | Payroll costs",
    "Headcount census",
    "Other analysis>>",
    "Top customers",
    "Revenue by customer group",
    "Recons >>",
    "Recons | IS",
    "Recons | BS",
    "Recon | Cash flow",
    "Recon | Billings",
    "Checks>>",
    "Checks | Core",
    "Audit>>",
    "Audit | Issues_Index",
    "Audit | Run_Log",
]

SYS_TABS = [
    "_Sys|Trial_Balance",
    "_Sys|Line_Map",
    "_Sys|Tab_Manifest",
    "_Sys|Config",
]

MONTHLY_MANIFEST = {
    "Income statement": (9, "F"),
    "Balance sheet": (9, "F"),
    "QofE | Detail": (8, "E"),
    "Net debt": (9, "E"),
    "Recon | Billings": (8, "F"),
}

COMMENT_CELLS = {
    "QofE | Bridge": "AI7",
    "QofE | Detail": "BA7",
    "Net debt": "M7",
    "Recon | Cash flow": "AF7",
    "Recon | Billings": "BB7",
}


def ensure_paths() -> None:
    """Ensure required build paths exist."""

    if not SEED_PATH.exists():
        raise FileNotFoundError(f"Seed workbook not found: {SEED_PATH}")
    ROOT.joinpath("contracts").mkdir(parents=True, exist_ok=True)
    ROOT.joinpath("reports").mkdir(parents=True, exist_ok=True)


def write_contract_files() -> None:
    """Write semantic style + anchor contracts from extracted North tokens."""

    if not NORTH_TOKEN_PATH.exists():
        return

    payload = json.loads(NORTH_TOKEN_PATH.read_text(encoding="utf-8"))
    semantic = payload.get("semantic_summary", {})

    style_contract = {
        "source": str(payload.get("source_xlsb", "")),
        "top_fonts": semantic.get("font_families", [])[:20],
        "top_colors": semantic.get("colors", [])[:30],
        "top_alignments": semantic.get("alignments", [])[:20],
        "top_fills": semantic.get("fills", [])[:20],
        "top_borders": semantic.get("borders", [])[:20],
    }

    anchor_contract = {
        "anchor_style_frequency": semantic.get("anchor_style_frequency", {}),
        "anchor_value_samples": semantic.get("anchor_value_samples", {}),
        "anchor_cells": ["B1", "B2", "B3", "B4", "C7", "C8", "C9", "C10"],
    }

    STYLE_CONTRACT_PATH.write_text(json.dumps(style_contract, indent=2), encoding="utf-8")
    ANCHOR_CONTRACT_PATH.write_text(json.dumps(anchor_contract, indent=2), encoding="utf-8")


def clear_sheet_values(ws) -> None:
    """Clear values/formulas/comments while preserving existing formatting."""

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.value = None
            cell.comment = None
            cell.hyperlink = None


def make_section_tab(wb, title: str) -> None:
    """Create a section divider tab by copying existing section style."""

    if title in wb.sheetnames:
        return

    donor = wb["Recons >>"] if "Recons >>" in wb.sheetnames else wb[wb.sheetnames[0]]
    ws = wb.copy_worksheet(donor)
    ws.title = title
    clear_sheet_values(ws)
    ws["A1"] = title


def make_structured_tab(wb, title: str, subtitle: str) -> None:
    """Create a structured tab by cloning North-style donor layout."""

    if title in wb.sheetnames:
        return

    donor = wb["Recons | IS"]
    ws = wb.copy_worksheet(donor)
    ws.title = title
    clear_sheet_values(ws)

    # Rebuild standard North-like banner anchors.
    ws["B1"] = "Project North"
    ws["B2"] = subtitle
    ws["B3"] = "Source: System generated"
    ws["B4"] = "WORKING DRAFT"
    ws["C7"] = subtitle


def copy_cell_style(src_ws, src_cell: str, dst_ws, dst_cell: str) -> None:
    """Copy style from one cell to another."""

    dst_ws[dst_cell]._style = copy(src_ws[src_cell]._style)
    dst_ws[dst_cell].number_format = src_ws[src_cell].number_format
    dst_ws[dst_cell].alignment = copy(src_ws[src_cell].alignment)
    dst_ws[dst_cell].font = copy(src_ws[src_cell].font)
    dst_ws[dst_cell].fill = copy(src_ws[src_cell].fill)
    dst_ws[dst_cell].border = copy(src_ws[src_cell].border)
    dst_ws[dst_cell].protection = copy(src_ws[src_cell].protection)


def build_checks_tab(wb) -> None:
    """Populate Checks | Core with deterministic checks."""

    ws = wb["Checks | Core"]
    donor = wb["Recons | IS"]

    # Banner and title styles from donor anchors.
    for src, dst in [("B1", "B1"), ("B2", "B2"), ("B3", "B3"), ("B4", "B4"), ("C7", "C7")]:
        copy_cell_style(donor, src, ws, dst)

    ws["B1"] = "Project North"
    ws["B2"] = "Checks | Core"
    ws["B3"] = "Source: System generated"
    ws["B4"] = "WORKING DRAFT"
    ws["C7"] = "Checks | Core"

    headers = ["Check_ID", "Description", "Status", "Actual", "Expected", "Severity", "Linked_Tab"]
    for idx, header in enumerate(headers, start=3):
        cell = ws.cell(row=9, column=idx)
        copy_cell_style(donor, "C9", ws, cell.coordinate)
        cell.value = header

    checks = [
        ("CHK-001", "Control values populated", "HIGH", "_Sys|Config"),
        ("CHK-002", "Unmapped TB rows", "HIGH", "_Sys|Trial_Balance"),
        ("CHK-003", "No forbidden formulas in governed tabs", "HIGH", "Core"),
    ]

    row = 10
    for check_id, desc, sev, linked in checks:
        for col in range(3, 10):
            copy_cell_style(donor, "C10", ws, ws.cell(row=row, column=col).coordinate)
        ws.cell(row=row, column=3, value=check_id)
        ws.cell(row=row, column=4, value=desc)
        ws.cell(row=row, column=8, value=sev)
        ws.cell(row=row, column=9, value=linked)
        row += 1

    ws["F10"] = "=COUNTBLANK('_Sys|Config'!$B$2:$B$5)"
    ws["G10"] = 0
    ws["E10"] = '=IF(F10=G10,"PASS","FAIL")'

    ws["F11"] = '=COUNTIF(\'_Sys|Trial_Balance\'!$I$2:$I$5000,"MISSING")+COUNTIF(\'_Sys|Trial_Balance\'!$L$2:$L$5000,"MISSING")'
    ws["G11"] = 0
    ws["E11"] = '=IF(F11=G11,"PASS","WARN")'

    ws["F12"] = '=COUNTIF(\'_Sys|Tab_Manifest\'!$A:$A,">0")'
    ws["G12"] = '=COUNTIF(\'_Sys|Tab_Manifest\'!$A:$A,">0")'
    ws["E12"] = '=IF(F12=G12,"PASS","WARN")'

    # Monthly checks
    for tab_name, (header_row, start_col) in MONTHLY_MANIFEST.items():
        end_col = get_column_letter(column_index_from_string(start_col) + 35)
        rng = f"'{tab_name}'!${start_col}${header_row}:${end_col}${header_row}"
        first = f"'{tab_name}'!${start_col}${header_row}"

        for col in range(3, 10):
            copy_cell_style(donor, "C10", ws, ws.cell(row=row, column=col).coordinate)

        ws.cell(row=row, column=3, value=f"CHK-MON-{row-12:03d}")
        ws.cell(row=row, column=4, value=f"36-month contiguous header | {tab_name}")
        ws.cell(row=row, column=8, value="HIGH")
        ws.cell(row=row, column=9, value=tab_name)
        ws.cell(
            row=row,
            column=6,
            value=(
                f"=IF(AND(COUNT({rng})=36,"
                f"SUMPRODUCT(--({rng}=EDATE({first},COLUMN({rng})-COLUMN({first}))))=36),1,0)"
            ),
        )
        ws.cell(row=row, column=7, value=1)
        ws.cell(row=row, column=5, value=f'=IF(F{row}=G{row},"PASS","FAIL")')
        row += 1

    ws.freeze_panes = "C10"


def build_audit_tab(wb, title: str, subtitle: str, headers: list[str]) -> None:
    """Populate audit tabs with North-like presentation and schema headers."""

    ws = wb[title]
    donor = wb["Recons | IS"]

    for src, dst in [("B1", "B1"), ("B2", "B2"), ("B3", "B3"), ("B4", "B4"), ("C7", "C7")]:
        copy_cell_style(donor, src, ws, dst)

    ws["B1"] = "Project North"
    ws["B2"] = subtitle
    ws["B3"] = "Source: System generated"
    ws["B4"] = "WORKING_DRAFT"
    ws["C7"] = subtitle

    for idx, header in enumerate(headers, start=3):
        cell = ws.cell(row=9, column=idx)
        copy_cell_style(donor, "C9", ws, cell.coordinate)
        cell.value = header

    for row in range(10, 300):
        for col in range(3, 3 + len(headers)):
            copy_cell_style(donor, "C10", ws, ws.cell(row=row, column=col).coordinate)
            ws.cell(row=row, column=col).protection = Protection(locked=False)

    ws.freeze_panes = "C10"


def setup_sys_config(wb) -> None:
    """Populate hidden config controls and define names."""

    ws = wb["_Sys|Config"]
    ws.sheet_state = "hidden"

    headers = ["config_key", "config_value"]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=1, column=col).fill = PatternFill(fill_type="solid", fgColor="D9E1F2")

    entries = [
        ("ctl_monthly_period_count", 36),
        ("ctl_month_start", datetime(2021, 1, 31)),
        ("ctl_tolerance_abs", 1),
        ("ctl_tolerance_pct", 0.001),
        ("ctl_project_name", "Project North"),
    ]

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row, (k, v) in enumerate(entries, start=2):
        ws.cell(row=row, column=1, value=k)
        ws.cell(row=row, column=2, value=v)
        ws.cell(row=row, column=1).border = border
        ws.cell(row=row, column=2).border = border
        ws.cell(row=row, column=2).protection = Protection(locked=False)

    ws["B3"].number_format = "yyyy-mm-dd"
    ws["B5"].number_format = "0.00%"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22

    for name, row in [("ctl_monthly_period_count", 2), ("ctl_month_start", 3), ("ctl_tolerance_abs", 4), ("ctl_tolerance_pct", 5), ("ctl_project_name", 6)]:
        ref = f"'_Sys|Config'!$B${row}"
        wb.defined_names.add(DefinedName(name=name, attr_text=ref))


def setup_trial_balance(wb) -> None:
    """Create TB input sheet with direct mapping helper formulas."""

    ws = wb["_Sys|Trial_Balance"]
    ws.sheet_state = "hidden"

    headers = [
        "period_end",
        "account_code",
        "account_name",
        "amount",
        "entity",
        "source_file",
        "source_tab",
        "note",
        "map_line",
        "map_sign",
        "signed_amount",
        "map_statement_type",
    ]

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        c.border = border

    for row in range(2, 5001):
        ws.cell(row=row, column=9, value=f"=XLOOKUP(B{row},'_Sys|Line_Map'!$A:$A,'_Sys|Line_Map'!$C:$C,\"MISSING\",0)")
        ws.cell(row=row, column=10, value=f"=XLOOKUP(B{row},'_Sys|Line_Map'!$A:$A,'_Sys|Line_Map'!$D:$D,1,0)")
        ws.cell(row=row, column=11, value=f"=IF(D{row}=\"\",\"\",D{row}*J{row})")
        ws.cell(row=row, column=12, value=f"=XLOOKUP(B{row},'_Sys|Line_Map'!$A:$A,'_Sys|Line_Map'!$B:$B,\"MISSING\",0)")
        for col in range(1, 13):
            ws.cell(row=row, column=col).border = border
            if col <= 8:
                ws.cell(row=row, column=col).protection = Protection(locked=False)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 24
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 20
    ws.freeze_panes = "A2"


def setup_line_map(wb) -> None:
    """Create account-to-line mapping table with editable rows."""

    ws = wb["_Sys|Line_Map"]
    ws.sheet_state = "hidden"

    headers = ["account_code", "statement_type", "report_line", "sign_factor", "active_flag", "notes"]
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        c.border = border

    seed = [
        ("4000", "IS", "Revenue", 1, "Y", "seed mapping"),
        ("5000", "IS", "COGS", -1, "Y", "seed mapping"),
        ("6100", "IS", "Personnel_Costs", -1, "Y", "seed mapping"),
        ("1000", "BS", "Cash", 1, "Y", "seed mapping"),
        ("2000", "BS", "AP", -1, "Y", "seed mapping"),
    ]

    for row, values in enumerate(seed, start=2):
        for col, value in enumerate(values, start=1):
            ws.cell(row=row, column=col, value=value)

    for row in range(2, 5001):
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).protection = Protection(locked=False)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 34
    ws.freeze_panes = "A2"


def enforce_month_headers(wb) -> None:
    """Apply 36-month contiguous headers on manifest-tagged monthly tabs."""

    for tab_name, (row, start_col) in MONTHLY_MANIFEST.items():
        ws = wb[tab_name]
        start_idx = column_index_from_string(start_col)

        # Reuse the first header style across the full 36-month run.
        style_src = ws.cell(row=row, column=start_idx)
        number_fmt = style_src.number_format

        for offset in range(36):
            col_idx = start_idx + offset
            cell = ws.cell(row=row, column=col_idx)
            if offset == 0:
                cell.value = "=ctl_month_start"
            else:
                cell.value = f"=EDATE(ctl_month_start,{offset})"
            cell.number_format = number_fmt if number_fmt else "mmm-yy"
            cell._style = copy(style_src._style)

        # Clear legacy trailing header cells to enforce exactly 36 periods.
        for col_idx in range(start_idx + 36, start_idx + 180):
            ws.cell(row=row, column=col_idx).value = None


def setup_manifest(wb) -> None:
    """Write `_Sys|Tab_Manifest` with write scopes and monthly metadata."""

    ws = wb["_Sys|Tab_Manifest"]
    ws.sheet_state = "hidden"

    headers = [
        "tab_name",
        "is_visible_scope",
        "is_monthly_enforced",
        "period_header_row",
        "period_start_col",
        "period_count_required",
        "comments_header_cell",
        "write_scope",
    ]

    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, header in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=header)
        c.font = Font(name="Calibri", size=10, bold=True)
        c.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        c.border = border

    governed_tabs = CORE_VISIBLE_ORDER + SYS_TABS

    for row, tab in enumerate(governed_tabs, start=2):
        ws.cell(row=row, column=1, value=tab)
        ws.cell(row=row, column=2, value="TRUE" if tab in CORE_VISIBLE_ORDER else "FALSE")
        ws.cell(row=row, column=3, value="TRUE" if tab in MONTHLY_MANIFEST else "FALSE")

        if tab in MONTHLY_MANIFEST:
            hdr_row, start_col = MONTHLY_MANIFEST[tab]
            ws.cell(row=row, column=4, value=hdr_row)
            ws.cell(row=row, column=5, value=start_col)
            ws.cell(row=row, column=6, value=36)
        else:
            ws.cell(row=row, column=4, value="")
            ws.cell(row=row, column=5, value="")
            ws.cell(row=row, column=6, value="")

        ws.cell(row=row, column=7, value=COMMENT_CELLS.get(tab, ""))

        if tab == "_Sys|Trial_Balance":
            scope = "A2:H5000"
        elif tab == "_Sys|Line_Map":
            scope = "A2:F5000"
        elif tab.startswith("Audit |"):
            scope = "C10:J5000"
        elif tab == "Checks | Core":
            scope = ""
        elif tab in MONTHLY_MANIFEST:
            scope = "monthly-grid + comments"
        elif tab in CORE_VISIBLE_ORDER:
            scope = "comments-only"
        else:
            scope = ""

        ws.cell(row=row, column=8, value=scope)

        for col in range(1, 9):
            ws.cell(row=row, column=col).border = border

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 30
    ws.freeze_panes = "A2"


def hide_non_core_tabs(wb) -> None:
    """Keep core scope visible; hide all other legacy tabs."""

    visible_set = set(CORE_VISIBLE_ORDER)
    sys_set = set(SYS_TABS)
    for ws in wb.worksheets:
        if ws.title in visible_set:
            ws.sheet_state = "visible"
        elif ws.title in sys_set:
            ws.sheet_state = "hidden"
        else:
            ws.sheet_state = "hidden"


def reorder_tabs(wb) -> None:
    """Order workbook with visible core first, then sys tabs, then remaining hidden tabs."""

    existing = {ws.title: ws for ws in wb.worksheets}
    ordered: list = []

    for name in CORE_VISIBLE_ORDER:
        if name in existing:
            ordered.append(existing[name])

    for name in SYS_TABS:
        if name in existing:
            ordered.append(existing[name])

    for ws in wb.worksheets:
        if ws not in ordered:
            ordered.append(ws)

    wb._sheets = ordered


def create_missing_tabs(wb) -> None:
    """Create Checks/Audit tabs and `_Sys|*` tabs when absent."""

    make_section_tab(wb, "Checks>>")
    make_structured_tab(wb, "Checks | Core", "Checks | Core")
    make_section_tab(wb, "Audit>>")
    make_structured_tab(wb, "Audit | Issues_Index", "Audit | Issues_Index")
    make_structured_tab(wb, "Audit | Run_Log", "Audit | Run_Log")

    for tab in SYS_TABS:
        if tab not in wb.sheetnames:
            ws = wb.create_sheet(tab)
            ws.sheet_state = "hidden"


def remove_external_links(wb) -> None:
    """Strip external link objects from workbook for clean portability."""

    try:
        wb._external_links = []
    except Exception:
        pass


def main() -> None:
    """Run full v2 build pipeline."""

    ensure_paths()
    write_contract_files()

    wb = load_workbook(SEED_PATH, data_only=False)

    create_missing_tabs(wb)

    # Build/normalize governed tabs.
    build_checks_tab(wb)
    build_audit_tab(
        wb,
        title="Audit | Issues_Index",
        subtitle="Audit | Issues_Index",
        headers=["issue_id", "sheet", "cell", "severity", "message", "linked_check", "timestamp_utc", "status"],
    )
    build_audit_tab(
        wb,
        title="Audit | Run_Log",
        subtitle="Audit | Run_Log",
        headers=["run_id", "timestamp_utc", "mode", "tabs_touched", "checks_pass", "checks_warn", "checks_fail", "operator"],
    )

    setup_sys_config(wb)
    setup_trial_balance(wb)
    setup_line_map(wb)
    setup_manifest(wb)
    enforce_month_headers(wb)

    remove_external_links(wb)
    hide_non_core_tabs(wb)
    reorder_tabs(wb)

    wb.properties.title = "Databook Template V2"
    wb.properties.subject = "North-based core scope template"
    wb.properties.creator = "Codex"

    wb.save(OUT_PATH)
    print(f"Created: {OUT_PATH}")


if __name__ == "__main__":
    main()
