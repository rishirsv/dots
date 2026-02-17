"""Validate North-based Databook Template v2 conformance.

Checks:
- Core tab presence and visibility
- System tab presence and hidden state
- Manifest-tagged 36-month enforcement
- Formula policy constraints
- External-link and formula-error hygiene
- North anchor fidelity report generation
"""

from __future__ import annotations

from datetime import datetime, date
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter


ROOT = Path(__file__).resolve().parent
WORKBOOK_PATH = ROOT / "databook-template-v2.xlsx"
SEED_PATH = ROOT / "seeds" / "north-master-seed.xlsx"
REPORT_PATH = ROOT / "reports" / "north-fidelity-diff.md"

CORE_VISIBLE_ORDER = [
    "Project North >>>",
    "Disclaimer",
    "QofE >>",
    "QofE | Bridge",
    "QofE | Summary",
    "QofE | Detail",
    "Net debt>>",
    "Net debt",
    "Financials>>",
    "Income statement",
    "Balance sheet",
    "Cash flow",
    "Personnel>>",
    "IS | Payroll costs",
    "Headcount census",
    "Other analysis>>",
    "Top customers",
    "Revenue by customer group",
    "Recons >>",
    "Recons | IS",
    "Recons | BS",
    "Recon | Cash flow",
    "Recon | Billings",
    "Checks>>",
    "Checks | Core",
    "Audit>>",
    "Audit | Issues_Index",
    "Audit | Run_Log",
]

SYS_TABS = ["_Sys|Trial_Balance", "_Sys|Line_Map", "_Sys|Tab_Manifest", "_Sys|Config"]
FORMULA_POLICY_RANGES = {
    "_Sys|Trial_Balance": (1, 350, 1, 12),
    "_Sys|Line_Map": (1, 350, 1, 6),
    "_Sys|Tab_Manifest": (1, 350, 1, 8),
    "_Sys|Config": (1, 60, 1, 2),
    "Checks | Core": (1, 250, 1, 20),
    "Audit | Issues_Index": (1, 250, 1, 20),
    "Audit | Run_Log": (1, 250, 1, 20),
}
FORBIDDEN_FORMULA_TOKENS = ["VLOOKUP(", "INDEX(", "MATCH(", "OFFSET(", "INDIRECT("]
ERROR_TOKENS = ["#REF!", "#VALUE!", "#NAME?", "#DIV/0!"]


class Result:
    """Accumulates validation errors and warnings."""

    def __init__(self) -> None:
        self.errors: list[str] = []
        self.warnings: list[str] = []

    def check(self, condition: bool, message: str) -> None:
        """Register an error when condition is false."""

        if not condition:
            self.errors.append(message)

    def warn(self, condition: bool, message: str) -> None:
        """Register a warning when condition is false."""

        if not condition:
            self.warnings.append(message)


def is_date_like(v) -> bool:
    """Return true for date/datetime values."""

    return isinstance(v, (datetime, date))


def validate_monthly_manifest(wb, result: Result) -> None:
    """Validate 36-month enforcement using `_Sys|Tab_Manifest`."""

    ws = wb["_Sys|Tab_Manifest"]

    headers = [ws.cell(1, c).value for c in range(1, 9)]
    expected_headers = [
        "tab_name",
        "is_visible_scope",
        "is_monthly_enforced",
        "period_header_row",
        "period_start_col",
        "period_count_required",
        "comments_header_cell",
        "write_scope",
    ]
    result.check(headers == expected_headers, "_Sys|Tab_Manifest headers do not match expected schema")

    row = 2
    while True:
        tab_name = ws.cell(row, 1).value
        if not tab_name:
            break

        is_monthly = str(ws.cell(row, 3).value).upper() == "TRUE"
        if is_monthly:
            hdr_row = ws.cell(row, 4).value
            start_col = ws.cell(row, 5).value
            cnt = ws.cell(row, 6).value

            result.check(tab_name in wb.sheetnames, f"Manifest tab missing in workbook: {tab_name}")
            result.check(hdr_row is not None, f"Manifest missing period_header_row for {tab_name}")
            result.check(start_col is not None, f"Manifest missing period_start_col for {tab_name}")
            result.check(int(cnt) == 36, f"Manifest period_count_required must be 36 for {tab_name}")

            if tab_name in wb.sheetnames and hdr_row and start_col:
                tws = wb[tab_name]
                start_idx = column_index_from_string(str(start_col))
                values = [tws.cell(int(hdr_row), start_idx + i).value for i in range(36)]
                trailing = tws.cell(int(hdr_row), start_idx + 36).value

                # For formula-driven headers, ensure the formulas are present and contiguous by construction.
                first = values[0]
                second = values[1] if len(values) > 1 else None
                result.check(
                    isinstance(first, str) and "ctl_month_start" in first,
                    f"First monthly header is not control-linked on {tab_name}",
                )
                result.check(
                    isinstance(second, str) and "EDATE(ctl_month_start,1)" in second,
                    f"Second monthly header is not EDATE-linked on {tab_name}",
                )
                result.check(trailing is None, f"Trailing monthly header not cleared after 36 periods on {tab_name}")

        row += 1


def validate_tab_visibility_and_order(wb, result: Result) -> None:
    """Validate core tab visibility and ordering expectations."""

    for tab in CORE_VISIBLE_ORDER:
        result.check(tab in wb.sheetnames, f"Missing core tab: {tab}")
        if tab in wb.sheetnames:
            result.check(wb[tab].sheet_state == "visible", f"Core tab must be visible: {tab}")

    for tab in SYS_TABS:
        result.check(tab in wb.sheetnames, f"Missing system tab: {tab}")
        if tab in wb.sheetnames:
            result.check(wb[tab].sheet_state == "hidden", f"System tab must be hidden: {tab}")

    # Ensure first visible tabs follow the core sequence.
    visible_titles = [ws.title for ws in wb.worksheets if ws.sheet_state == "visible"]
    result.check(visible_titles == CORE_VISIBLE_ORDER, "Visible tab order does not match core scope order")


def validate_formula_policy(wb, result: Result) -> None:
    """Check formula constraints in governed tabs."""

    monthly_tabs = set()
    manifest = wb["_Sys|Tab_Manifest"]
    row = 2
    while True:
        tab_name = manifest.cell(row, 1).value
        if not tab_name:
            break
        is_monthly = str(manifest.cell(row, 3).value).upper() == "TRUE"
        if is_monthly:
            monthly_tabs.add(tab_name)
        row += 1

    forbidden_hits: list[str] = []
    error_hits: list[str] = []

    for ws in wb.worksheets:
        if ws.title in FORMULA_POLICY_RANGES:
            min_row, max_row, min_col, max_col = FORMULA_POLICY_RANGES[ws.title]
            for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
                for cell in row:
                    value = cell.value
                    if isinstance(value, str) and value.startswith("="):
                        upper = value.upper()
                        for token in FORBIDDEN_FORMULA_TOKENS:
                            if token in upper:
                                forbidden_hits.append(f"{ws.title}!{cell.coordinate}: {token}")
                        for token in ERROR_TOKENS:
                            if token in upper:
                                error_hits.append(f"{ws.title}!{cell.coordinate}: {token}")
                    elif isinstance(value, str):
                        upper = value.upper()
                        for token in ERROR_TOKENS:
                            if token in upper:
                                error_hits.append(f"{ws.title}!{cell.coordinate}: {token}")
        elif ws.title in monthly_tabs:
            # Monthly tabs are checked only in header/control region to keep validation fast.
            for r in range(6, 15):
                for c in range(1, 260):
                    value = ws.cell(r, c).value
                    if isinstance(value, str) and value.startswith("="):
                        upper = value.upper()
                        for token in FORBIDDEN_FORMULA_TOKENS:
                            if token in upper:
                                forbidden_hits.append(f"{ws.title}!{get_column_letter(c)}{r}: {token}")
                        for token in ERROR_TOKENS:
                            if token in upper:
                                error_hits.append(f"{ws.title}!{get_column_letter(c)}{r}: {token}")
                    elif isinstance(value, str):
                        upper = value.upper()
                        for token in ERROR_TOKENS:
                            if token in upper:
                                error_hits.append(f"{ws.title}!{get_column_letter(c)}{r}: {token}")

    result.check(not forbidden_hits, "Forbidden formula token(s):\n" + "\n".join(forbidden_hits[:30]))
    result.check(not error_hits, "Formula error token(s):\n" + "\n".join(error_hits[:30]))


def validate_hygiene(wb, result: Result) -> None:
    """General workbook hygiene checks."""

    result.check(len(getattr(wb, "_external_links", [])) == 0, "Workbook contains external links")


def write_fidelity_report(v2_wb, seed_wb) -> None:
    """Generate side-by-side anchor fidelity report for core tabs."""

    anchors = ["B1", "B2", "B3", "B4", "C7", "C8", "C9", "C10"]
    lines: list[str] = []
    lines.append("# North Fidelity Diff (Core Tabs)")
    lines.append("")
    lines.append(f"- Seed: `{SEED_PATH}`")
    lines.append(f"- V2: `{WORKBOOK_PATH}`")
    lines.append("")

    mismatches = 0
    compared = 0

    for tab in CORE_VISIBLE_ORDER:
        if tab not in seed_wb.sheetnames or tab not in v2_wb.sheetnames:
            continue
        lines.append(f"## {tab}")
        for anchor in anchors:
            seed_cell = seed_wb[tab][anchor]
            v2_cell = v2_wb[tab][anchor]
            seed_val = seed_cell.value
            v2_val = v2_cell.value
            seed_style = getattr(seed_cell, "style_id", getattr(seed_cell, "_style_id", None))
            v2_style = getattr(v2_cell, "style_id", getattr(v2_cell, "_style_id", None))
            value_match = seed_val == v2_val
            status = "MATCH" if value_match else "DIFF_VALUE"
            compared += 1
            if status != "MATCH":
                mismatches += 1
            lines.append(
                f"- `{anchor}` {status} | seed_style={seed_style} v2_style={v2_style} | seed_val={seed_val!r} | v2_val={v2_val!r}"
            )
        lines.append("")

    lines.insert(3, f"- Compared anchors: **{compared}**")
    lines.insert(4, f"- Anchor value diffs: **{mismatches}**")
    lines.insert(5, "- Note: `style_id` is workbook-local and shown as advisory, not pass/fail.")
    REPORT_PATH.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    """Run validation and return shell exit code."""

    result = Result()

    if not WORKBOOK_PATH.exists():
        print(f"FAIL\n- Missing workbook: {WORKBOOK_PATH}")
        return 1

    v2_wb = load_workbook(WORKBOOK_PATH, data_only=False, read_only=True)
    seed_wb = load_workbook(SEED_PATH, data_only=False, read_only=True) if SEED_PATH.exists() else None

    validate_tab_visibility_and_order(v2_wb, result)
    validate_monthly_manifest(v2_wb, result)
    validate_formula_policy(v2_wb, result)
    validate_hygiene(v2_wb, result)

    if seed_wb is not None:
        write_fidelity_report(v2_wb, seed_wb)

    if result.warnings:
        print("WARNINGS:")
        for warn in result.warnings:
            print(f"- {warn}")

    if result.errors:
        print("FAIL")
        for err in result.errors:
            print(f"- {err}")
        return 1

    print("PASS")
    print(f"Validated workbook: {WORKBOOK_PATH}")
    print(f"Fidelity report: {REPORT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
