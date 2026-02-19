#!/usr/bin/env python3
"""Normalize trial balance exports into canonical long-form CSV."""

from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path

from openpyxl import load_workbook

CANONICAL_FIELDS = [
    "period",
    "entity",
    "account_number",
    "account_name",
    "amount_raw",
    "units_scale_applied",
    "amount",
    "source_column",
    "source_row",
    "is_derived",
    "notes",
]

LONG_AMOUNT_FIELDS = ("amount", "balance", "ending_balance", "trial_balance")
LONG_ACCOUNT_NUMBER_FIELDS = (
    "account",
    "account_number",
    "account_no",
    "gl_account",
    "gl_code",
    "acct",
    "acct_no",
)
LONG_ACCOUNT_NAME_FIELDS = ("account_name", "description", "account_description", "gl_name", "name")
PERIOD_FIELDS = ("period", "period_end", "as_of", "date", "month_end")
YEAR_FIELDS = ("year", "fiscal_year")
MONTH_FIELDS = ("month", "fiscal_month")
DERIVED_PREFIXES = ("total_", "subtotal_", "variance_", "pct_", "percent_")
DERIVED_EXACT = {
    "gross_profit",
    "net_income",
    "ebitda",
    "operating_income",
    "operating_profit",
    "total_assets",
    "total_liabilities",
    "total_equity",
}


def _normalize_header(name: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]+", "_", (name or "").strip().lower())
    return re.sub(r"_+", "_", s).strip("_")


def _normalize_key(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", (value or "").strip().lower())


def _to_amount(raw: object) -> float:
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip()
    if not s:
        return 0.0
    s = s.replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = f"-{s[1:-1]}"
    return float(s)


def _read_csv(path: Path) -> tuple[list[str], list[dict[str, object]]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            raise SystemExit(f"Input CSV has no header row: {path}")
        fields = [_normalize_header(n) for n in reader.fieldnames]
        rows: list[dict[str, object]] = []
        for row in reader:
            rows.append({_normalize_header(k): v for k, v in row.items()})
    return fields, rows


def _read_xlsx(path: Path, sheet: str | None) -> tuple[list[str], list[dict[str, object]]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    headers = [_normalize_header(str(ws.cell(1, c).value or "")) for c in range(1, ws.max_column + 1)]
    if not any(headers):
        raise SystemExit(f"Input XLSX first row has no headers: {path}")
    rows: list[dict[str, object]] = []
    for r in range(2, ws.max_row + 1):
        out: dict[str, object] = {}
        empty = True
        for c, h in enumerate(headers, start=1):
            v = ws.cell(r, c).value
            if v not in (None, ""):
                empty = False
            out[h] = v
        if not empty:
            rows.append(out)
    return headers, rows


def _is_derived_col(name: str) -> bool:
    return name in DERIVED_EXACT or name.startswith(DERIVED_PREFIXES)


def _account_columns(headers: list[str]) -> list[str]:
    return [h for h in headers if h.startswith("acct_") or h.startswith("account_")]


def _detect_layout(headers: list[str]) -> str:
    account_cols = _account_columns(headers)
    amount_like = set(headers) & set(LONG_AMOUNT_FIELDS)
    period_candidates = set(PERIOD_FIELDS) | set(YEAR_FIELDS) | set(MONTH_FIELDS)
    has_period_axis = bool(set(headers) & period_candidates)
    if len(account_cols) >= 2 and not amount_like and has_period_axis:
        return "wide"
    return "long"


def _derive_period(row: dict[str, object]) -> str:
    for col in PERIOD_FIELDS:
        if col in row and row[col] not in (None, ""):
            return str(row[col]).strip()
    year_val = next((row[c] for c in YEAR_FIELDS if c in row and row[c] not in (None, "")), None)
    month_val = next((row[c] for c in MONTH_FIELDS if c in row and row[c] not in (None, "")), None)
    if year_val is not None and month_val is not None:
        try:
            return f"{int(year_val):04d}-{int(month_val):02d}"
        except Exception:
            return f"{year_val}-{month_val}"
    return ""


def _scaled(raw_amount: float, scale: float) -> float:
    return raw_amount * scale


def _wide_to_long(
    rows: list[dict[str, object]],
    headers: list[str],
    default_entity: str,
    include_derived: bool,
    scale: float,
) -> list[dict[str, object]]:
    out: list[dict[str, object]] = []
    acct_cols = _account_columns(headers)
    cols = list(acct_cols)
    if include_derived:
        cols.extend([h for h in headers if _is_derived_col(h) and h not in cols])
    for row_idx, row in enumerate(rows, start=2):
        period = _derive_period(row)
        entity = str(row.get("entity") or default_entity)
        for col in cols:
            derived = _is_derived_col(col)
            if derived and not include_derived:
                continue
            raw_val = row.get(col)
            if raw_val in (None, ""):
                continue
            token = col.split("_", 1)[1] if "_" in col else col
            acct_num = token if token.isdigit() else ""
            acct_name = token if not token.isdigit() else ""
            amount_raw = _to_amount(raw_val)
            amount_scaled = _scaled(amount_raw, scale)
            out.append(
                {
                    "period": period,
                    "entity": entity,
                    "account_number": acct_num,
                    "account_name": acct_name,
                    "amount_raw": f"{amount_raw:.6f}",
                    "units_scale_applied": f"{scale:.6f}",
                    "amount": f"{amount_scaled:.6f}",
                    "source_column": col,
                    "source_row": str(row_idx),
                    "is_derived": "1" if derived else "0",
                    "notes": "" if not derived else "Derived source column",
                }
            )
    return out


def _pick(row: dict[str, object], options: tuple[str, ...]) -> str:
    for key in options:
        if key in row and row[key] not in (None, ""):
            return str(row[key]).strip()
    return ""


def _long_normalize(rows: list[dict[str, object]], default_entity: str, scale: float) -> list[dict[str, object]]:
    out: list[dict[str, object]] = []
    for row_idx, row in enumerate(rows, start=2):
        period = _derive_period(row)
        acct_num = _pick(row, LONG_ACCOUNT_NUMBER_FIELDS)
        acct_name = _pick(row, LONG_ACCOUNT_NAME_FIELDS)
        amount_raw = _to_amount(_pick(row, LONG_AMOUNT_FIELDS) or "0")
        amount_scaled = _scaled(amount_raw, scale)
        out.append(
            {
                "period": period,
                "entity": _pick(row, ("entity", "legal_entity", "company")) or default_entity,
                "account_number": acct_num,
                "account_name": acct_name,
                "amount_raw": f"{amount_raw:.6f}",
                "units_scale_applied": f"{scale:.6f}",
                "amount": f"{amount_scaled:.6f}",
                "source_column": "row",
                "source_row": str(row_idx),
                "is_derived": "0",
                "notes": "",
            }
        )
    return out


def _write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=CANONICAL_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Normalize trial balance input into canonical long-form CSV.")
    parser.add_argument("--input", required=True, help="Source TB file (.csv or .xlsx).")
    parser.add_argument("--output", required=True, help="Canonical CSV output path.")
    parser.add_argument("--sheet", help="Optional sheet name for XLSX input.")
    parser.add_argument("--entity", default="Entity 1", help="Default entity when source lacks entity column.")
    parser.add_argument(
        "--include-derived",
        action="store_true",
        help="Include derived total/variance columns from wide exports.",
    )
    parser.add_argument(
        "--scale",
        type=float,
        default=0.001,
        help="Scale applied to source amounts (default 0.001 for $'000 outputs).",
    )
    parser.add_argument("--profile", help="Optional JSON profile output.")
    args = parser.parse_args()

    src = Path(args.input)
    suffix = src.suffix.lower()
    if suffix == ".csv":
        headers, rows = _read_csv(src)
    elif suffix in {".xlsx", ".xlsm"}:
        headers, rows = _read_xlsx(src, args.sheet)
    else:
        raise SystemExit(f"Unsupported TB input type: {src.suffix}")

    layout = _detect_layout(headers)
    if layout == "wide":
        out_rows = _wide_to_long(rows, headers, args.entity, args.include_derived, args.scale)
    else:
        out_rows = _long_normalize(rows, args.entity, args.scale)

    if not out_rows:
        raise SystemExit("No output rows produced after normalization.")
    if any(not r["period"] for r in out_rows):
        raise SystemExit("Missing period values after normalization. Provide period/year+month fields.")

    _write_csv(Path(args.output), out_rows)

    profile = {
        "input": str(src),
        "layout": layout,
        "scale": args.scale,
        "source_rows": len(rows),
        "output_rows": len(out_rows),
        "distinct_periods": len({r["period"] for r in out_rows}),
        "distinct_accounts": len(
            {
                _normalize_key((r["account_number"] or r["account_name"]))
                for r in out_rows
                if r["account_number"] or r["account_name"]
            }
        ),
        "derived_rows": sum(1 for r in out_rows if r["is_derived"] == "1"),
    }
    if args.profile:
        Path(args.profile).write_text(json.dumps(profile, indent=2), encoding="utf-8")

    print(f"Wrote normalized TB: {args.output}")
    print(f"Layout: {layout}, rows={len(out_rows)}, scale={args.scale}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
