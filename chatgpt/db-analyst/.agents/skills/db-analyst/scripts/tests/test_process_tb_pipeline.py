#!/usr/bin/env python3
"""Integration tests for the Process-TB pipeline."""

from __future__ import annotations

import csv
import json
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook


SCRIPTS = Path(__file__).resolve().parents[1]
SKILL_ROOT = Path(__file__).resolve().parents[2]
PROJECT_ROOT = Path(__file__).resolve().parents[5]
PY = sys.executable


def _run(args: list[str]) -> None:
    subprocess.run(args, check=True)


def _write_mapping(path: Path) -> None:
    rows = [
        ("4000", "revenue", "IS", "IS_REV", "Revenue", "IS_REV_CORE", "Revenue", "IS_REV_CORE", "Revenue", "IS_REV", "Revenue", "10", "-1"),
        ("4010", "service_revenue", "IS", "IS_REV", "Revenue", "IS_REV_CORE", "Revenue", "IS_REV_CORE", "Revenue", "IS_REV", "Revenue", "10", "-1"),
        ("4020", "insurance_revenue", "IS", "IS_REV", "Revenue", "IS_REV_CORE", "Revenue", "IS_REV_CORE", "Revenue", "IS_REV", "Revenue", "10", "-1"),
        ("4090", "other_revenue", "IS", "IS_REV", "Revenue", "IS_REV_OTHER", "Other Revenue", "IS_REV_OTHER", "Other Revenue", "IS_REV", "Revenue", "10", "-1"),
        ("5000", "clinical_labor", "IS", "IS_COGS", "Cost of goods sold", "IS_COGS_CORE", "COGS", "IS_COGS_CORE", "COGS", "IS_COGS", "Cost of goods sold", "20", "-1"),
        ("5100", "labs_supplies", "IS", "IS_COGS", "Cost of goods sold", "IS_COGS_CORE", "COGS", "IS_COGS_CORE", "COGS", "IS_COGS", "Cost of goods sold", "20", "-1"),
        ("5200", "sterilization", "IS", "IS_COGS", "Cost of goods sold", "IS_COGS_CORE", "COGS", "IS_COGS_CORE", "COGS", "IS_COGS", "Cost of goods sold", "20", "-1"),
        ("5300", "other_direct_costs", "IS", "IS_COGS", "Cost of goods sold", "IS_COGS_CORE", "COGS", "IS_COGS_CORE", "COGS", "IS_COGS", "Cost of goods sold", "20", "-1"),
        ("6000", "payroll", "IS", "IS_OPEX", "Operating expenses", "IS_OPEX_CORE", "Opex", "IS_OPEX_CORE", "Opex", "IS_OPEX", "Operating expenses", "30", "-1"),
        ("6100", "occupancy", "IS", "IS_OPEX", "Operating expenses", "IS_OPEX_CORE", "Opex", "IS_OPEX_CORE", "Opex", "IS_OPEX", "Operating expenses", "30", "-1"),
        ("6200", "marketing", "IS", "IS_OPEX", "Operating expenses", "IS_OPEX_CORE", "Opex", "IS_OPEX_CORE", "Opex", "IS_OPEX", "Operating expenses", "30", "-1"),
        ("6300", "g_and_a", "IS", "IS_OPEX", "Operating expenses", "IS_OPEX_CORE", "Opex", "IS_OPEX_CORE", "Opex", "IS_OPEX", "Operating expenses", "30", "-1"),
        ("6400", "other_opex", "IS", "IS_OPEX", "Operating expenses", "IS_OPEX_CORE", "Opex", "IS_OPEX_CORE", "Opex", "IS_OPEX", "Operating expenses", "30", "-1"),
        ("1000", "cash", "BS", "BS_ASSET", "Assets", "BS_CURRENT", "Current Assets", "BS_CASH", "Cash", "BS_CASH", "Cash", "100", "1"),
        ("1100", "accounts_receivable", "BS", "BS_ASSET", "Assets", "BS_CURRENT", "Current Assets", "BS_AR", "Accounts Receivable", "BS_AR", "Accounts receivable", "110", "1"),
        ("", "prepaid", "BS", "BS_ASSET", "Assets", "BS_CURRENT", "Current Assets", "BS_PREPAID", "Prepaid", "BS_PREPAID", "Prepaid expenses", "120", "1"),
        ("", "fixed_assets", "BS", "BS_ASSET", "Assets", "BS_NONCURRENT", "Non-current Assets", "BS_PPE", "PPE", "BS_PPE", "Fixed assets", "130", "1"),
        ("", "accum_depr", "BS", "BS_ASSET", "Assets", "BS_NONCURRENT", "Non-current Assets", "BS_ACCUM", "Accumulated Depreciation", "BS_ACCUM_DEPR", "Accumulated depreciation", "140", "1"),
        ("2000", "accounts_payable", "BS", "BS_LIAB", "Liabilities", "BS_CURRENT_L", "Current Liabilities", "BS_AP", "Accounts Payable", "BS_AP", "Accounts payable", "200", "1"),
        ("", "accrued", "BS", "BS_LIAB", "Liabilities", "BS_CURRENT_L", "Current Liabilities", "BS_ACCRUED", "Accrued", "BS_ACCRUED", "Accrued liabilities", "210", "1"),
        ("", "deferred", "BS", "BS_LIAB", "Liabilities", "BS_CURRENT_L", "Current Liabilities", "BS_DEFERRED", "Deferred", "BS_DEFERRED", "Deferred revenue", "220", "1"),
        ("2500", "long_term_debt", "BS", "BS_LIAB", "Liabilities", "BS_NONCURRENT_L", "Non-current Liabilities", "BS_DEBT", "Debt", "BS_DEBT", "Long-term debt", "230", "1"),
        ("3000", "retained_earnings", "BS", "BS_EQ", "Equity", "BS_EQ_CORE", "Equity", "BS_RE", "Retained Earnings", "BS_RE", "Retained earnings", "300", "1"),
        ("3200", "common_equity", "BS", "BS_EQ", "Equity", "BS_EQ_CORE", "Equity", "BS_COMMON", "Common Equity", "BS_EQUITY", "Common equity", "310", "1"),
    ]
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(
            [
                "account_number",
                "account_name",
                "statement",
                "Level1Key",
                "Level1Name",
                "Level2Key",
                "Level2Name",
                "Level3Key",
                "Level3Name",
                "LineKey",
                "LineName",
                "SortOrder",
                "SignMultiplier",
            ]
        )
        writer.writerows(rows)


class ProcessTBPipelineTests(unittest.TestCase):
    def test_end_to_end_on_simulated_data_room_tb(self) -> None:
        with tempfile.TemporaryDirectory(prefix="process_tb_it_") as td:
            tmp = Path(td)
            tb_src = PROJECT_ROOT / "data-room" / "trial_balance.xlsx"
            tb_work = tmp / "trial_balance.xlsx"
            shutil.copy2(tb_src, tb_work)

            mapping = tmp / "mapping.csv"
            _write_mapping(mapping)

            out_tb = tmp / "tb_canonical.csv"
            out_mapped = tmp / "tb_mapped.csv"
            out_is = tmp / "is.csv"
            out_bs = tmp / "bs.csv"
            out_wb = tmp / "databook_fs.xlsx"
            qc = tmp / "qc.json"

            _run(
                [
                    PY,
                    str(SCRIPTS / "ProcessTB.py"),
                    "--tb",
                    str(tb_work),
                    "--mapping",
                    str(mapping),
                    "--out-tb",
                    str(out_tb),
                    "--out-mapped",
                    str(out_mapped),
                    "--out-is",
                    str(out_is),
                    "--out-bs",
                    str(out_bs),
                    "--qc-report",
                    str(qc),
                    "--out-workbook",
                    str(out_wb),
                    "--template",
                    str(SKILL_ROOT / "assets" / "databook-template-v2.xlsx"),
                    "--scope",
                    "fs-only",
                ]
            )

            report = json.loads(qc.read_text(encoding="utf-8"))
            self.assertIn(report["result"], {"pass", "warn"})
            checks = {c["check"]: c for c in report["checks"]}
            self.assertEqual(checks["period_raw_tieout"]["status"], "pass")
            self.assertEqual(checks["derived_rows_included"]["derived_rows"], 0)
            self.assertEqual(checks["unmapped_rows"]["unmapped"], 0)
            self.assertEqual(checks["workbook_scope_sheet_set"]["status"], "pass")
            self.assertEqual(checks["Control | Setup_required"]["status"], "pass")
            self.assertEqual(checks["Combined | IS_formula_policy"]["status"], "pass")
            self.assertEqual(checks["Combined | BS_formula_policy"]["status"], "pass")

            wb = load_workbook(out_wb, data_only=False)
            self.assertEqual(
                set(wb.sheetnames),
                {
                    "Control | Setup",
                    "Data | TB",
                    "Map | COA to Lines",
                    "Combined | IS",
                    "Combined | BS",
                    "Control | QC",
                },
            )

            ws_bs = wb["Combined | BS"]
            header_vals = [ws_bs.cell(8, c).value for c in range(7, min(ws_bs.max_column, 70) + 1)]
            self.assertFalse(any(isinstance(v, str) and v.strip().upper().startswith("FY") for v in header_vals))

            with out_is.open(newline="", encoding="utf-8") as fh:
                rows = list(csv.DictReader(fh))
            jan_rev = next(float(r["amount"]) for r in rows if r["line_key"] == "IS_REV" and r["period"] == "2023-01")
            jan_cogs = next(float(r["amount"]) for r in rows if r["line_key"] == "IS_COGS" and r["period"] == "2023-01")
            self.assertNotEqual(jan_rev, 0.0)
            self.assertNotEqual(jan_cogs, 0.0)

    def test_ingest_excludes_derived_columns_and_applies_default_scale(self) -> None:
        with tempfile.TemporaryDirectory(prefix="process_tb_ingest_") as td:
            tmp = Path(td)
            src = tmp / "wide.csv"
            out_default = tmp / "out_default.csv"
            out_incl = tmp / "out_incl.csv"
            with src.open("w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                writer.writerow(["period", "acct_1000", "acct_2000", "total_assets"])
                writer.writerow(["2025-01", "100", "-40", "60"])

            _run(
                [
                    PY,
                    str(SCRIPTS / "ingest_tb.py"),
                    "--input",
                    str(src),
                    "--output",
                    str(out_default),
                ]
            )
            _run(
                [
                    PY,
                    str(SCRIPTS / "ingest_tb.py"),
                    "--input",
                    str(src),
                    "--output",
                    str(out_incl),
                    "--include-derived",
                ]
            )

            with out_default.open(newline="", encoding="utf-8") as fh:
                rows_default = list(csv.DictReader(fh))
            with out_incl.open(newline="", encoding="utf-8") as fh:
                rows_incl = list(csv.DictReader(fh))

            self.assertEqual(len(rows_default), 2)
            self.assertEqual(len(rows_incl), 3)
            self.assertEqual(sum(1 for r in rows_incl if r["is_derived"] == "1"), 1)
            self.assertTrue(all(r["units_scale_applied"] == "0.001000" for r in rows_default))
            scaled_sum = sum(float(r["amount"]) for r in rows_default)
            raw_sum = sum(float(r["amount_raw"]) for r in rows_default)
            self.assertAlmostEqual(raw_sum, 60.0, delta=0.01)
            self.assertAlmostEqual(scaled_sum, 0.06, delta=0.0001)


if __name__ == "__main__":
    unittest.main()
