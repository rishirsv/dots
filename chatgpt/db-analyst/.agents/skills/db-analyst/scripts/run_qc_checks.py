#!/usr/bin/env python3
"""Run QC checks for the Process-TB pipeline."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


def _read_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        headers = reader.fieldnames or []
        rows = list(reader)
    return headers, rows


def _to_float(raw: str) -> float:
    s = (raw or "0").strip().replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = f"-{s[1:-1]}"
    try:
        return float(s or 0)
    except ValueError:
        return 0.0


def _check_required_columns(headers: list[str], required: set[str], name: str) -> dict[str, object]:
    missing = sorted(required - set(headers))
    return {
        "check": f"{name}_required_columns",
        "status": "pass" if not missing else "fail",
        "missing": missing,
    }


def _check_unmapped(mapped_rows: list[dict[str, str]]) -> dict[str, object]:
    total = len(mapped_rows)
    unmapped = sum(1 for r in mapped_rows if r.get("mapping_status") != "mapped")
    pct = 0 if total == 0 else unmapped / total
    status = "pass" if unmapped == 0 else ("warn" if pct <= 0.03 else "fail")
    return {
        "check": "unmapped_rows",
        "status": status,
        "unmapped": unmapped,
        "rows": total,
        "pct": round(pct, 6),
    }


def _check_period_tieout(tb_rows: list[dict[str, str]], mapped_rows: list[dict[str, str]]) -> dict[str, object]:
    tb_totals: dict[str, float] = defaultdict(float)
    map_totals: dict[str, float] = defaultdict(float)
    for row in tb_rows:
        tb_totals[row.get("period", "")] += _to_float(row.get("amount", row.get("amount_raw", "0")))
    for row in mapped_rows:
        map_totals[row.get("period", "")] += _to_float(row.get("amount_raw", "0"))

    periods = sorted(set(tb_totals) | set(map_totals))
    max_abs = 0.0
    breaks = 0
    details = []
    for p in periods:
        diff = tb_totals[p] - map_totals[p]
        max_abs = max(max_abs, abs(diff))
        if abs(diff) > 0.01:
            breaks += 1
            details.append({"period": p, "diff": round(diff, 6)})
    return {
        "check": "period_raw_tieout",
        "status": "pass" if breaks == 0 else "fail",
        "periods": len(periods),
        "breaks": breaks,
        "max_abs_diff": round(max_abs, 6),
        "details": details[:20],
    }


def _check_units_scaling(tb_rows: list[dict[str, str]]) -> dict[str, object]:
    scales = sorted({(r.get("units_scale_applied") or "").strip() for r in tb_rows if r.get("units_scale_applied")})
    if not scales:
        return {
            "check": "units_scale_recorded",
            "status": "warn",
            "scales": scales,
            "note": "units_scale_applied not found",
        }
    status = "pass" if len(scales) == 1 else "warn"
    if scales == ["1.000000"]:
        status = "warn"
    return {
        "check": "units_scale_recorded",
        "status": status,
        "scales": scales,
    }


def _check_sign_status(mapped_rows: list[dict[str, str]]) -> dict[str, object]:
    provided = sum(1 for r in mapped_rows if r.get("sign_multiplier_status") == "provided")
    derived = sum(1 for r in mapped_rows if r.get("sign_multiplier_status") == "derived")
    missing = sum(1 for r in mapped_rows if r.get("sign_multiplier_status") == "missing_assumption")
    status = "pass" if missing == 0 else "warn"
    return {
        "check": "sign_multiplier_status",
        "status": status,
        "provided": provided,
        "derived": derived,
        "missing_assumption": missing,
    }


def _check_account_completeness(mapped_rows: list[dict[str, str]]) -> dict[str, object]:
    mapped = [r for r in mapped_rows if r.get("mapping_status") == "mapped"]
    account_to_statement: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in mapped:
        key = (row.get("account_number", "").strip(), row.get("account_name", "").strip())
        account_to_statement[key].add(row.get("statement", "").strip())

    both_or_none = []
    for key, stmts in account_to_statement.items():
        stmts = {s for s in stmts if s}
        if len(stmts) != 1:
            both_or_none.append({"account": key, "statements": sorted(stmts)})

    return {
        "check": "account_statement_uniqueness",
        "status": "pass" if not both_or_none else "fail",
        "accounts_checked": len(account_to_statement),
        "violations": both_or_none[:20],
    }


def _classify_bs_group(row: dict[str, str]) -> str:
    text = f"{row.get('level1_name','')} {row.get('level2_name','')} {row.get('line_name','')}".lower()
    if "asset" in text:
        if any(k in text for k in ("accum", "allowance", "contra", "reserve")):
            return "contra_asset"
        return "asset"
    if "equity" in text:
        return "equity"
    if "liab" in text or "debt" in text or "payable" in text:
        return "liability"
    return "unknown"


def _check_bs_balance(mapped_rows: list[dict[str, str]]) -> dict[str, object]:
    bs_rows = [r for r in mapped_rows if r.get("mapping_status") == "mapped" and r.get("statement") == "BS"]
    if not bs_rows:
        return {"check": "bs_balance_a_equals_l_plus_e", "status": "warn", "note": "No BS rows"}

    by_period: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for row in bs_rows:
        p = row.get("period", "")
        grp = _classify_bs_group(row)
        by_period[p][grp] += _to_float(row.get("amount_signed", "0"))

    breaks = []
    for p, vals in by_period.items():
        assets = vals.get("asset", 0.0) + vals.get("contra_asset", 0.0)
        liab = vals.get("liability", 0.0)
        eq = vals.get("equity", 0.0)
        diff = assets + liab + eq
        if abs(diff) > 0.05:
            breaks.append(
                {
                    "period": p,
                    "assets": round(assets, 4),
                    "liabilities": round(liab, 4),
                    "equity": round(eq, 4),
                    "diff": round(diff, 4),
                }
            )

    return {
        "check": "bs_balance_a_equals_l_plus_e",
        "status": "pass" if not breaks else "warn",
        "periods": len(by_period),
        "breaks": breaks[:20],
    }


def _check_statement_coverage(mapped_rows: list[dict[str, str]]) -> dict[str, object]:
    periods: dict[str, set[str]] = defaultdict(set)
    for row in mapped_rows:
        if row.get("mapping_status") == "mapped":
            periods[row.get("period", "")].add(row.get("statement", ""))
    missing = [p for p, stmts in periods.items() if not {"IS", "BS"}.issubset(stmts)]
    return {
        "check": "statement_coverage",
        "status": "pass" if not missing else "warn",
        "periods": len(periods),
        "missing_periods": missing[:20],
    }


def _check_derived_row_inclusion(mapped_rows: list[dict[str, str]]) -> dict[str, object]:
    derived = sum(1 for r in mapped_rows if r.get("is_derived") in {"1", "true", "True"})
    return {
        "check": "derived_rows_included",
        "status": "pass" if derived == 0 else "warn",
        "derived_rows": derived,
    }


def _check_trend_file(path: Path, statement: str) -> dict[str, object]:
    headers, rows = _read_rows(path)
    col_check = _check_required_columns(
        headers,
        {"statement", "line_key", "line_name", "period", "amount"},
        f"{statement.lower()}_trend",
    )
    if col_check["status"] == "fail":
        return col_check
    bad_stmt = sum(1 for r in rows if r.get("statement") != statement)
    return {
        "check": f"{statement.lower()}_trend_statement_label",
        "status": "pass" if bad_stmt == 0 else "fail",
        "rows": len(rows),
        "bad_statement_rows": bad_stmt,
    }


def _formula_policy_check(ws, sheet_name: str) -> dict[str, object]:
    offenders = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 500), min_col=1, max_col=min(ws.max_column, 180)):
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            upper = v.upper()
            if "IF(COUNT" in upper or "IF(COUNTA" in upper or "IF(OR(" in upper:
                offenders.append({"cell": cell.coordinate, "formula": v})
                if len(offenders) >= 20:
                    break
        if len(offenders) >= 20:
            break

    return {
        "check": f"{sheet_name}_formula_policy",
        "status": "pass" if not offenders else "fail",
        "offenders": offenders,
    }


def _check_bs_header_no_fy(ws) -> dict[str, object]:
    offenders = []
    for col in range(7, min(ws.max_column, 180) + 1):
        v = ws.cell(8, col).value
        if isinstance(v, str) and re.match(r"^\s*FY\d{0,2}\s*$", v, flags=re.IGNORECASE):
            offenders.append({"cell": ws.cell(8, col).coordinate, "value": v})
    return {
        "check": "Combined | BS_header_no_fy_totals",
        "status": "pass" if not offenders else "fail",
        "offenders": offenders,
    }


def _check_total_border_top_black(ws, sheet_name: str) -> dict[str, object]:
    bad = 0
    checked = 0
    for row in range(9, min(ws.max_row, 260) + 1):
        label = ws.cell(row, 2).value
        if not isinstance(label, str):
            continue
        ll = label.lower()
        if "total" in ll or "net income" in ll or "ebitda" in ll:
            checked += 1
            b = ws.cell(row, 7).border
            top = b.top
            color = (top.color.rgb if top and top.color else None) or ""
            if top.style != "thin" or color not in {"FF000000", "00000000"}:
                bad += 1
    status = "pass" if checked == 0 or bad == 0 else "warn"
    return {
        "check": f"{sheet_name}_total_top_black_border",
        "status": status,
        "checked_rows": checked,
        "bad_rows": bad,
    }


def _check_workbook(path: Path, expected_sheets: list[str]) -> list[dict[str, object]]:
    wb = load_workbook(path)
    checks: list[dict[str, object]] = []

    if expected_sheets:
        actual = set(wb.sheetnames)
        expected = set(expected_sheets)
        checks.append(
            {
                "check": "workbook_scope_sheet_set",
                "status": "pass" if actual == expected else "fail",
                "expected": sorted(expected),
                "actual": sorted(actual),
            }
        )

    checks.append(
        {
            "check": "Control | Setup_required",
            "status": "pass" if "Control | Setup" in wb.sheetnames else "fail",
        }
    )

    for sheet_name in ("Combined | IS", "Combined | BS"):
        if sheet_name not in wb.sheetnames:
            checks.append({"check": f"{sheet_name}_exists", "status": "fail"})
            continue
        ws = wb[sheet_name]
        row7 = ws.row_dimensions[7].height
        checks.append(
            {
                "check": f"{sheet_name}_row7_height",
                "status": "pass" if abs((row7 or 0) - 19.5) < 0.01 else "warn",
                "value": row7,
            }
        )

        max_col = min(ws.max_column, 18)
        bad_align = 0
        for col in range(7, max_col + 1):
            if ws.cell(8, col).alignment.horizontal != "right":
                bad_align += 1
        checks.append(
            {
                "check": f"{sheet_name}_period_header_alignment",
                "status": "pass" if bad_align == 0 else "warn",
                "bad_cells": bad_align,
            }
        )

        bad_nf = 0
        for row in range(9, min(ws.max_row, 60) + 1):
            for col in range(7, max_col + 1):
                v = ws.cell(row, col).value
                if isinstance(v, (int, float)):
                    nf = ws.cell(row, col).number_format or ""
                    if "#,##0" not in nf:
                        bad_nf += 1
        checks.append(
            {
                "check": f"{sheet_name}_number_format",
                "status": "pass" if bad_nf == 0 else "warn",
                "bad_cells": bad_nf,
            }
        )

        checks.append(_formula_policy_check(ws, sheet_name))
        checks.append(_check_total_border_top_black(ws, sheet_name))

    if "Combined | BS" in wb.sheetnames:
        checks.append(_check_bs_header_no_fy(wb["Combined | BS"]))

    cf_count = 0
    for sheet in wb.worksheets:
        cf_count += len(sheet.conditional_formatting)
    checks.append(
        {
            "check": "workbook_conditional_formatting_count",
            "status": "pass" if cf_count == 0 else "warn",
            "count": cf_count,
        }
    )
    return checks


def main() -> int:
    parser = argparse.ArgumentParser(description="Run Process-TB QC checks and emit JSON report.")
    parser.add_argument("--tb", required=True, help="Canonical TB CSV path.")
    parser.add_argument("--mapped", required=True, help="Mapped TB CSV path.")
    parser.add_argument("--is-trend", help="Optional IS trend CSV path.")
    parser.add_argument("--bs-trend", help="Optional BS trend CSV path.")
    parser.add_argument("--workbook", help="Optional workbook path for formatting/scope checks.")
    parser.add_argument("--expected-sheets", help="Comma-separated expected workbook sheets.")
    parser.add_argument("--report", required=True, help="QC report JSON output path.")
    args = parser.parse_args()

    tb_headers, tb_rows = _read_rows(Path(args.tb))
    mapped_headers, mapped_rows = _read_rows(Path(args.mapped))

    checks = [
        _check_required_columns(
            tb_headers,
            {"period", "account_number", "account_name", "amount", "amount_raw", "units_scale_applied", "is_derived"},
            "tb",
        ),
        _check_required_columns(
            mapped_headers,
            {
                "period",
                "amount_raw",
                "amount_signed",
                "statement",
                "line_key",
                "mapping_status",
                "sign_multiplier_status",
            },
            "mapped",
        ),
        _check_unmapped(mapped_rows),
        _check_period_tieout(tb_rows, mapped_rows),
        _check_units_scaling(tb_rows),
        _check_sign_status(mapped_rows),
        _check_account_completeness(mapped_rows),
        _check_statement_coverage(mapped_rows),
        _check_bs_balance(mapped_rows),
        _check_derived_row_inclusion(mapped_rows),
    ]

    if args.is_trend:
        checks.append(_check_trend_file(Path(args.is_trend), "IS"))
    if args.bs_trend:
        checks.append(_check_trend_file(Path(args.bs_trend), "BS"))

    if args.workbook:
        expected = [s.strip() for s in (args.expected_sheets or "").split(",") if s.strip()]
        checks.extend(_check_workbook(Path(args.workbook), expected))

    result = "pass"
    for c in checks:
        if c["status"] == "fail":
            result = "fail"
            break
        if c["status"] == "warn" and result != "fail":
            result = "warn"

    report = {
        "inputs": {
            "tb": args.tb,
            "mapped": args.mapped,
            "is_trend": args.is_trend,
            "bs_trend": args.bs_trend,
            "workbook": args.workbook,
        },
        "checks": checks,
        "result": result,
    }

    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, indent=2), encoding="utf-8")
    print(f"Wrote QC report: {report_path} (result={result})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
