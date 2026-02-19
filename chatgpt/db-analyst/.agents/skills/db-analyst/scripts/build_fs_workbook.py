#!/usr/bin/env python3
"""Build FS workbook outputs from Process-TB pipeline artifacts."""

from __future__ import annotations

import argparse
import csv
import json
from calendar import monthrange
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

TITLE_ROW = 7
HEADER_ROW = 8
DATA_START_ROW = 9
MONTH_START_COL = 7  # G

NUMBER_FMT = "#,##0_);(#,##0);\"-\"_);@"
DATE_FMT_IS = "mmm-yy"
DATE_FMT_BS = "dd-mmm-yy"

NO_BORDER = Border()
THIN_TOP_BLACK = Border(top=Side(style="thin", color="FF000000"))
NO_FILL = PatternFill(fill_type=None)


@dataclass(frozen=True)
class RowKey:
    level1: str
    level2: str
    level3: str
    line_key: str
    line_name: str
    account_number: str
    account_name: str


@dataclass(frozen=True)
class SummaryCol:
    label: object
    col_type: str  # sum | as_at | avg
    start_idx: int
    end_idx: int


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8-sig") as fh:
        return list(csv.DictReader(fh))


def _ensure_sheet(wb: Workbook, name: str) -> None:
    if name not in wb.sheetnames:
        wb.create_sheet(name)


def _ensure_required_sheets(wb: Workbook, required: list[str]) -> None:
    for name in required:
        _ensure_sheet(wb, name)


def _delete_non_required_sheets(wb: Workbook, required: set[str]) -> None:
    for name in list(wb.sheetnames):
        if name not in required:
            del wb[name]


def _clear_sheet_data(ws, start_row: int = HEADER_ROW, max_cols: int = 140, max_rows: int = 400) -> None:
    last_row = max(ws.max_row, max_rows)
    last_col = max(ws.max_column, max_cols)
    for row in ws.iter_rows(min_row=start_row, max_row=last_row, min_col=1, max_col=last_col):
        for cell in row:
            cell.value = None


def _period_sort_key(value: str) -> tuple[int, int, str]:
    v = (value or "").strip()
    if len(v) >= 7 and v[4] == "-":
        try:
            return int(v[:4]), int(v[5:7]), v
        except ValueError:
            return 9999, 99, v
    return 9999, 99, v


def _period_to_eom(period: str) -> datetime:
    y = int(period[:4])
    m = int(period[5:7])
    return datetime(y, m, monthrange(y, m)[1])


def _year_label(year: int) -> str:
    return f"FY{str(year)[-2:]}"


def _year_groups(periods: list[str]) -> list[tuple[int, list[int]]]:
    groups: dict[int, list[int]] = defaultdict(list)
    for idx, p in enumerate(periods):
        groups[int(p[:4])].append(idx)
    return [(y, groups[y]) for y in sorted(groups)]


def _populate_flat_table(ws, rows: list[dict[str, str]], headers: list[str], start_row: int = HEADER_ROW) -> None:
    for idx, header in enumerate(headers, start=1):
        ws.cell(start_row, idx).value = header
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, header in enumerate(headers, start=1):
            ws.cell(r_idx, c_idx).value = row.get(header, "")


def _to_float(raw: str) -> float:
    s = (raw or "0").strip().replace(",", "")
    if s.startswith("(") and s.endswith(")"):
        s = f"-{s[1:-1]}"
    try:
        return float(s)
    except ValueError:
        return 0.0


def _account_rollup_rows(mapped_rows: list[dict[str, str]], statement: str) -> tuple[list[str], list[dict[str, object]]]:
    rows = [r for r in mapped_rows if r.get("mapping_status") == "mapped" and r.get("statement") == statement]
    periods = sorted({r.get("period", "") for r in rows if r.get("period")}, key=_period_sort_key)

    account_totals: dict[RowKey, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    row_order: dict[RowKey, float] = {}

    for r in rows:
        key = RowKey(
            level1=r.get("level1_name", "") or "Unclassified",
            level2=r.get("level2_name", "") or "Unclassified",
            level3=r.get("level3_name", "") or "Unclassified",
            line_key=r.get("line_key", ""),
            line_name=r.get("line_name", ""),
            account_number=r.get("account_number", ""),
            account_name=r.get("account_name", ""),
        )
        p = r.get("period", "")
        if not p:
            continue
        account_totals[key][p] += _to_float(r.get("amount_signed", "0"))
        try:
            sort_order = float((r.get("sort_order") or "").strip())
        except ValueError:
            sort_order = 1_000_000.0
        row_order[key] = min(row_order.get(key, sort_order), sort_order)

    keys = sorted(
        account_totals,
        key=lambda k: (
            row_order.get(k, 1_000_000.0),
            k.level1,
            k.level2,
            k.level3,
            k.line_key,
            k.account_number,
            k.account_name,
        ),
    )

    output_rows: list[dict[str, object]] = []

    def _emit_subtotal(label: str, subtotal_rows: list[dict[str, object]], level_tag: str) -> None:
        if not subtotal_rows:
            return
        vals: dict[str, float] = defaultdict(float)
        for r in subtotal_rows:
            for p in periods:
                vals[p] += float(r.get(p, 0.0) or 0.0)
        row: dict[str, object] = {
            "line_key": f"{level_tag}_SUBTOTAL",
            "line_name": label,
            "account_number": "",
            "sign": "+",
            "source": "Formula",
            "comments": f"{level_tag} subtotal",
            "is_total": True,
        }
        for p in periods:
            row[p] = vals[p]
        output_rows.append(row)

    current_l1 = current_l2 = current_l3 = current_line = None
    l1_bucket: list[dict[str, object]] = []
    l2_bucket: list[dict[str, object]] = []
    l3_bucket: list[dict[str, object]] = []
    line_bucket: list[dict[str, object]] = []

    for key in keys:
        if current_line is not None and key.line_name != current_line:
            _emit_subtotal(f"{current_line} subtotal", line_bucket, "LINE")
            line_bucket = []
        if current_l3 is not None and key.level3 != current_l3:
            _emit_subtotal(f"{current_l3} subtotal", l3_bucket, "L3")
            l3_bucket = []
        if current_l2 is not None and key.level2 != current_l2:
            _emit_subtotal(f"{current_l2} subtotal", l2_bucket, "L2")
            l2_bucket = []
        if current_l1 is not None and key.level1 != current_l1:
            _emit_subtotal(f"{current_l1} subtotal", l1_bucket, "L1")
            l1_bucket = []

        current_l1, current_l2, current_l3, current_line = key.level1, key.level2, key.level3, key.line_name

        line_label = key.account_name or key.line_name or key.account_number
        row: dict[str, object] = {
            "line_key": key.line_key,
            "line_name": line_label,
            "account_number": key.account_number,
            "sign": "+",
            "source": "TB/Map",
            "comments": "",
            "is_total": False,
            "level1": key.level1,
            "level2": key.level2,
            "level3": key.level3,
        }
        for p in periods:
            row[p] = account_totals[key].get(p, 0.0)

        output_rows.append(row)
        line_bucket.append(row)
        l3_bucket.append(row)
        l2_bucket.append(row)
        l1_bucket.append(row)

    if current_line is not None:
        _emit_subtotal(f"{current_line} subtotal", line_bucket, "LINE")
    if current_l3 is not None:
        _emit_subtotal(f"{current_l3} subtotal", l3_bucket, "L3")
    if current_l2 is not None:
        _emit_subtotal(f"{current_l2} subtotal", l2_bucket, "L2")
    if current_l1 is not None:
        _emit_subtotal(f"{current_l1} subtotal", l1_bucket, "L1")

    return periods, output_rows


def _is_total_row(row: dict[str, object]) -> bool:
    label = str(row.get("line_name", "")).lower()
    return bool(row.get("is_total")) or "total" in label or "net income" in label or "ebitda" in label


def _build_summary_cols(periods: list[str], statement: str) -> list[SummaryCol]:
    groups = _year_groups(periods)
    cols: list[SummaryCol] = []

    if statement == "IS":
        for year, idxs in groups:
            cols.append(SummaryCol(label=_year_label(year), col_type="sum", start_idx=min(idxs), end_idx=max(idxs)))
        return cols

    # BS: present period-end dates for as-at columns, plus average presentation by fiscal year.
    for year, idxs in groups:
        as_at_date = _period_to_eom(periods[max(idxs)])
        cols.append(SummaryCol(label=as_at_date, col_type="as_at", start_idx=max(idxs), end_idx=max(idxs)))
    for year, idxs in groups:
        cols.append(SummaryCol(label=f"Average {_year_label(year)}", col_type="avg", start_idx=min(idxs), end_idx=max(idxs)))
    return cols


def _apply_statement_layout(
    ws,
    periods: list[str],
    rows: list[dict[str, object]],
    units_text: str,
    date_format: str,
    statement: str,
) -> tuple[int, int, int]:
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[TITLE_ROW].height = 19.5
    ws.freeze_panes = "G9"

    ws.cell(HEADER_ROW, 1).value = "LineKey"
    ws.cell(HEADER_ROW, 2).value = units_text
    ws.cell(HEADER_ROW, 3).value = "TB Account No."
    ws.cell(HEADER_ROW, 4).value = "Sign"
    ws.cell(HEADER_ROW, 5).value = "Source"
    ws.cell(HEADER_ROW, 6).value = "Comments"

    # Month headers from Control | Setup anchor cells.
    for idx, _period in enumerate(periods, start=MONTH_START_COL):
        c = ws.cell(HEADER_ROW, idx)
        setup_row = 22 + (idx - MONTH_START_COL)
        c.value = f"='Control | Setup'!$A${setup_row}"
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = date_format

    # Data rows.
    for r_off, row in enumerate(rows):
        r = DATA_START_ROW + r_off
        ws.cell(r, 1).value = row.get("line_key")
        ws.cell(r, 2).value = row.get("line_name")
        ws.cell(r, 3).value = row.get("account_number")
        ws.cell(r, 4).value = row.get("sign")
        ws.cell(r, 5).value = row.get("source")
        ws.cell(r, 6).value = row.get("comments")
        ws.cell(r, 2).alignment = Alignment(horizontal="left", vertical="center")

        for idx, period in enumerate(periods, start=MONTH_START_COL):
            c = ws.cell(r, idx)
            c.value = float(row.get(period, 0.0) or 0.0)
            c.number_format = NUMBER_FMT
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.border = NO_BORDER
            c.fill = NO_FILL

    month_end_col = MONTH_START_COL + len(periods) - 1

    # Summary columns (no IF wrappers, only SUM/AVERAGE/direct references).
    summary_cols = _build_summary_cols(periods, statement)
    summary_start_col = month_end_col + 1
    for i, sc in enumerate(summary_cols):
        col = summary_start_col + i
        hc = ws.cell(HEADER_ROW, col)
        hc.value = sc.label
        hc.alignment = Alignment(horizontal="right", vertical="center")
        if statement == "BS" and isinstance(sc.label, datetime):
            hc.number_format = DATE_FMT_BS
        else:
            hc.number_format = "@"

        start_col = MONTH_START_COL + sc.start_idx
        end_col = MONTH_START_COL + sc.end_idx

        for r_off, _row in enumerate(rows):
            r = DATA_START_ROW + r_off
            dc = ws.cell(r, col)
            if sc.col_type == "sum":
                dc.value = f"=SUM({ws.cell(r, start_col).coordinate}:{ws.cell(r, end_col).coordinate})"
                dc.number_format = NUMBER_FMT
            elif sc.col_type == "as_at":
                dc.value = f"={ws.cell(r, end_col).coordinate}"
                dc.number_format = NUMBER_FMT
            else:
                dc.value = f"=AVERAGE({ws.cell(r, start_col).coordinate}:{ws.cell(r, end_col).coordinate})"
                dc.number_format = NUMBER_FMT
            dc.alignment = Alignment(horizontal="right", vertical="center")
            dc.border = NO_BORDER
            dc.fill = NO_FILL

    # BS check block with corrected sign logic: Assets + Liabilities+Equity should be zero under sign convention.
    bs_check_row = 0
    if statement == "BS":
        label_to_row = {str(row.get("line_name", "")).lower(): DATA_START_ROW + i for i, row in enumerate(rows)}
        assets_row = next((v for k, v in label_to_row.items() if "total assets" in k), None)
        tle_row = next((v for k, v in label_to_row.items() if "total liabilities + equity" in k), None)
        if assets_row and tle_row:
            sec_row = DATA_START_ROW + len(rows) + 2
            bs_check_row = sec_row + 1
            ws.cell(sec_row, 1).value = "BS_CHECKS"
            ws.cell(bs_check_row, 1).value = "Balance check (Assets + Liabilities & Equity)"

            for col in range(MONTH_START_COL, summary_start_col + len(summary_cols)):
                ws.cell(bs_check_row, col).value = f"={ws.cell(assets_row, col).coordinate}+{ws.cell(tle_row, col).coordinate}"
                ws.cell(bs_check_row, col).number_format = NUMBER_FMT
                ws.cell(bs_check_row, col).alignment = Alignment(horizontal="right", vertical="center")
                ws.cell(bs_check_row, col).border = NO_BORDER

    # Totals: only thin black top border in numeric area.
    for r_off, row in enumerate(rows):
        if not _is_total_row(row):
            continue
        r = DATA_START_ROW + r_off
        for col in range(MONTH_START_COL, summary_start_col + len(summary_cols)):
            ws.cell(r, col).border = THIN_TOP_BLACK
        ws.cell(r, 2).font = Font(bold=True)

    # Header numeric alignment.
    last_col = summary_start_col + len(summary_cols) - 1
    for col in range(MONTH_START_COL, last_col + 1):
        ws.cell(HEADER_ROW, col).alignment = Alignment(horizontal="right", vertical="center")

    return DATA_START_ROW, last_col, bs_check_row


def _write_qc_sheet(wb, report_path: Path | None) -> None:
    if "Control | QC" not in wb.sheetnames or report_path is None or not report_path.exists():
        return
    ws = wb["Control | QC"]
    _clear_sheet_data(ws, HEADER_ROW, max_cols=20, max_rows=300)
    report = json.loads(report_path.read_text(encoding="utf-8"))
    ws.cell(8, 1).value = "Check"
    ws.cell(8, 2).value = "Status"
    ws.cell(8, 3).value = "Details"
    row = 9
    for chk in report.get("checks", []):
        ws.cell(row, 1).value = chk.get("check")
        ws.cell(row, 2).value = chk.get("status")
        ws.cell(row, 3).value = json.dumps(chk, ensure_ascii=True)
        row += 1


def _write_setup_sheet(ws, source: str, units_text: str, scale: float, periods: list[str]) -> None:
    _clear_sheet_data(ws, 1, max_cols=8, max_rows=250)
    ws.cell(1, 1).value = "Control Setup"
    ws.cell(2, 1).value = f"Source: {source}"
    ws.cell(3, 1).value = f"Units: {units_text}"
    ws.cell(4, 1).value = f"Scale applied: {scale:.6f}"
    ws.cell(20, 1).value = "Period End Dates"
    ws.cell(21, 1).value = "Period"
    ws.cell(21, 2).value = "Fiscal Year"

    for idx, period in enumerate(periods):
        row = 22 + idx
        dt = _period_to_eom(period)
        ws.cell(row, 1).value = dt
        ws.cell(row, 1).number_format = DATE_FMT_BS
        ws.cell(row, 1).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row, 2).value = _year_label(dt.year)  # explicit string fiscal-year labels


def main() -> int:
    parser = argparse.ArgumentParser(description="Build FS workbook from Process-TB outputs.")
    parser.add_argument("--template", required=True, help="Template workbook path.")
    parser.add_argument("--tb", required=True, help="Canonical TB CSV.")
    parser.add_argument("--mapped", required=True, help="Mapped TB CSV.")
    parser.add_argument("--is-trend", help="IS trend CSV (optional, for audit).")
    parser.add_argument("--bs-trend", help="BS trend CSV (optional, for audit).")
    parser.add_argument("--qc-report", help="QC report JSON path (optional).")
    parser.add_argument("--output", required=True, help="Output workbook path.")
    parser.add_argument("--scope", choices=["fs-only", "full"], default="fs-only")
    parser.add_argument("--source-label", default="trial balance input")
    parser.add_argument("--units", default="$'000")
    parser.add_argument("--scale", type=float, default=0.001)
    args = parser.parse_args()

    template = Path(args.template)
    wb = load_workbook(template)

    required_fs = [
        "Control | Setup",
        "Data | TB",
        "Map | COA to Lines",
        "Combined | IS",
        "Combined | BS",
        "Control | QC",
    ]
    _ensure_required_sheets(wb, required_fs)

    if args.scope == "fs-only":
        _delete_non_required_sheets(wb, set(required_fs))

    tb_rows = _read_csv(Path(args.tb))
    mapped_rows = _read_csv(Path(args.mapped))

    if "Data | TB" in wb.sheetnames:
        ws = wb["Data | TB"]
        _clear_sheet_data(ws, HEADER_ROW)
        tb_headers = list(tb_rows[0].keys()) if tb_rows else []
        _populate_flat_table(ws, tb_rows, tb_headers, HEADER_ROW)

    if "Map | COA to Lines" in wb.sheetnames:
        ws = wb["Map | COA to Lines"]
        _clear_sheet_data(ws, HEADER_ROW)
        seen = set()
        map_rows: list[dict[str, str]] = []
        for r in mapped_rows:
            k = (r.get("account_number", ""), r.get("account_name", ""), r.get("line_key", ""), r.get("statement", ""))
            if k in seen:
                continue
            seen.add(k)
            map_rows.append(
                {
                    "account_number": r.get("account_number", ""),
                    "account_name": r.get("account_name", ""),
                    "statement": r.get("statement", ""),
                    "Level1Key": r.get("level1_key", ""),
                    "Level1Name": r.get("level1_name", ""),
                    "Level2Key": r.get("level2_key", ""),
                    "Level2Name": r.get("level2_name", ""),
                    "Level3Key": r.get("level3_key", ""),
                    "Level3Name": r.get("level3_name", ""),
                    "LineKey": r.get("line_key", ""),
                    "LineName": r.get("line_name", ""),
                    "SortOrder": r.get("sort_order", ""),
                    "SignMultiplier": r.get("sign_multiplier", ""),
                    "Notes": r.get("mapping_note", ""),
                }
            )
        map_headers = list(map_rows[0].keys()) if map_rows else [
            "account_number",
            "account_name",
            "statement",
            "Level1Key",
            "Level1Name",
            "Level2Key",
            "Level2Name",
            "Level3Key",
            "Level3Name",
            "LineKey",
            "LineName",
            "SortOrder",
            "SignMultiplier",
            "Notes",
        ]
        _populate_flat_table(ws, map_rows, map_headers, HEADER_ROW)

    is_periods: list[str] = []
    if "Combined | IS" in wb.sheetnames:
        ws = wb["Combined | IS"]
        _clear_sheet_data(ws, HEADER_ROW)
        is_periods, is_rows = _account_rollup_rows(mapped_rows, "IS")
        _apply_statement_layout(ws, is_periods, is_rows, args.units, DATE_FMT_IS, "IS")

    bs_periods: list[str] = []
    if "Combined | BS" in wb.sheetnames:
        ws = wb["Combined | BS"]
        _clear_sheet_data(ws, HEADER_ROW)
        bs_periods, bs_rows = _account_rollup_rows(mapped_rows, "BS")
        _apply_statement_layout(ws, bs_periods, bs_rows, args.units, DATE_FMT_BS, "BS")

    all_periods = sorted(set(is_periods) | set(bs_periods), key=_period_sort_key)
    _write_setup_sheet(wb["Control | Setup"], args.source_label, args.units, args.scale, all_periods)
    _write_qc_sheet(wb, Path(args.qc_report) if args.qc_report else None)

    out = Path(args.output)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote workbook: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
