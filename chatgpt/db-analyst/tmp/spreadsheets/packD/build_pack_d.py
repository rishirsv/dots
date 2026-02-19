from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

SRC = 'output/spreadsheet/databook-template-v1-packC-clean-v3.xlsx'
OUT = 'output/spreadsheet/databook-template-v1-packD-clean-v1.xlsx'

wb = load_workbook(SRC)

# Pull canonical style tokens from existing tabs.
style_ws = wb['Combined | IS']
setup_ws = wb['Control | Setup']
index_ws = wb['Template | Index']

STYLE = {
    'title': setup_ws['A1'],
    'source': setup_ws['A2'],
    'note_label': setup_ws['A4'],
    'section_bar': style_ws['A7'],   # dark navy (no border)
    'header': style_ws['A8'],        # cobalt header
    'data': style_ws['A9'],          # thin border base
    'input': style_ws['G9'],         # light-blue input / linked
    'check': style_ws['G28'],        # yellow check cell
    'manifest_data': index_ws['A10'],
    'index_data': index_ws['A34'],
}

SUBTITLE_FONT = Font(name='Arial', size=10, italic=True, color='FF666666')
LEFT_CENTER = Alignment(horizontal='left', vertical='center')
LEFT_TOP = Alignment(horizontal='left', vertical='top')


def apply_style(cell, template_key):
    template = STYLE[template_key]
    cell._style = copy(template._style)


def style_range(ws, r1, r2, c1, c2, template_key):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            apply_style(ws.cell(r, c), template_key)


def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def set_defaults(ws):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.zoomScaleNormal = 100


def clear_sheet_if_exists(name):
    if name in wb.sheetnames:
        wb.remove(wb[name])


for sheet_name in ['Control | QC', 'Recons', 'Agings']:
    clear_sheet_if_exists(sheet_name)

# Create Control | QC after Control | Setup
qc_index = wb.sheetnames.index('Control | Setup') + 1
ws_qc = wb.create_sheet('Control | QC', qc_index)
set_defaults(ws_qc)
set_col_widths(ws_qc, {
    'A': 16, 'B': 34, 'C': 18, 'D': 28, 'E': 10, 'F': 12, 'G': 10, 'H': 10, 'I': 42,
})
ws_qc.row_dimensions[1].height = 24

ws_qc['A1'] = 'Control | QC'
apply_style(ws_qc['A1'], 'title')
ws_qc['A2'] = 'Source: <document_source>'
apply_style(ws_qc['A2'], 'source')

ws_qc['A4'] = 'Workbook QC Status'
apply_style(ws_qc['A4'], 'note_label')
ws_qc['B4'] = '=IF(COUNTIF(G9:G200,FALSE)>0,"FAIL",IF(COUNTIF(G9:G200,TRUE)>0,"PASS","PENDING"))'
apply_style(ws_qc['B4'], 'check')
ws_qc['B4'].font = Font(name='Arial', size=10, bold=True, color='FF000000')
ws_qc['B4'].alignment = LEFT_CENTER

ws_qc['A6'] = 'Centralized pass/fail matrix linked to check cells across all core sheets.'
ws_qc['A6'].font = SUBTITLE_FONT
ws_qc['A6'].alignment = LEFT_CENTER

style_range(ws_qc, 7, 7, 1, 9, 'section_bar')
ws_qc['A7'] = 'QC Matrix'
ws_qc['A7'].font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
ws_qc['A7'].alignment = LEFT_CENTER

headers = ['CheckID', 'CheckName', 'Sheet', 'CellRef', 'Expected', 'Actual', 'PassFlag', 'Severity', 'Notes']
style_range(ws_qc, 8, 8, 1, 9, 'header')
for idx, h in enumerate(headers, start=1):
    cell = ws_qc.cell(8, idx)
    cell.value = h
    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
    cell.alignment = LEFT_CENTER

qc_rows = [
    ('QC_IS_NI', 'Combined IS net income tie-out', 'Combined | IS', "'Combined | IS'!G28:K28", 'High', 'Should resolve to 0'),
    ('QC_IS_EBITDA', 'Combined IS EBITDA tie-out', 'Combined | IS', "'Combined | IS'!G29:K29", 'High', 'Should resolve to 0'),
    ('QC_BS_BAL', 'Combined BS balance check', 'Combined | BS', "'Combined | BS'!G28:K28", 'High', 'Should resolve to 0'),
    ('QC_CF_NET', 'Combined CF net cash tie-out', 'Combined | CF', "'Combined | CF'!G28:K28", 'High', 'Should resolve to 0'),
    ('QC_QOFE_REV', 'QofE Summary revenue tie-out', 'QofE | Summary', "'QofE | Summary'!G20:J20", 'Med', 'Should resolve to 0 once linked'),
    ('QC_QOFE_EB', 'QofE Summary EBITDA tie-out', 'QofE | Summary', "'QofE | Summary'!G21:J21", 'Med', 'Should resolve to 0 once linked'),
    ('QC_QOFE_DET', 'QofE Detail adjusted EBITDA tie-out', 'QofE | Detail', "'QofE | Detail'!G28:K28", 'Med', 'Should resolve to 0 once linked'),
    ('QC_NWC_SUM', 'NWC Summary tie-out', 'NWC | Summary', "'NWC | Summary'!G17:K17", 'High', 'Should resolve to 0 once linked'),
    ('QC_NWC_DET', 'NWC Detail tie-out', 'NWC | Detail', "'NWC | Detail'!G28:K28", 'High', 'Should resolve to 0 once linked'),
    ('QC_STRAT_BS', 'Stratified BS tie-out', 'Stratified | BS', "'Stratified | BS'!G28:K28", 'High', 'Should resolve to 0 once linked'),
    ('QC_NETDEBT', 'Net debt tie-out', 'Net_debt', "'Net_debt'!G28:K28", 'High', 'Should resolve to 0 once linked'),
    ('QC_RECONS_IS', 'Recons IS section checks', 'Recons', "'Recons'!H9:H13", 'Med', 'Should resolve to 0 once linked'),
    ('QC_RECONS_BS', 'Recons BS section checks', 'Recons', "'Recons'!H18:H22", 'Med', 'Should resolve to 0 once linked'),
    ('QC_RECONS_CF', 'Recons CF section checks', 'Recons', "'Recons'!H27:H31", 'Med', 'Should resolve to 0 once linked'),
    ('QC_RECONS_OTH', 'Recons Other section checks', 'Recons', "'Recons'!H36:H40", 'Med', 'Should resolve to 0 once linked'),
    ('QC_AGING_AR', 'AR aging tie-out check', 'Agings', "'Agings'!J17", 'High', 'Should resolve to 0 once linked'),
    ('QC_AGING_AP', 'AP aging tie-out check', 'Agings', "'Agings'!J33", 'High', 'Should resolve to 0 once linked'),
    ('QC_AGING_ALL', 'Combined aging tie-out check', 'Agings', "'Agings'!J40", 'High', 'Should resolve to 0 once linked'),
]

start = 9
for i, (check_id, check_name, sheet_name, cell_ref, severity, note) in enumerate(qc_rows):
    r = start + i
    style_range(ws_qc, r, r, 1, 9, 'data')
    ws_qc.cell(r, 1).value = check_id
    ws_qc.cell(r, 2).value = check_name
    ws_qc.cell(r, 3).value = sheet_name
    ws_qc.cell(r, 4).value = cell_ref
    ws_qc.cell(r, 5).value = 0
    ws_qc.cell(r, 6).value = f'=IF(D{r}="","",SUMPRODUCT(ABS(INDIRECT(D{r}))))'
    ws_qc.cell(r, 7).value = f'=IF(F{r}="","",IF(ABS(F{r}-E{r})<=0.01,TRUE,FALSE))'
    ws_qc.cell(r, 8).value = severity
    ws_qc.cell(r, 9).value = note
    for c in range(1, 10):
        ws_qc.cell(r, c).alignment = LEFT_CENTER

# Create Recons + Agings after Net_debt.
net_idx = wb.sheetnames.index('Net_debt')
ws_recons = wb.create_sheet('Recons', net_idx + 1)
ws_agings = wb.create_sheet('Agings', net_idx + 2)

for ws in [ws_recons, ws_agings]:
    set_defaults(ws)
    ws.row_dimensions[1].height = 24

set_col_widths(ws_recons, {
    'A': 16, 'B': 22, 'C': 16, 'D': 14, 'E': 14, 'F': 14, 'G': 24, 'H': 12, 'I': 12,
})

ws_recons['A1'] = 'Recons'
apply_style(ws_recons['A1'], 'title')
ws_recons['A2'] = 'Source: <document_source>'
apply_style(ws_recons['A2'], 'source')
ws_recons['A4'] = 'Cross-tab reconciliation mini-grids for IS, BS, CF, and other tie-outs.'
ws_recons['A4'].font = SUBTITLE_FONT
ws_recons['A4'].alignment = LEFT_CENTER

recon_headers = ['LineKey', "$'000", 'TB Account No.', 'Source value', 'Databook value', 'Variance', 'Explanation', 'Check (=0)', 'Status']
recon_sections = [
    (7, 'Recons | IS', [
        ('R_IS_REV', 'Revenue'),
        ('R_IS_NI', 'Net income'),
        ('R_IS_EBITDA', 'EBITDA'),
        ('R_IS_OPEX', 'Operating expenses'),
        ('R_IS_OTHER', 'Other IS line'),
    ]),
    (16, 'Recons | BS', [
        ('R_BS_TA', 'Total assets'),
        ('R_BS_TL', 'Total liabilities'),
        ('R_BS_EQ', 'Equity'),
        ('R_BS_TLE', 'Liabilities + equity'),
        ('R_BS_OTHER', 'Other BS line'),
    ]),
    (25, 'Recons | CF', [
        ('R_CF_NET', 'Net cash flow'),
        ('R_CF_OP', 'Operating cash flow'),
        ('R_CF_INV', 'Investing cash flow'),
        ('R_CF_FIN', 'Financing cash flow'),
        ('R_CF_OTHER', 'Other CF line'),
    ]),
    (34, 'Recons | Other', [
        ('R_OTH_AR', 'AR aging to BS'),
        ('R_OTH_AP', 'AP aging to BS'),
        ('R_OTH_NWC', 'NWC Summary to NWC Detail'),
        ('R_OTH_ND', 'Net debt to Stratified BS'),
        ('R_OTH_MISC', 'Other tie-out'),
    ]),
]

for start_row, title, items in recon_sections:
    style_range(ws_recons, start_row, start_row, 1, 9, 'section_bar')
    ws_recons.cell(start_row, 1).value = title
    ws_recons.cell(start_row, 1).font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
    ws_recons.cell(start_row, 1).alignment = LEFT_CENTER

    style_range(ws_recons, start_row + 1, start_row + 1, 1, 9, 'header')
    for c, header in enumerate(recon_headers, start=1):
        cell = ws_recons.cell(start_row + 1, c)
        cell.value = header
        cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
        cell.alignment = LEFT_CENTER

    for idx, (line_key, line_name) in enumerate(items, start=0):
        r = start_row + 2 + idx
        style_range(ws_recons, r, r, 1, 9, 'data')
        ws_recons.cell(r, 1).value = line_key
        ws_recons.cell(r, 2).value = line_name
        ws_recons.cell(r, 5).value = None
        ws_recons.cell(r, 6).value = f'=IF(OR(D{r}="",E{r}=""),"",E{r}-D{r})'
        ws_recons.cell(r, 8).value = f'=IF(F{r}="","",F{r})'
        ws_recons.cell(r, 9).value = f'=IF(H{r}="","",IF(ABS(H{r})<=0.01,"PASS","FAIL"))'

        apply_style(ws_recons.cell(r, 4), 'input')
        apply_style(ws_recons.cell(r, 5), 'input')
        apply_style(ws_recons.cell(r, 8), 'check')

        for c in range(1, 10):
            ws_recons.cell(r, c).alignment = LEFT_CENTER

style_range(ws_recons, 43, 43, 1, 9, 'section_bar')
ws_recons['A43'] = 'RECONS_CHECKS'
ws_recons['A43'].font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
ws_recons['A43'].alignment = LEFT_CENTER

style_range(ws_recons, 44, 45, 1, 9, 'data')
ws_recons['A44'] = 'Open recon fails (should be 0)'
ws_recons['H44'] = '=COUNTIF(I9:I40,"FAIL")'
apply_style(ws_recons['H44'], 'check')
ws_recons['I44'] = '=IF(H44=0,"PASS","FAIL")'

ws_recons['A45'] = 'Absolute variance sum (should be 0)'
ws_recons['H45'] = '=SUMPRODUCT(ABS(F9:F40))'
apply_style(ws_recons['H45'], 'check')
ws_recons['I45'] = '=IF(H45=0,"PASS","FAIL")'

for r in [44, 45]:
    for c in range(1, 10):
        ws_recons.cell(r, c).alignment = LEFT_CENTER

# Build Agings tab.
set_col_widths(ws_agings, {
    'A': 16, 'B': 22, 'C': 16, 'D': 10, 'E': 10, 'F': 10, 'G': 10, 'H': 10, 'I': 10, 'J': 12, 'K': 12,
})

ws_agings['A1'] = 'Agings'
apply_style(ws_agings['A1'], 'title')
ws_agings['A2'] = 'Source: <document_source>'
apply_style(ws_agings['A2'], 'source')
ws_agings['A4'] = 'AR/AP aging scaffolds with canonical buckets and control tie-out checks.'
ws_agings['A4'].font = SUBTITLE_FONT
ws_agings['A4'].alignment = LEFT_CENTER

aging_headers = ['LineKey', "$'000", 'Counterparty ID', 'Current', '1-30', '31-60', '61-90', '91-120', '>120', 'Total', 'Source']

# AR section
style_range(ws_agings, 7, 7, 1, 11, 'section_bar')
ws_agings['A7'] = 'Aging | AR'
ws_agings['A7'].font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
ws_agings['A7'].alignment = LEFT_CENTER

style_range(ws_agings, 8, 8, 1, 11, 'header')
for c, h in enumerate(aging_headers, start=1):
    cell = ws_agings.cell(8, c)
    cell.value = h
    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
    cell.alignment = LEFT_CENTER

for i in range(6):
    r = 9 + i
    style_range(ws_agings, r, r, 1, 11, 'data')
    ws_agings.cell(r, 1).value = f'AR_{i+1:03d}'
    ws_agings.cell(r, 2).value = f'Top customer {i+1}'
    ws_agings.cell(r, 3).value = f'AR{100+i+1}'
    ws_agings.cell(r, 10).value = f'=IF(COUNTA(D{r}:I{r})=0,"",SUM(D{r}:I{r}))'
    ws_agings.cell(r, 11).value = 'Input'
    for c in range(4, 10):
        apply_style(ws_agings.cell(r, c), 'input')
    ws_agings.cell(r, 10).alignment = LEFT_CENTER

style_range(ws_agings, 15, 15, 1, 11, 'data')
ws_agings['A15'] = 'AR_TOTAL'
ws_agings['B15'] = 'Total AR aging'
ws_agings['J15'] = '=IF(COUNT(J9:J14)=0,"",SUM(J9:J14))'
ws_agings['K15'] = 'Formula'

style_range(ws_agings, 16, 16, 1, 11, 'section_bar')
ws_agings['A16'] = 'AGING_AR_CHECKS'
ws_agings['A16'].font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
ws_agings['A16'].alignment = LEFT_CENTER

style_range(ws_agings, 17, 18, 1, 11, 'data')
ws_agings['A17'] = 'AR aging tie-out vs AR control'
ws_agings['J17'] = '=IF(OR(J15="",J18=""),"",J15-J18)'
apply_style(ws_agings['J17'], 'check')
ws_agings['K17'] = 'Should be 0'
ws_agings['A18'] = 'AR_CONTROL'
ws_agings['B18'] = 'AR control account (from Combined | BS)'
apply_style(ws_agings['J18'], 'input')
ws_agings['K18'] = 'Input/Link'

# AP section
style_range(ws_agings, 23, 23, 1, 11, 'section_bar')
ws_agings['A23'] = 'Aging | AP'
ws_agings['A23'].font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
ws_agings['A23'].alignment = LEFT_CENTER

style_range(ws_agings, 24, 24, 1, 11, 'header')
for c, h in enumerate(aging_headers, start=1):
    cell = ws_agings.cell(24, c)
    cell.value = h
    cell.font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
    cell.alignment = LEFT_CENTER

for i in range(6):
    r = 25 + i
    style_range(ws_agings, r, r, 1, 11, 'data')
    ws_agings.cell(r, 1).value = f'AP_{i+1:03d}'
    ws_agings.cell(r, 2).value = f'Top vendor {i+1}'
    ws_agings.cell(r, 3).value = f'AP{100+i+1}'
    ws_agings.cell(r, 10).value = f'=IF(COUNTA(D{r}:I{r})=0,"",SUM(D{r}:I{r}))'
    ws_agings.cell(r, 11).value = 'Input'
    for c in range(4, 10):
        apply_style(ws_agings.cell(r, c), 'input')
    ws_agings.cell(r, 10).alignment = LEFT_CENTER

style_range(ws_agings, 31, 31, 1, 11, 'data')
ws_agings['A31'] = 'AP_TOTAL'
ws_agings['B31'] = 'Total AP aging'
ws_agings['J31'] = '=IF(COUNT(J25:J30)=0,"",SUM(J25:J30))'
ws_agings['K31'] = 'Formula'

style_range(ws_agings, 32, 32, 1, 11, 'section_bar')
ws_agings['A32'] = 'AGING_AP_CHECKS'
ws_agings['A32'].font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
ws_agings['A32'].alignment = LEFT_CENTER

style_range(ws_agings, 33, 34, 1, 11, 'data')
ws_agings['A33'] = 'AP aging tie-out vs AP control'
ws_agings['J33'] = '=IF(OR(J31="",J34=""),"",J31-J34)'
apply_style(ws_agings['J33'], 'check')
ws_agings['K33'] = 'Should be 0'
ws_agings['A34'] = 'AP_CONTROL'
ws_agings['B34'] = 'AP control account (from Combined | BS)'
apply_style(ws_agings['J34'], 'input')
ws_agings['K34'] = 'Input/Link'

style_range(ws_agings, 39, 39, 1, 11, 'section_bar')
ws_agings['A39'] = 'AGING_CHECKS'
ws_agings['A39'].font = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
ws_agings['A39'].alignment = LEFT_CENTER

style_range(ws_agings, 40, 40, 1, 11, 'data')
ws_agings['A40'] = 'Combined AR/AP tie-out check'
ws_agings['J40'] = '=IF(OR(J17="",J33=""),"",J17+J33)'
apply_style(ws_agings['J40'], 'check')
ws_agings['K40'] = 'Should be 0'

# Consistent alignment for Agings used range.
for r in range(7, 41):
    for c in range(1, 12):
        ws_agings.cell(r, c).alignment = LEFT_CENTER

# Keep divider tabs in page-break preview mode at 100% as requested.
for divider in ['Financials>>', 'QofE>>', 'NWC>>']:
    if divider in wb.sheetnames:
        ws = wb[divider]
        ws.sheet_view.view = 'pageBreakPreview'
        set_defaults(ws)

# Ensure all sheets keep global defaults.
for ws in wb.worksheets:
    set_defaults(ws)

# Update Template | Index manifest + block rows.
idx = wb['Template | Index']

# Combined-sheet references and dependencies.
idx['C5'] = 'Control | Setup; Control | QC'
idx['F6'] = 'Control | Setup'
idx['E8'] = 'Control | Setup; Data | TB; Map | COA to Lines'
idx['E9'] = 'Control | Setup; Data | TB; Map | COA to Lines'
idx['E10'] = 'Control | Setup; Data | TB; Map | NWC & NetDebt Class'

# Add manifest rows for Pack D modules.
manifest_rows = {
    12: ['CORE_RECONS', 'Reconciliations', 'Recons', 'Cross-tab reconciliation mini-grids', 'Core output checks', 'Combined | IS; Combined | BS; Combined | CF', 'All deals', 'FALSE'],
    13: ['CORE_AGINGS', 'Agings', 'Agings', 'AR/AP aging blocks with canonical buckets and tie-outs', 'Data | AR Aging; Data | AP Aging', 'Combined | BS', 'All deals', 'FALSE'],
    14: ['CORE_QC', 'QC dashboard', 'Control | QC', 'Centralized workbook check status', 'Check cells across core sheets', 'All core modules', 'All deals', 'FALSE'],
}
for r, vals in manifest_rows.items():
    style_range(idx, r, r, 1, 8, 'manifest_data')
    for c, v in enumerate(vals, start=1):
        idx.cell(r, c).value = v
        idx.cell(r, c).alignment = LEFT_CENTER

# Fix combined Index entries for merged setup/index layout.
idx['A35'] = 'Control | Setup'
idx['B35'] = 'PERIODS_BLOCK'
idx['C35'] = 'A20'
idx['D35'] = 20
idx['E35'] = 22
idx['F35'] = 'A'
idx['G35'] = 21
idx['H35'] = None
idx['I35'] = 'grid'
idx['J35'] = 'FALSE'
idx['K35'] = 'A22:A299 must be month-end dates'

idx['A36'] = 'Template | Index'
idx['B36'] = 'MANIFEST_MAIN'
idx['C36'] = 'A3'
idx['D36'] = 4
idx['E36'] = 5
idx['F36'] = None
idx['G36'] = None
idx['H36'] = None
idx['I36'] = 'grid'
idx['J36'] = 'FALSE'
idx['K36'] = 'Module keys should be stable'

idx['A37'] = 'Template | Index'
idx['B37'] = 'INDEX_MAIN'
idx['C37'] = 'A32'
idx['D37'] = 33
idx['E37'] = 34
idx['F37'] = None
idx['G37'] = None
idx['H37'] = None
idx['I37'] = 'grid'
idx['J37'] = 'FALSE'
idx['K37'] = 'Update when block anchors move'

idx['K38'] = 'Month columns map to Control | Setup A22:A57'

# Append Pack D index rows.
pack_d_rows = [
    ['Control | QC', 'QC_MAIN', 'A8', 8, 9, None, None, None, 'grid', 'FALSE', 'PassFlag TRUE when check resolves'],
    ['Recons', 'RECONS_IS', 'A7', 8, 9, None, None, None, 'grid', 'FALSE', "B8 stores units as $'000"],
    ['Recons', 'RECONS_BS', 'A16', 17, 18, None, None, None, 'grid', 'FALSE', 'Section-level variances should resolve to 0'],
    ['Recons', 'RECONS_CF', 'A25', 26, 27, None, None, None, 'grid', 'FALSE', 'Section-level variances should resolve to 0'],
    ['Recons', 'RECONS_OTHER', 'A34', 35, 36, None, None, None, 'grid', 'FALSE', 'Section-level variances should resolve to 0'],
    ['Recons', 'RECONS_CHECKS', 'A43', 43, 44, None, None, None, 'grid', 'FALSE', 'Overall open fails should resolve to 0'],
    ['Agings', 'AGING_AR', 'A7', 8, 9, None, None, None, 'grid', 'FALSE', "B8 stores units as $'000"],
    ['Agings', 'AGING_AP', 'A23', 24, 25, None, None, None, 'grid', 'FALSE', "B24 stores units as $'000"],
    ['Agings', 'AGING_CHECKS', 'A39', 39, 40, None, None, None, 'grid', 'FALSE', 'AR/AP tie-out should resolve to 0'],
]

start_row = 70
for offset, row_vals in enumerate(pack_d_rows):
    r = start_row + offset
    style_range(idx, r, r, 1, 11, 'index_data')
    for c, v in enumerate(row_vals, start=1):
        idx.cell(r, c).value = v
        idx.cell(r, c).alignment = LEFT_CENTER

# Extend helper columns if present in row 34 style footprint.
for r in range(70, 70 + len(pack_d_rows)):
    for c in range(12, 45):
        idx.cell(r, c)._style = copy(idx.cell(34, c)._style)

wb.save(OUT)
print(f'Wrote {OUT}')
