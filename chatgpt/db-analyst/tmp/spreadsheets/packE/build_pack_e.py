from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

SRC = 'output/spreadsheet/databook-template-v1-packD-clean-v1.xlsx'
OUT = 'output/spreadsheet/databook-template-v1-packE-clean-v2.xlsx'

wb = load_workbook(SRC)

# Style tokens from existing tabs.
style_ws = wb['Combined | IS']
setup_ws = wb['Control | Setup']
index_ws = wb['Template | Index']

STYLE = {
    'title': setup_ws['A1'],
    'source': setup_ws['A2'],
    'note_label': setup_ws['A4'],
    'section_bar': style_ws['A7'],     # dark navy, no borders
    'header': style_ws['A8'],          # cobalt header row
    'data': style_ws['A9'],            # base data row style
    'input': style_ws['G9'],           # light-blue input style
    'check': style_ws['G28'],          # yellow check style
    'manifest_data': index_ws['A10'],
    'index_data': index_ws['A34'],
}

LEFT_CENTER = Alignment(horizontal='left', vertical='center')
SUBTITLE_FONT = Font(name='Arial', size=10, italic=True, color='FF666666')
WHITE_BOLD = Font(name='Arial', size=10, bold=True, color='FFFFFFFF')
BLACK_BOLD = Font(name='Arial', size=10, bold=True, color='FF000000')


def apply_style(cell, key):
    cell._style = copy(STYLE[key]._style)


def style_range(ws, r1, r2, c1, c2, key):
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            apply_style(ws.cell(r, c), key)


def set_defaults(ws):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.zoomScaleNormal = 100


def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def remove_if_exists(name):
    if name in wb.sheetnames:
        wb.remove(wb[name])


for s in ['Cover', 'Personnel', 'Capex', 'Leases', 'Other | Analysis']:
    remove_if_exists(s)

# Cover goes first.
ws_cover = wb.create_sheet('Cover', 0)
set_defaults(ws_cover)
set_col_widths(ws_cover, {
    'A': 20, 'B': 30, 'C': 16, 'D': 16, 'E': 22, 'F': 18, 'G': 16, 'H': 16,
})
ws_cover.row_dimensions[1].height = 24

ws_cover['A1'] = 'Cover'
apply_style(ws_cover['A1'], 'title')
ws_cover['A2'] = 'Source: <document_source>'
apply_style(ws_cover['A2'], 'source')
ws_cover['A4'] = 'Template dashboard with navigation and QC status-at-a-glance.'
ws_cover['A4'].font = SUBTITLE_FONT
ws_cover['A4'].alignment = LEFT_CENTER

style_range(ws_cover, 7, 7, 1, 8, 'section_bar')
ws_cover['A7'] = 'Navigation'
ws_cover['A7'].font = WHITE_BOLD
ws_cover['A7'].alignment = LEFT_CENTER

style_range(ws_cover, 8, 8, 1, 8, 'header')
cover_headers = ['Section', 'Go to sheet', 'Status', 'Notes', 'Last refresh', 'Owner', 'Priority', 'Complete']
for c, h in enumerate(cover_headers, start=1):
    cell = ws_cover.cell(8, c)
    cell.value = h
    cell.font = WHITE_BOLD
    cell.alignment = LEFT_CENTER

nav_rows = [
    ('Control', 'Go to Control | Setup', 'PENDING', 'Template controls and periods'),
    ('Financials', 'Go to Combined | IS', 'PENDING', 'IS/BS/CF scaffolds'),
    ('QofE', 'Go to QofE | Summary', 'PENDING', 'QofE summary and detail'),
    ('NWC', 'Go to NWC | Summary', 'PENDING', 'NWC summary/detail/days'),
    ('Net debt', 'Go to Net_debt', 'PENDING', 'Net debt scaffold'),
    ('Recons', 'Go to Recons', 'PENDING', 'Cross-tab reconciliations'),
    ('Agings', 'Go to Agings', 'PENDING', 'AR/AP aging tie-outs'),
    ('Modules', 'Go to Personnel', 'PENDING', 'Base module library tabs'),
]

for i, (section, link, status, notes) in enumerate(nav_rows):
    r = 9 + i
    style_range(ws_cover, r, r, 1, 8, 'data')
    ws_cover.cell(r, 1).value = section
    ws_cover.cell(r, 2).value = link
    ws_cover.cell(r, 3).value = status
    ws_cover.cell(r, 4).value = notes
    ws_cover.cell(r, 5).value = '<last_refresh>'
    ws_cover.cell(r, 6).value = '<preparer>'
    ws_cover.cell(r, 7).value = 'Core'
    ws_cover.cell(r, 8).value = False
    for c in range(1, 9):
        ws_cover.cell(r, c).alignment = LEFT_CENTER

style_range(ws_cover, 20, 20, 1, 8, 'section_bar')
ws_cover['A20'] = 'Workbook QC'
ws_cover['A20'].font = WHITE_BOLD
ws_cover['A20'].alignment = LEFT_CENTER

style_range(ws_cover, 21, 23, 1, 8, 'data')
ws_cover['A21'] = 'Overall QC status'
ws_cover['B21'] = 'See Control | QC!B4'
apply_style(ws_cover['B21'], 'check')
ws_cover['A22'] = 'Open check count'
ws_cover['B22'] = 'Count FALSE in Control | QC!G9:G200'
apply_style(ws_cover['B22'], 'check')
ws_cover['A23'] = 'Last source used'
ws_cover['B23'] = '<document_source>'
for r in [21, 22, 23]:
    for c in range(1, 9):
        ws_cover.cell(r, c).alignment = LEFT_CENTER

# Module scaffolds.
module_defs = [
    ('Personnel', 'Personnel module scaffold for headcount and compensation analytics.', [
        ('PER_HC', 'Headcount', 'Input', ''),
        ('PER_PAYROLL', 'Payroll', 'Input', ''),
        ('PER_PAY_PER_HC', 'Payroll per HC', 'Formula', 'Derived'),
        ('PER_BONUS', 'Bonus', 'Input', ''),
        ('PER_CONTRACT', 'Contractors', 'Input', ''),
        ('PER_TOTAL_COMP', 'Total personnel cost', 'Formula', 'Derived'),
    ]),
    ('Capex', 'Capex module scaffold for maintenance/growth trends and intensity checks.', [
        ('CAP_MAINT', 'Maintenance capex', 'Input', ''),
        ('CAP_GROWTH', 'Growth capex', 'Input', ''),
        ('CAP_TOTAL', 'Total capex', 'Formula', 'Derived'),
        ('CAP_SALES', 'Sales', 'Input', ''),
        ('CAP_PCT', 'Capex % sales', 'Formula', 'Derived'),
        ('CAP_NONREC', 'Non-recurring capex', 'Input', ''),
    ]),
    ('Leases', 'Leases module scaffold for lease run-rate and balance tie-outs.', [
        ('LEASE_EXP', 'Lease expense', 'Input', ''),
        ('LEASE_ST_LIAB', 'Lease liability - current', 'Input', ''),
        ('LEASE_LT_LIAB', 'Lease liability - non-current', 'Input', ''),
        ('LEASE_TOTAL_LIAB', 'Total lease liability', 'Formula', 'Derived'),
        ('LEASE_INT', 'Lease interest', 'Input', ''),
        ('LEASE_CASH', 'Lease cash burden', 'Formula', 'Derived'),
    ]),
    ('Other | Analysis', 'Other analysis scaffold for bespoke deal-specific review items.', [
        ('OTH_ADJ_1', 'Other adjustment 1', 'Input', ''),
        ('OTH_ADJ_2', 'Other adjustment 2', 'Input', ''),
        ('OTH_ADJ_TOTAL', 'Total adjustments', 'Formula', 'Derived'),
        ('OTH_NONREC', 'Non-recurring impact', 'Input', ''),
        ('OTH_RUNRATE', 'Run-rate impact', 'Formula', 'Derived'),
        ('OTH_REVIEW', 'Reviewer note line', 'Input', 'Review'),
    ]),
]


def create_module_sheet(name, subtitle, lines):
    ws = wb.create_sheet(name)
    set_defaults(ws)
    set_col_widths(ws, {
        'A': 16, 'B': 24, 'C': 16, 'D': 7, 'E': 10, 'F': 18, 'G': 10, 'H': 10, 'I': 10, 'J': 10,
    })
    ws.row_dimensions[1].height = 24

    ws['A1'] = name
    apply_style(ws['A1'], 'title')
    ws['A2'] = 'Source: <document_source>'
    apply_style(ws['A2'], 'source')
    ws['A4'] = subtitle
    ws['A4'].font = SUBTITLE_FONT
    ws['A4'].alignment = LEFT_CENTER

    style_range(ws, 7, 7, 1, 10, 'section_bar')
    ws['A7'] = f'{name} | Main'
    ws['A7'].font = WHITE_BOLD
    ws['A7'].alignment = LEFT_CENTER

    headers = ['LineKey', "$'000", 'TB Account No.', 'Sign', 'Source', 'Comments', 'FY-3', 'FY-2', 'FY-1', 'TTM']
    style_range(ws, 8, 8, 1, 10, 'header')
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(8, c)
        cell.value = h
        cell.font = WHITE_BOLD
        cell.alignment = LEFT_CENTER

    start_row = 9
    for idx, (line_key, line_name, source, comment) in enumerate(lines):
        r = start_row + idx
        style_range(ws, r, r, 1, 10, 'data')
        ws.cell(r, 1).value = line_key
        ws.cell(r, 2).value = line_name
        ws.cell(r, 4).value = '+'
        ws.cell(r, 5).value = source
        ws.cell(r, 6).value = comment

        # Period columns as editable inputs by default.
        for c in [7, 8, 9, 10]:
            apply_style(ws.cell(r, c), 'input')

        for c in range(1, 11):
            ws.cell(r, c).alignment = LEFT_CENTER

    # Lightweight formulas for derived rows.
    if name == 'Personnel':
        for col in ['G', 'H', 'I', 'J']:
            ws[f'{col}11'] = f'=IF(OR({col}9="",{col}10="",{col}9=0),"",{col}10/{col}9)'
            ws[f'{col}14'] = f'=IF(COUNT({col}10:{col}13)=0,"",SUM({col}10,{col}12,{col}13))'
    elif name == 'Capex':
        for col in ['G', 'H', 'I', 'J']:
            ws[f'{col}11'] = f'=IF(COUNT({col}9:{col}10)=0,"",SUM({col}9:{col}10))'
            ws[f'{col}13'] = f'=IF(OR({col}11="",{col}12="",{col}12=0),"",{col}11/{col}12)'
    elif name == 'Leases':
        for col in ['G', 'H', 'I', 'J']:
            ws[f'{col}12'] = f'=IF(COUNT({col}10:{col}11)=0,"",SUM({col}10:{col}11))'
            ws[f'{col}14'] = f'=IF(COUNT({col}9:{col}13)=0,"",SUM({col}9,{col}13))'
    elif name == 'Other | Analysis':
        for col in ['G', 'H', 'I', 'J']:
            ws[f'{col}11'] = f'=IF(COUNT({col}9:{col}10)=0,"",SUM({col}9:{col}10))'
            ws[f'{col}13'] = f'=IF(COUNT({col}11:{col}12)=0,"",SUM({col}11:{col}12))'
    elif name == 'GM | LOB':
        for col in ['G', 'H', 'I', 'J']:
            ws[f'{col}11'] = f'=IF(COUNT({col}9:{col}10)=0,"",{col}9+{col}10)'
            ws[f'{col}14'] = f'=IF(COUNT({col}12:{col}13)=0,"",{col}12+{col}13)'
            ws[f'{col}15'] = f'=IF(COUNT({col}11:{col}14)=0,"",SUM({col}11,{col}14))'
            ws[f'{col}16'] = f'=IF(OR({col}15="",{col}9="",{col}12="",({col}9+{col}12)=0),"",{col}15/({col}9+{col}12))'

    # Checks block.
    style_range(ws, 27, 27, 1, 10, 'section_bar')
    ws['A27'] = f"{name.replace(' | ', '_').replace(' ', '_').upper()}_CHECKS"
    ws['A27'].font = WHITE_BOLD
    ws['A27'].alignment = LEFT_CENTER

    style_range(ws, 28, 29, 1, 10, 'data')
    ws['A28'] = 'Primary tie-out check'
    ws['A29'] = 'Secondary tie-out check'
    for col in ['G', 'H', 'I', 'J']:
        if name == 'Personnel':
            ws[f'{col}28'] = f'=IF({col}14="","",{col}14-({col}10+{col}12+{col}13))'
            ws[f'{col}29'] = f'=IF(OR({col}11="",{col}10="",{col}9="",{col}9=0),"",{col}11-({col}10/{col}9))'
        elif name == 'Capex':
            ws[f'{col}28'] = f'=IF({col}11="","",{col}11-({col}9+{col}10))'
            ws[f'{col}29'] = f'=IF(OR({col}13="",{col}11="",{col}12="",{col}12=0),"",{col}13-({col}11/{col}12))'
        elif name == 'Leases':
            ws[f'{col}28'] = f'=IF({col}12="","",{col}12-({col}10+{col}11))'
            ws[f'{col}29'] = f'=IF({col}14="","",{col}14-({col}9+{col}13))'
        elif name == 'Other | Analysis':
            ws[f'{col}28'] = f'=IF({col}11="","",{col}11-({col}9+{col}10))'
            ws[f'{col}29'] = f'=IF({col}13="","",{col}13-({col}11+{col}12))'
        elif name == 'GM | LOB':
            ws[f'{col}28'] = f'=IF({col}15="","",{col}15-({col}11+{col}14))'
            ws[f'{col}29'] = f'=IF(OR({col}16="",{col}15="",{col}9="",{col}12="",({col}9+{col}12)=0),"",{col}16-({col}15/({col}9+{col}12)))'

        apply_style(ws[f'{col}28'], 'check')
        apply_style(ws[f'{col}29'], 'check')

    for r in [28, 29]:
        for c in range(1, 11):
            ws.cell(r, c).alignment = LEFT_CENTER

    return ws


for name, subtitle, lines in module_defs:
    create_module_sheet(name, subtitle, lines)

# Keep divider sheets in page-break preview, 100%.
for divider in ['Financials>>', 'QofE>>', 'NWC>>']:
    if divider in wb.sheetnames:
        ws = wb[divider]
        ws.sheet_view.view = 'pageBreakPreview'
        set_defaults(ws)

# Ensure defaults globally.
for ws in wb.worksheets:
    set_defaults(ws)

# Update Template | Index manifest + block registry.
idx = wb['Template | Index']

# Add manifest rows for Pack E modules.
manifest_rows = {
    15: ['CORE_COVER', 'Cover', 'Cover', 'Workbook landing and navigation dashboard', 'Control | QC', 'All modules', 'All deals', 'FALSE'],
    16: ['MOD_PERSONNEL', 'Personnel', 'Personnel', 'Personnel analytics scaffold', 'Data | TB', 'Control | Setup; Map | COA to Lines', 'Deals with payroll workstreams', 'FALSE'],
    17: ['MOD_CAPEX', 'Capex', 'Capex', 'Capex analytics scaffold', 'Data | TB', 'Control | Setup; Map | COA to Lines', 'Capex-heavy deals', 'FALSE'],
    18: ['MOD_LEASES', 'Leases', 'Leases', 'Leases analytics scaffold', 'Lease schedules; Data | TB', 'Control | Setup; Map | COA to Lines', 'Deals with material leasing', 'FALSE'],
    19: ['MOD_OTHER_ANALYSIS', 'Other analysis', 'Other | Analysis', 'Flexible analysis scaffold', 'Deal-specific supporting data', 'Control | Setup', 'Custom analyses', 'FALSE'],
}
for r, vals in manifest_rows.items():
    style_range(idx, r, r, 1, 8, 'manifest_data')
    for c, v in enumerate(vals, start=1):
        idx.cell(r, c).value = v
        idx.cell(r, c).alignment = LEFT_CENTER

# Append Pack E index rows after current content.
last_content_row = 1
for r in range(1, idx.max_row + 1):
    if any(idx.cell(r, c).value is not None for c in range(1, 12)):
        last_content_row = r
start = max(last_content_row + 1, 79)

index_rows = [
    ['Cover', 'COVER_NAV', 'A7', 8, 9, None, None, None, 'grid', 'FALSE', 'Navigation links should open core sheets'],
    ['Cover', 'COVER_STATUS', 'A20', 20, 21, None, None, None, 'grid', 'FALSE', "Overall status should tie to Control | QC!B4"],
    ['Personnel', 'PERSONNEL_MAIN', 'A8', 8, 9, 'G', 8, None, 'grid', 'FALSE', "B8 stores units as $'000"],
    ['Personnel', 'PERSONNEL_CHECKS', 'A27', 27, 28, 'G', 8, None, 'grid', 'FALSE', 'Checks should resolve to 0 once linked'],
    ['Capex', 'CAPEX_MAIN', 'A8', 8, 9, 'G', 8, None, 'grid', 'FALSE', "B8 stores units as $'000"],
    ['Capex', 'CAPEX_CHECKS', 'A27', 27, 28, 'G', 8, None, 'grid', 'FALSE', 'Checks should resolve to 0 once linked'],
    ['Leases', 'LEASES_MAIN', 'A8', 8, 9, 'G', 8, None, 'grid', 'FALSE', "B8 stores units as $'000"],
    ['Leases', 'LEASES_CHECKS', 'A27', 27, 28, 'G', 8, None, 'grid', 'FALSE', 'Checks should resolve to 0 once linked'],
    ['Other | Analysis', 'OTHER_ANALYSIS_MAIN', 'A8', 8, 9, 'G', 8, None, 'grid', 'FALSE', "B8 stores units as $'000"],
    ['Other | Analysis', 'OTHER_ANALYSIS_CHECKS', 'A27', 27, 28, 'G', 8, None, 'grid', 'FALSE', 'Checks should resolve to 0 once linked'],
]

for i, vals in enumerate(index_rows):
    r = start + i
    style_range(idx, r, r, 1, 11, 'index_data')
    for c, v in enumerate(vals, start=1):
        idx.cell(r, c).value = v
        idx.cell(r, c).alignment = LEFT_CENTER
    # Extend helper columns style footprint.
    for c in range(12, 45):
        idx.cell(r, c)._style = copy(idx.cell(34, c)._style)

wb.save(OUT)
print(f'Wrote {OUT}')
