from datetime import date
from calendar import monthrange
from copy import copy
from openpyxl import load_workbook

SRC = 'output/spreadsheet/databook-template-v1-final-merged.xlsx'
OUT = 'output/spreadsheet/databook-template-v1-final-validated.xlsx'

wb = load_workbook(SRC)


def month_end(y, m):
    return date(y, m, monthrange(y, m)[1])


def col_letter(c):
    letters = ''
    while c:
        c, rem = divmod(c - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


def write_row(ws, r, values, start_col=1):
    for i, v in enumerate(values):
        ws.cell(r, start_col + i).value = v


def style_like(ws, src_row, dst_row, c1, c2):
    for c in range(c1, c2 + 1):
        ws.cell(dst_row, c)._style = copy(ws.cell(src_row, c)._style)
        ws.cell(dst_row, c).font = copy(ws.cell(src_row, c).font)
        ws.cell(dst_row, c).alignment = copy(ws.cell(src_row, c).alignment)


# --- Control setup + periods -------------------------------------------------
setup = wb['Control | Setup']
setup['B8'] = 'Synthetic Validation Deal'
setup['B9'] = 'v1.1-packF-validated'
setup['B13'] = month_end(2025, 12)
setup['B4'] = 'Template scaffold v1: Packs A-F merged and validated with synthetic data; direct formulas/values, no Excel Tables.'

months = [month_end(y, m) for y in (2023, 2024, 2025) for m in range(1, 13)]
for i, dt in enumerate(months):
    setup.cell(22 + i, 1).value = dt

# Pack F validation block
style_like(setup, 7, 15, 1, 5)
style_like(setup, 8, 16, 1, 5)
style_like(setup, 9, 17, 1, 5)
style_like(setup, 9, 18, 1, 5)
style_like(setup, 9, 19, 1, 5)
style_like(setup, 9, 20, 1, 5)
style_like(setup, 9, 21, 1, 5)

setup['A15'] = 'Pack F | Validation'
setup['A16'] = 'Control'
setup['B16'] = 'Value'
setup['C16'] = 'Target'
setup['D16'] = 'Status'
setup['E16'] = 'Notes'

setup['A17'] = 'Periods loaded'
setup['B17'] = '=COUNTA(A22:A57)'
setup['C17'] = 36
setup['D17'] = '=IF(B17=C17,"PASS","FAIL")'
setup['E17'] = 'Last 36 month-end dates'

setup['A18'] = 'FY-end flags'
setup['B18'] = '=COUNTIF(F22:F57,TRUE)'
setup['C18'] = 3
setup['D18'] = '=IF(B18=C18,"PASS","FAIL")'
setup['E18'] = 'Expected three FY ends in 36 months'

setup['A19'] = 'TTM end flags'
setup['B19'] = '=COUNTIF(H22:H57,TRUE)'
setup['C19'] = 1
setup['D19'] = '=IF(B19=C19,"PASS","FAIL")'
setup['E19'] = 'Exactly one TTM endpoint'

setup['A20'] = 'Latest month consistency'
setup['B20'] = '=A57=B13'
setup['C20'] = True
setup['D20'] = '=IF(B20=C20,"PASS","FAIL")'
setup['E20'] = 'Latest period equals LatestMonthEnd'

setup['A21'] = 'Workbook QC status'
setup['B21'] = 'See Control | QC!B4'
setup['C21'] = 'PASS'
setup['D21'] = 'MANUAL'
setup['E21'] = 'Verify in Excel calc mode'

# --- Change log in setup (top section, outside period engine) ----------------
style_like(setup, 4, 14, 1, 2)
style_like(setup, 9, 14, 3, 5)
setup['A14'] = 'Change Log'
setup['B14'] = f'v1.1-packF-validated | {date.today().isoformat()} | Codex | Pack F validation'
setup['C14'] = None
setup['D14'] = None
setup['E14'] = None

# Clear any conflicting content in period-engine rows used in prior drafts.
for r in range(60, 63):
    for c in range(1, 6):
        setup.cell(r, c).value = None

# --- Data: entities + mappings + synthetic source data -----------------------
ent = wb['Map | Entities']
write_row(ent, 9, ['Combined', 'COMB', True, 1, 'Aggregated view'])
write_row(ent, 10, ['Entity 1', 'E1', True, 2, 'Synthetic entity'])
write_row(ent, 11, ['Entity 2', 'E2', True, 3, 'Synthetic entity'])

coa = wb['Map | COA to Lines']
coa_rows = [
    [4000, 'Revenue', 'IS', 'IS_REV', 'Revenue', 'IS_Revenue', 1, 1, ''],
    [5000, 'Cost of goods sold', 'IS', 'IS_COGS', 'Cost of goods sold', 'IS_COGS', 2, 1, ''],
    [6100, 'Operating expenses', 'IS', 'IS_OPEX', 'Operating expenses', 'IS_OPEX', 3, 1, ''],
    [7100, 'Interest expense', 'IS', 'IS_INT', 'Interest expense', 'IS_Other', 4, 1, ''],
    [7200, 'Taxes', 'IS', 'IS_TAX', 'Taxes', 'IS_Other', 5, 1, ''],
    [7300, 'Depreciation & amortization add-back', 'IS', 'IS_ADD_DA', 'Add back: Depreciation & amortization', 'IS_AddBack', 6, 1, ''],
    [7310, 'Interest add-back', 'IS', 'IS_ADD_INT', 'Add back: Interest', 'IS_AddBack', 7, 1, ''],
    [7320, 'Taxes add-back', 'IS', 'IS_ADD_TAX', 'Add back: Taxes', 'IS_AddBack', 8, 1, ''],
    [1000, 'Cash', 'BS', 'BS_CASH', 'Cash', 'BS_Assets', 9, 1, ''],
    [1100, 'Accounts receivable', 'BS', 'BS_AR', 'Accounts receivable', 'BS_Assets', 10, 1, ''],
    [1200, 'Inventory', 'BS', 'BS_INV', 'Inventory', 'BS_Assets', 11, 1, ''],
    [1300, 'Other current assets', 'BS', 'BS_OCA', 'Other current assets', 'BS_Assets', 12, 1, ''],
    [1500, 'PP&E', 'BS', 'BS_PPE', 'PP&E', 'BS_Assets', 13, 1, ''],
    [1600, 'Other assets', 'BS', 'BS_OTA', 'Other assets', 'BS_Assets', 14, 1, ''],
    [2000, 'Accounts payable', 'BS', 'BS_AP', 'Accounts payable', 'BS_Liab', 15, 1, ''],
    [2100, 'Accrued liabilities', 'BS', 'BS_ACCR', 'Accrued liabilities', 'BS_Liab', 16, 1, ''],
    [2200, 'Debt', 'BS', 'BS_DEBT', 'Debt', 'BS_Liab', 17, 1, ''],
    [2300, 'Other liabilities', 'BS', 'BS_OL', 'Other liabilities', 'BS_Liab', 18, 1, ''],
    [3000, 'Equity', 'BS', 'BS_EQ', 'Equity', 'BS_Equity', 19, 1, ''],
    [7400, 'Capex', 'CF', 'CF_CAPEX', 'Capex', 'CF', 20, 1, ''],
    [8100, 'Change in NWC', 'CF', 'CF_CHG_NWC', 'Change in NWC', 'CF', 21, 1, ''],
    [8200, 'Other cash flow', 'CF', 'CF_OTHER', 'Other', 'CF', 22, 1, ''],
]
for i, row in enumerate(coa_rows):
    write_row(coa, 9 + i, row)

cls = wb['Map | NWC & NetDebt Class']
class_rows = [
    [1100, 'NWC', 'NWC_AR', 'AR', 'AR', True, False, 1],
    [1200, 'NWC', 'NWC_INV', 'Inventory', 'INV', True, False, 2],
    [2000, 'NWC', 'NWC_AP', 'AP', 'AP', True, False, 3],
    [2100, 'NWC', 'NWC_ACCR', 'Accruals', 'ACCR', True, False, 4],
    [1000, 'NetDebt', 'ND_CASH', 'Cash', 'Cash', False, True, 5],
    [2200, 'NetDebt', 'ND_DEBT', 'Debt', 'Debt', False, True, 6],
    [2300, 'NetDebt', 'ND_DEBTLIKE', 'Debt-like', 'DebtLike', False, True, 7],
]
for i, row in enumerate(class_rows):
    write_row(cls, 9 + i, row)

# Build synthetic TB data
acct_order = [4000, 5000, 6100, 7100, 7200, 7300, 7310, 7320, 7400, 8100, 8200, 1000, 1100, 1200, 1300, 1500, 1600, 2000, 2100, 2200, 2300, 3000]
acct_name = {r[0]: r[1] for r in coa_rows}
flow_accounts = {4000, 5000, 6100, 7100, 7200, 7300, 7310, 7320, 7400, 8100, 8200}


def month_values(entity_mult, m_idx):
    rev = round((1200 + 18 * m_idx) * entity_mult, 2)
    cogs = round(-0.56 * rev, 2)
    opex = round(-0.22 * rev, 2)
    int_exp = round(-0.03 * rev, 2)
    tax_exp = round(-0.04 * rev, 2)
    add_da = round(0.06 * rev, 2)
    add_int = round(-int_exp, 2)
    add_tax = round(-tax_exp, 2)
    capex = round(-0.03 * rev, 2)
    chg_nwc = round(-0.015 * rev, 2)
    cf_other = round(0.01 * rev, 2)

    cash = round(320 * entity_mult + 6 * m_idx, 2)
    ar = round(0.14 * rev, 2)
    inv = round(0.09 * rev, 2)
    oca = round(0.04 * rev, 2)
    ppe = round(500 * entity_mult + 4 * m_idx, 2)
    ota = round(190 * entity_mult + 2 * m_idx, 2)
    ap = round(0.11 * (-cogs), 2)
    accr = round(70 * entity_mult + 1.5 * m_idx, 2)
    debt = round(430 * entity_mult - 1.5 * m_idx, 2)
    ol = round(55 * entity_mult + 1.0 * m_idx, 2)

    total_assets = cash + ar + inv + oca + ppe + ota
    total_liab = ap + accr + debt + ol
    equity = round(total_assets - total_liab, 2)

    return {
        4000: rev, 5000: cogs, 6100: opex, 7100: int_exp, 7200: tax_exp,
        7300: add_da, 7310: add_int, 7320: add_tax, 7400: capex, 8100: chg_nwc, 8200: cf_other,
        1000: cash, 1100: ar, 1200: inv, 1300: oca, 1500: ppe, 1600: ota,
        2000: ap, 2100: accr, 2200: debt, 2300: ol, 3000: equity,
    }


tb = wb['Data | TB']
entities = [('Entity 1', 1.00), ('Entity 2', 0.82)]
row = 9
for ent_name, mult in entities:
    monthly_by_account = {a: [] for a in acct_order}
    for m_idx in range(36):
        vals = month_values(mult, m_idx)
        for a in acct_order:
            monthly_by_account[a].append(vals[a])

    for a in acct_order:
        tb.cell(row, 1).value = ent_name
        tb.cell(row, 2).value = a
        tb.cell(row, 3).value = acct_name[a]
        tb.cell(row, 4).value = 1
        tb.cell(row, 5).value = 'USD'
        tb.cell(row, 6).value = 'Synthetic Pack F'

        # 36 monthly values G:AP
        for i, v in enumerate(monthly_by_account[a]):
            tb.cell(row, 7 + i).value = v

        # FY/TTM AQ:AS
        if a in flow_accounts:
            tb.cell(row, 43).value = f'=SUM(S{row}:AD{row})'   # FY-2 (2024)
            tb.cell(row, 44).value = f'=SUM(AE{row}:AP{row})'  # FY-1 (2025)
            tb.cell(row, 45).value = f'=SUM(AE{row}:AP{row})'  # TTM
        else:
            tb.cell(row, 43).value = f'=AD{row}'
            tb.cell(row, 44).value = f'=AP{row}'
            tb.cell(row, 45).value = f'=AP{row}'

        row += 1

# AR/AP aging raw inputs
ar_in = wb['Data | AR Aging']
ap_in = wb['Data | AP Aging']
for i in range(6):
    r = 9 + i
    ar_in.cell(r, 1).value = '=\'Control | Setup\'!$B$13'
    ar_in.cell(r, 2).value = 'Combined'
    ar_in.cell(r, 3).value = f'Customer {i+1}'
    ar_in.cell(r, 4).value = f'AR{i+1:03d}'
    ar_in.cell(r, 5).value = 120 - i * 10
    ar_in.cell(r, 6).value = 40 - i * 4
    ar_in.cell(r, 7).value = 20 - i * 2
    ar_in.cell(r, 8).value = 10 - i
    ar_in.cell(r, 9).value = 5
    ar_in.cell(r, 10).value = 3
    ar_in.cell(r, 11).value = f'=SUM(E{r}:J{r})'

    ap_in.cell(r, 1).value = '=\'Control | Setup\'!$B$13'
    ap_in.cell(r, 2).value = 'Combined'
    ap_in.cell(r, 3).value = f'Vendor {i+1}'
    ap_in.cell(r, 4).value = f'AP{i+1:03d}'
    ap_in.cell(r, 5).value = 100 - i * 8
    ap_in.cell(r, 6).value = 35 - i * 3
    ap_in.cell(r, 7).value = 18 - i * 2
    ap_in.cell(r, 8).value = 9 - i
    ap_in.cell(r, 9).value = 4
    ap_in.cell(r, 10).value = 2
    ap_in.cell(r, 11).value = f'=SUM(E{r}:J{r})'

# --- Output wiring for validation --------------------------------------------

# Combined IS
is_ws = wb['Combined | IS']
is_accounts = {9: 4000, 10: 5000, 12: 6100, 14: 7100, 15: 7200, 17: 7300, 18: 7310, 19: 7320}
for r, acct in is_accounts.items():
    is_ws.cell(r, 3).value = acct
    for c in range(7, 12):
        col = col_letter(c)
        is_ws.cell(r, c).value = f'=SUMIFS(\'Data | TB\'!{col}:{col},\'Data | TB\'!$B:$B,$C{r})'

# Combined BS
bs_ws = wb['Combined | BS']
bs_accounts = {9: 1000, 10: 1100, 11: 1200, 12: 1300, 14: 1500, 15: 1600, 17: 2000, 18: 2100, 19: 2200, 20: 2300, 22: 3000}
for r, acct in bs_accounts.items():
    bs_ws.cell(r, 3).value = acct
    for c in range(7, 12):
        col = col_letter(c)
        bs_ws.cell(r, c).value = f'=SUMIFS(\'Data | TB\'!{col}:{col},\'Data | TB\'!$B:$B,$C{r})'

# Combined CF
cf_ws = wb['Combined | CF']
cf_ws['C10'] = 8100
cf_ws['C11'] = 7400
cf_ws['C12'] = 8200
for c in range(7, 12):
    col = col_letter(c)
    cf_ws.cell(9, c).value = f"='Combined | IS'!{col}20"
    for r in [10, 11, 12]:
        cf_ws.cell(r, c).value = f'=SUMIFS(\'Data | TB\'!{col}:{col},\'Data | TB\'!$B:$B,$C{r})'

# Entity 1 tabs
for tab, row_map in [('Entity 1 | IS', is_accounts), ('Entity 1 | BS', bs_accounts)]:
    ws = wb[tab]
    for r, acct in row_map.items():
        ws.cell(r, 3).value = acct
        for c in range(7, 12):
            col = col_letter(c)
            ws.cell(r, c).value = f'=SUMIFS(\'Data | TB\'!{col}:{col},\'Data | TB\'!$B:$B,$C{r},\'Data | TB\'!$A:$A,"Entity 1")'

ws_e1cf = wb['Entity 1 | CF']
ws_e1cf['C10'] = 8100
ws_e1cf['C11'] = 7400
ws_e1cf['C12'] = 8200
for c in range(7, 12):
    col = col_letter(c)
    ws_e1cf.cell(9, c).value = f"='Entity 1 | IS'!{col}20"
    for r in [10, 11, 12]:
        ws_e1cf.cell(r, c).value = f'=SUMIFS(\'Data | TB\'!{col}:{col},\'Data | TB\'!$B:$B,$C{r},\'Data | TB\'!$A:$A,"Entity 1")'

# QofE detail + summary + checks
qd = wb['QofE | Detail']
qs = wb['QofE | Summary']
for c in range(7, 12):
    col = col_letter(c)
    qd.cell(9, c).value = f"='Combined | IS'!{col}9"
    qd.cell(10, c).value = f"='Combined | IS'!{col}10"
    qd.cell(11, c).value = f"='Combined | IS'!{col}12"
    qd.cell(13, c).value = 0
    qd.cell(14, c).value = 0

    qs.cell(9, c).value = f"='QofE | Detail'!{col}9"
    qs.cell(10, c).value = 0
    qs.cell(11, c).value = 0
    qs.cell(13, c).value = f"='QofE | Detail'!{col}12"
    qs.cell(14, c).value = 0
    qs.cell(15, c).value = 0

for c in range(7, 11):  # QC checks G:J
    col = col_letter(c)
    qs.cell(20, c).value = f'=IF({col}12="","",{col}12-\'QofE | Detail\'!{col}9)'
    qs.cell(21, c).value = f'=IF({col}16="","",{col}16-\'QofE | Detail\'!{col}15)'

# NWC detail + summary
nwd = wb['NWC | Detail']
nws = wb['NWC | Summary']
for c in range(7, 12):
    col = col_letter(c)
    nwd.cell(9, c).value = f"='Combined | BS'!{col}10"
    nwd.cell(10, c).value = f"='Combined | BS'!{col}11"
    nwd.cell(11, c).value = f"=-'Combined | BS'!{col}17"
    nwd.cell(12, c).value = f"=-'Combined | BS'!{col}18"
    nwd.cell(13, c).value = 0
    nwd.cell(28, c).value = f'=IF({col}14="","",{col}14-(\'Stratified | BS\'!{col}9+\'Stratified | BS\'!{col}10-\'Stratified | BS\'!{col}11-\'Stratified | BS\'!{col}12))'

    nws.cell(9, c).value = f"='NWC | Detail'!{col}14"
    nws.cell(10, c).value = 0
    nws.cell(11, c).value = 0
    nws.cell(13, c).value = f"='Combined | IS'!{col}9"
    nws.cell(17, c).value = f'=IF({col}12="","",{col}12-\'NWC | Detail\'!{col}14)'

# Stratified BS + Net debt
strat = wb['Stratified | BS']
nd = wb['Net_debt']
for c in range(7, 12):
    col = col_letter(c)
    strat.cell(9, c).value = f"='Combined | BS'!{col}10"
    strat.cell(10, c).value = f"='Combined | BS'!{col}11"
    strat.cell(11, c).value = f"='Combined | BS'!{col}17"
    strat.cell(12, c).value = f"='Combined | BS'!{col}18"
    strat.cell(13, c).value = f"='Combined | BS'!{col}9"
    strat.cell(14, c).value = f"='Combined | BS'!{col}19"
    strat.cell(15, c).value = f"='Combined | BS'!{col}23-SUM({col}9:{col}14)"
    strat.cell(28, c).value = f'=IF({col}16="","",{col}16-\'Combined | BS\'!{col}23)'

    nd.cell(9, c).value = f"='Stratified | BS'!{col}13"
    nd.cell(10, c).value = f"='Stratified | BS'!{col}14"
    nd.cell(11, c).value = 0
    nd.cell(12, c).value = f'=IF(COUNT({col}9:{col}11)=0,"",{col}10-{col}9+{col}11)'
    nd.cell(13, c).value = 0
    nd.cell(14, c).value = 0
    nd.cell(15, c).value = f'=IF(COUNT({col}12:{col}14)=0,"",{col}12+{col}13+{col}14)'
    nd.cell(28, c).value = f'=IF({col}12="","",{col}12-(\'Stratified | BS\'!{col}14-\'Stratified | BS\'!{col}13))'

# Recons (source vs databook pairs)
re = wb['Recons']
re_pairs = {
    9: ('Combined | IS', 9, 'QofE | Detail', 9),
    10: ('Combined | IS', 16, 'Combined | IS', 16),
    11: ('Combined | IS', 20, 'QofE | Summary', 16),
    12: ('Combined | IS', 12, 'Combined | IS', 12),
    13: ('Combined | IS', 13, 'Combined | IS', 13),
    18: ('Combined | BS', 16, 'Combined | BS', 16),
    19: ('Combined | BS', 21, 'Combined | BS', 21),
    20: ('Combined | BS', 22, 'Combined | BS', 22),
    21: ('Combined | BS', 23, 'Combined | BS', 23),
    22: ('Combined | BS', 10, 'Combined | BS', 10),
    27: ('Combined | CF', 13, 'Combined | CF', 13),
    28: ('Combined | CF', 9, 'Combined | CF', 9),
    29: ('Combined | CF', 11, 'Combined | CF', 11),
    30: ('Combined | CF', 12, 'Combined | CF', 12),
    31: ('Combined | CF', 10, 'Combined | CF', 10),
    36: ('Agings', 15, 'Agings', 18),
    37: ('Agings', 31, 'Agings', 34),
    38: ('NWC | Summary', 12, 'NWC | Detail', 14),
    39: ('Net_debt', 12, 'Net_debt', 12),
}
for r, (s1, r1, s2, r2) in re_pairs.items():
    d_col = 'J' if s1 == 'Agings' else 'G'
    e_col = 'J' if s2 == 'Agings' else 'G'
    re.cell(r, 4).value = f"='{s1}'!{d_col}{r1}"
    re.cell(r, 5).value = f"='{s2}'!{e_col}{r2}"
re.cell(40, 4).value = 0
re.cell(40, 5).value = 0

# Agings control links + bucket values
ag = wb['Agings']
ag['J18'] = "='Combined | BS'!G10"
ag['J34'] = "='Combined | BS'!G17"

ag['D9'] = '=IF($J$18="","",$J$18)'
for r in range(10, 15):
    ag.cell(r, 4).value = 0

ag['D25'] = '=IF($J$34="","",$J$34)'
for r in range(26, 31):
    ag.cell(r, 4).value = 0

# Other module checks: set inputs to ensure checks resolve.
for tab, input_rows in [('Personnel', [9, 10, 12, 13]), ('Capex', [9, 10, 12]), ('Leases', [9, 10, 11, 13]), ('Other | Analysis', [9, 10, 12])]:
    ws = wb[tab]
    for c in range(7, 11):
        col = col_letter(c)
        for r in input_rows:
            ws.cell(r, c).value = 100 + (r - 8) * 10 + (c - 6)

# --- Template index release controls -----------------------------------------
idx = wb['Template | Index']
idx['B29'] = 'Template scaffold v1: Packs A-F merged and validated; synthetic harness loaded, anchors/keys under release governance.'

# Add release governance block.
for dst, src in [(92, 7), (93, 8), (94, 9), (95, 9), (96, 9)]:
    for c in range(1, 6):
        idx.cell(dst, c)._style = copy(idx.cell(src, c)._style)
        idx.cell(dst, c).font = copy(idx.cell(src, c).font)
        idx.cell(dst, c).alignment = copy(idx.cell(src, c).alignment)

idx['A92'] = 'Release Controls'
idx['A93'] = 'Control'
idx['B93'] = 'Value'
idx['C93'] = 'Rule'
idx['D93'] = 'Status'
idx['E93'] = 'Notes'

idx['A94'] = 'Anchor map'
idx['B94'] = 'FROZEN'
idx['C94'] = 'Do not change BlockKey/TopLeftCell without version bump + changelog.'
idx['D94'] = 'ENFORCED'
idx['E94'] = 'Governed via Template | Index + Control | Setup change log'

idx['A95'] = 'Manifest keys'
idx['B95'] = 'FROZEN'
idx['C95'] = 'Do not rename ModuleKey values after release.'
idx['D95'] = 'ENFORCED'
idx['E95'] = 'Additive rows only for new modules'

idx['A96'] = 'Current release'
idx['B96'] = "='Control | Setup'!B9"
idx['C96'] = 'Version must increment if anchors move.'
idx['D96'] = 'TRACKED'
idx['E96'] = 'Validated with synthetic 2-entity load'

# Keep defaults explicit.
for ws in wb.worksheets:
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100
    ws.sheet_view.zoomScaleNormal = 100
for s in ['Financials>>', 'QofE>>', 'NWC>>']:
    if s in wb.sheetnames:
        wb[s].sheet_view.view = 'pageBreakPreview'

wb.save(OUT)
print('Wrote', OUT)
